#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从事件表 Excel 中回查典型案例候选事件。

输入:
  extract_case_study_candidates.py <incident.xlsx> <c2_ids_json> <virus_ids_json> <exploit_ids_json>

输出 JSON:
{
  "candidateCount": 3,
  "matchedCandidates": [
    {
      "incidentId": "alert-001",
      "severity": "严重",
      "dealStatus": "处置完成",
      "pushStatus": "已推送",
      "sourceType": "c2"
    }
  ]
}
"""

import json
import sys

from openpyxl import load_workbook

from _path_helper import decode_argv, reset_read_only_dimensions
decode_argv()


def normalize(value):
    return "" if value is None else str(value).strip()


def build_col_map(ws):
    header = [normalize(cell) for cell in next(ws.iter_rows(values_only=True))]
    return {name: i for i, name in enumerate(header) if name}


def find_column(col_map, aliases):
    for alias in aliases:
        if alias in col_map:
            return col_map[alias]
    return None


def resolve_source_type(incident_id, row, source_type_map, columns):
    mapped = source_type_map.get(incident_id, "")
    if mapped:
        return mapped

    gpt_result_col = columns.get("gpt_result_col")
    class_col = columns.get("class_col")
    gpt_result = normalize(row[gpt_result_col]) if gpt_result_col is not None and len(row) > gpt_result_col else ""
    class_value = normalize(row[class_col]) if class_col is not None and len(row) > class_col else ""

    if gpt_result == "主机失陷活动":
        return "c2"
    if gpt_result == "病毒木马活动":
        return "virus"
    if class_value == "漏洞利用":
        return "exploit"
    return ""


def collect_candidates(sheet, columns, source_type_map, candidate_ids=None):
    matched_candidates = []
    id_col = columns["id_col"]
    severity_col = columns["severity_col"]
    deal_status_col = columns["deal_status_col"]
    push_status_col = columns["push_status_col"]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        incident_id = normalize(row[id_col]) if len(row) > id_col else ""
        if not incident_id:
            continue
        if candidate_ids is not None and incident_id not in candidate_ids:
            continue

        deal_status = normalize(row[deal_status_col]) if len(row) > deal_status_col else ""
        push_status = normalize(row[push_status_col]) if len(row) > push_status_col else ""
        if deal_status != "处置完成" or push_status != "已推送":
            continue

        severity = normalize(row[severity_col]) if len(row) > severity_col else ""
        matched_candidates.append({
            "incidentId": incident_id,
            "severity": severity,
            "dealStatus": deal_status,
            "pushStatus": push_status,
            "sourceType": resolve_source_type(incident_id, row, source_type_map, columns)
        })

    return matched_candidates


def main():
    if len(sys.argv) < 5:
        raise SystemExit(
            "Usage: extract_case_study_candidates.py <incident.xlsx> <c2_ids_json> <virus_ids_json> <exploit_ids_json>"
        )

    incident_path = sys.argv[1]
    c2_ids = json.loads(sys.argv[2])
    virus_ids = json.loads(sys.argv[3])
    exploit_ids = json.loads(sys.argv[4])

    source_type_map = {}
    for source_type, ids in (
        ("c2", c2_ids),
        ("virus", virus_ids),
        ("exploit", exploit_ids),
    ):
        for incident_id in ids:
            key = normalize(incident_id)
            if key and key not in source_type_map:
                source_type_map[key] = source_type

    candidate_ids = set(source_type_map.keys())

    workbook = load_workbook(incident_path, read_only=True, data_only=True)
    sheet = reset_read_only_dimensions(workbook.active)
    col_map = build_col_map(sheet)

    id_col = find_column(col_map, ["事件ID", "事件Id", "incident_id", "incidentId", "id", "ID", "uuId"])
    severity_col = find_column(col_map, ["等级", "severity", "Severity"])
    deal_status_col = find_column(col_map, ["处置状态", "deal_status", "dealStatus"])
    push_status_col = find_column(col_map, ["推送状态", "event_push_label", "pushStatus"])
    gpt_result_col = find_column(col_map, ["GPT研判结论", "gpt_result", "gptResult"])
    class_col = find_column(col_map, ["安全事件一级分类", "事件一级分类", "class1", "categoryLevel1"])

    if id_col is None:
        raise SystemExit(f"事件表缺少事件ID列，可用列: {list(col_map.keys())}")
    if severity_col is None:
        raise SystemExit(f"事件表缺少等级列，可用列: {list(col_map.keys())}")
    if deal_status_col is None:
        raise SystemExit(f"事件表缺少处置状态列，可用列: {list(col_map.keys())}")
    if push_status_col is None:
        raise SystemExit(f"事件表缺少推送状态列，可用列: {list(col_map.keys())}")

    columns = {
        "id_col": id_col,
        "severity_col": severity_col,
        "deal_status_col": deal_status_col,
        "push_status_col": push_status_col,
        "gpt_result_col": gpt_result_col,
        "class_col": class_col
    }

    matched_candidates = collect_candidates(sheet, columns, source_type_map, candidate_ids)
    fallback_used = False
    if not matched_candidates:
        matched_candidates = collect_candidates(sheet, columns, source_type_map, None)
        fallback_used = True

    workbook.close()
    print(json.dumps({
        "candidateCount": len(candidate_ids),
        "matchedCandidates": matched_candidates,
        "fallbackUsed": fallback_used
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
