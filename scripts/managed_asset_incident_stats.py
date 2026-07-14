import json
import re
import sys
from datetime import datetime
from collections import Counter

from openpyxl import load_workbook

from _path_helper import decode_argv
decode_argv()

# 确保 stdout 使用 UTF-8 编码（Windows 兼容）
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

IP_PATTERN = re.compile(r'(?<!\d)(?:\d{1,3}\.){3}\d{1,3}(?!\d)')
SEVERITY_MAP = {
    "严重": "critical",
    "超危": "critical",
    "超危 改为  严重": "critical",
    "高危": "high",
    "中危": "medium",
    "低危": "low",
}


def normalize(value):
    return "" if value is None else str(value).strip()


def normalize_header(value):
    return normalize(value).lower().replace(" ", "").replace("_", "")


def find_column(headers, aliases):
    for index, header in enumerate(headers):
        normalized = normalize_header(header)
        if not normalized:
            continue
        for alias in aliases:
            if alias in normalized:
                return index
    return None


def extract_ips(text):
    return IP_PATTERN.findall(normalize(text))


def top1_name(counter):
    """取 Counter 第1项的名称, 无可返回空字符串"""
    items = counter.most_common(1)
    return items[0][0] if items else ""


def top_n(counter, n):
    """取 Counter 前n项, 返回 [{name, value}] 格式"""
    return [{"name": k, "value": v} for k, v in counter.most_common(n)]


def rank_business_systems(records, n):
    """
    按业务系统聚合事件并排序：
    1. 严重
    2. 高危
    3. 中危
    4. 低危
    5. total
    6. system（升序兜底，保证稳定）
    """
    system_counts = {}
    for record in records:
        system_name = record.get("system")
        if not system_name:
            system_name = "未知"
        bucket = system_counts.setdefault(system_name, {
            "critical": 0,
            "high": 0,
            "medium": 0,
            "low": 0,
            "total": 0,
        })
        severity = record.get("severity")
        if severity in ("critical", "high", "medium", "low"):
            bucket[severity] += 1
        bucket["total"] += 1

    ranking = []
    for system_name, counts in system_counts.items():
        ranking.append({
            "name": system_name,
            "value": counts["total"],
            "critical": counts["critical"],
            "high": counts["high"],
            "highRisk": counts["critical"] + counts["high"],
            "medium": counts["medium"],
            "low": counts["low"],
        })

    ranking.sort(key=lambda item: (
        -item["critical"],
        -item["high"],
        -item["medium"],
        -item["low"],
        -item["value"],
        item["name"],
    ))
    return ranking[:n]


def parse_datetime_value(value):
    """解析 Excel/字符串时间，返回 datetime"""
    if value is None:
        return None

    if isinstance(value, datetime):
        return value

    text = normalize(value)
    if not text:
        return None

    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue

    return None


def main():
    if len(sys.argv) < 3:
        raise SystemExit("Usage: managed_asset_incident_stats.py <asset.xlsx> <incident.xlsx>")

    asset_path = sys.argv[1]
    incident_path = sys.argv[2]

    # ====== 读取资产表 ======
    managed_asset_ips = set()
    ip_to_business = {}  # IP -> 所属业务

    try:
        asset_wb = load_workbook(asset_path, read_only=True, data_only=True)
        asset_ws = asset_wb.active
        header_row = next(asset_ws.iter_rows(min_row=2, max_row=2, values_only=True), ())
        headers = [normalize(h) for h in header_row]

        managed_col = find_column(headers, ["托管状态"])
        ip_col = find_column(headers, ["ip地址"])
        biz_col = find_column(headers, ["所属业务"])

        for row in asset_ws.iter_rows(min_row=3, values_only=True):
            if not any(normalize(cell) for cell in row):
                continue

            # 提取 IP
            ips = extract_ips(row[ip_col]) if (ip_col is not None and ip_col < len(row)) else []

            if not ips:
                continue

            ip = ips[0]

            # 收集业务系统映射
            if biz_col is not None and biz_col < len(row):
                biz = normalize(row[biz_col])
                if biz:
                    ip_to_business[ip] = biz

            # 收集托管资产 IP
            is_managed = False
            if managed_col is not None and managed_col < len(row):
                val = normalize(row[managed_col])
                if "已托管" in val or "托管" == val:
                    is_managed = True

            if is_managed:
                managed_asset_ips.add(ip)

        asset_wb.close()
    except Exception as e:
        print(json.dumps({
            "managedAssetEvents": 0,
            "managedAssetContainedEvents": 0,
            "managedAssetDisposedEvents": 0,
            "managedEventCloseRate": 0,
            "managedAssetCount": 0,
            "managedAvgResponseTime": 0,
            "topEventType": "",
            "top3BusinessSystems": "",
            "businessSystemEventDistribution": [],
            "error": f"读取资产表失败: {str(e)}"
        }, ensure_ascii=False))
        return

    # ====== 读取事件表 ======
    total_managed = 0
    contained_managed = 0
    disposed_managed = 0
    disposed_response_times = []
    event_type_counter = Counter()
    business_risk_records = []

    try:
        incident_wb = load_workbook(incident_path, data_only=True)
        incident_ws = incident_wb.active

        header_row = next(incident_ws.iter_rows(values_only=True))
        incident_headers = [normalize(h) for h in header_row]

        status_col = find_column(incident_headers, ["处置状态"])
        asset_col = find_column(incident_headers, ["影响资产"])
        event_type_col = find_column(incident_headers, ["安全事件一级分类"])
        completed_at_col = find_column(incident_headers, ["完成时间"])
        created_at_col = find_column(incident_headers, ["事件创建时间"])
        level_col = find_column(incident_headers, ["等级"])

        for row in incident_ws.iter_rows(min_row=2, values_only=True):
            if not any(normalize(cell) for cell in row):
                continue

            # 安全事件类型分布（所有事件）
            if event_type_col is not None and event_type_col < len(row):
                et = normalize(row[event_type_col])
                if et:
                    event_type_counter[et] += 1

            # 提取影响资产 IP
            incident_ip = None
            if asset_col is not None and asset_col < len(row):
                ips = extract_ips(row[asset_col])
                if ips:
                    incident_ip = ips[0]

            # 业务系统安全事件分布：资产表所属业务为空时归入"未知"，正常计入分布
            if incident_ip:
                biz = ip_to_business.get(incident_ip) or "未知"
                severity_raw = normalize(row[level_col] if level_col is not None and level_col < len(row) else None)
                business_risk_records.append({
                    "system": biz,
                    "severity": SEVERITY_MAP.get(severity_raw, ""),
                })

            # 托管资产事件统计（只统计影响资产为托管 IP 的事件）
            if not incident_ip or incident_ip not in managed_asset_ips:
                continue

            total_managed += 1

            if status_col is not None and status_col < len(row):
                status = normalize(row[status_col])
                if status == "已遏制":
                    contained_managed += 1
                elif status == "处置完成":
                    contained_managed += 1
                    disposed_managed += 1
                    # 收集处置完成事件的响应时间（完成时间 - 事件创建时间）
                    completed_at = row[completed_at_col] if completed_at_col is not None and completed_at_col < len(row) else None
                    created_at = row[created_at_col] if created_at_col is not None and created_at_col < len(row) else None
                    completed_dt = parse_datetime_value(completed_at)
                    created_dt = parse_datetime_value(created_at)
                    if completed_dt is not None and created_dt is not None:
                        diff_minutes = (completed_dt - created_dt).total_seconds() / 60
                        if diff_minutes >= 0:
                            disposed_response_times.append(diff_minutes)

        incident_wb.close()
    except Exception as e:
        print(json.dumps({
            "managedAssetEvents": 0,
            "managedAssetContainedEvents": 0,
            "managedAssetDisposedEvents": 0,
            "managedEventCloseRate": 0,
            "managedAssetCount": len(managed_asset_ips),
            "managedAvgResponseTime": 0,
            "topEventType": "",
            "top3BusinessSystems": "",
            "businessSystemEventDistribution": [],
            "error": f"读取事件表失败: {str(e)}"
        }, ensure_ascii=False))
        return

    close_rate = round((disposed_managed / total_managed) * 100, 2) if total_managed else 0
    avg_response_time = round(sum(disposed_response_times) / len(disposed_response_times), 1) if disposed_response_times else 0
    business_ranking = rank_business_systems(business_risk_records, 5)

    result = {
        "managedAssetEvents": total_managed,
        "managedAssetContainedEvents": contained_managed,
        "managedAssetDisposedEvents": disposed_managed,
        "managedEventCloseRate": close_rate,
        "managedAssetCount": len(managed_asset_ips),
        "managedAvgResponseTime": avg_response_time,
        "topEventType": top1_name(event_type_counter),
        "top3BusinessSystems": "、".join(item["name"] for item in business_ranking[:3]),
        "businessSystemEventDistribution": [
            {
                "name": item["name"],
                "value": item["value"],
                "highRisk": item["highRisk"],
            }
            for item in business_ranking
        ]
    }

    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
