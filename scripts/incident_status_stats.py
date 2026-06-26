import json
import re
import sys

from openpyxl import load_workbook

IP_PATTERN = re.compile(r'(?<!\d)(?:\d{1,3}\.){3}\d{1,3}(?!\d)')


def normalize(value):
    return "" if value is None else str(value).strip()


def parse_number(value):
    text = normalize(value)
    if not text:
        return None

    try:
        return float(text)
    except ValueError:
        return None


def extract_ips(value):
    text = normalize(value)
    if not text:
        return []

    return IP_PATTERN.findall(text)


def _build_col_map(ws):
    """读取表头行，返回 列名 → 列索引(0-based) 的映射"""
    header = [normalize(cell) for cell in next(ws.iter_rows(values_only=True))]
    return {name: i for i, name in enumerate(header) if name}


def main():
    if len(sys.argv) < 2:
        raise SystemExit("Usage: incident_status_stats.py <incident.xlsx>")

    workbook = load_workbook(sys.argv[1], read_only=True, data_only=True)
    sheet = workbook.active

    col_map = _build_col_map(sheet)
    level_col = col_map.get("等级")
    status_col = col_map.get("处置状态")
    ip_col = col_map.get("影响资产")

    total = 0
    severe = 0
    high = 0
    closed = 0
    contained = 0
    processing = 0
    unique_ips = set()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(normalize(cell) for cell in row):
            continue

        total += 1
        if level_col is not None and len(row) > level_col:
            level = normalize(row[level_col])
            if level == "严重":
                severe += 1
            elif level == "高危":
                high += 1

        if status_col is not None and len(row) > status_col:
            status = normalize(row[status_col])
            if status == "处置完成":
                closed += 1
            elif status == "已遏制":
                contained += 1
            elif status == "处置中":
                processing += 1

        if ip_col is not None and len(row) > ip_col:
            for ip in extract_ips(row[ip_col]):
                unique_ips.add(ip)

    close_rate = round((closed / total) * 100) if total else 0
    print(json.dumps({
        "totalEvents": total,
        "severeEvents": severe,
        "highEvents": high,
        "closedEvents": closed,
        "containedEvents": contained,
        "processingEvents": processing,
        "closeRate": close_rate,
        "uniqueAssetCount": len(unique_ips)
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
