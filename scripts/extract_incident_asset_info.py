#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从事件表 Excel 和资产表 Excel 中提取三类资产信息：
1. 病毒攻击资产    - 第一个病毒有效事件的影响资产(IP)
2. 未被AES覆盖资产 - 在资产表中"数据源"列不含EDR的资产IP，最多2个
3. 未标注资产      - 在资产表中"责任人"列为空的资产IP，最多2个

用法:
  extract_incident_asset_info.py <incident.xlsx> <asset.xlsx> <confirmed_ids_json> <virus_ids_json>

输出 JSON:
{
  "virusAttackAsset": "10.5.40.62",
  "nonAesCoveredAssets": ["10.5.40.63", "10.5.40.64"],
  "unlabeledAssets": ["10.5.40.65", "10.5.40.66"]
}
"""

import json
import re
import sys

from openpyxl import load_workbook


def normalize(value):
    return "" if value is None else str(value).strip()


def extract_ip(raw):
    """从文本中提取第一个 IPv4 地址"""
    if not raw:
        return None
    m = re.search(r'(\d+\.\d+\.\d+\.\d+)', str(raw))
    return m.group(1) if m else None


def build_col_map(ws):
    """读取表头行，返回 列名 → 列索引(0-based) 的映射"""
    header = [normalize(cell) for cell in next(ws.iter_rows(values_only=True))]
    return {name: i for i, name in enumerate(header) if name}


def find_column(col_map, aliases):
    """在列名映射中查找第一个匹配的列索引"""
    for alias in aliases:
        if alias in col_map:
            return col_map[alias]
    return None


def main():
    if len(sys.argv) < 5:
        raise SystemExit(
            "Usage: extract_incident_asset_info.py <incident.xlsx> <asset.xlsx> "
            "<confirmed_ids_json> <virus_ids_json>"
        )

    incident_path = sys.argv[1]
    asset_path = sys.argv[2]
    confirmed_ids = json.loads(sys.argv[3])
    virus_ids = json.loads(sys.argv[4])

    # ====== Phase 1: 读取事件表 Excel ======
    # 收集：第一个病毒事件的IP、所有已确认事件的有序IP列表(去重)
    incident_wb = load_workbook(incident_path, read_only=True, data_only=True)
    incident_ws = incident_wb.active
    inc_col_map = build_col_map(incident_ws)

    id_col = find_column(inc_col_map, ["事件ID", "incident_id"])
    asset_col = find_column(inc_col_map, ["影响资产", "host_ip", "hostIp", "主机IP", "ip"])

    if id_col is None:
        print(json.dumps({
            "error": f"事件表缺少事件ID列，可用列: {list(inc_col_map.keys())}"
        }, ensure_ascii=False))
        sys.exit(1)
    if asset_col is None:
        print(json.dumps({
            "error": f"事件表缺少影响资产列，可用列: {list(inc_col_map.keys())}"
        }, ensure_ascii=False))
        sys.exit(1)

    virus_set = set(virus_ids)
    confirmed_set = set(confirmed_ids)

    first_virus_asset = None       # 第一个病毒事件的影响资产(IP)
    confirmed_ips_ordered = []     # 所有已确认事件的影响资产IP，保持事件表遍历顺序，去重
    seen_ips = set()

    for row in incident_ws.iter_rows(min_row=2, values_only=True):
        incident_id = normalize(row[id_col]) if len(row) > id_col else ""
        if not incident_id or incident_id not in confirmed_set:
            continue

        raw_asset = normalize(row[asset_col]) if len(row) > asset_col else ""
        ip = extract_ip(raw_asset) if raw_asset else None
        if not ip:
            continue

        # 第一个病毒事件
        if not first_virus_asset and incident_id in virus_set:
            first_virus_asset = ip

        # 有序去重记录IP
        if ip not in seen_ips:
            seen_ips.add(ip)
            confirmed_ips_ordered.append(ip)

    incident_wb.close()

    # ====== Phase 2: 读取资产表 Excel ======
    # 建映射: IP → {数据源, 责任人}
    ip_to_data_source = {}
    ip_to_responsible = {}

    if asset_path and confirmed_ips_ordered:
        try:
            asset_wb = load_workbook(asset_path, read_only=True, data_only=True)
            asset_ws = asset_wb.active
            asset_col_map = build_col_map(asset_ws)

            ip_col = find_column(asset_col_map, [
                "IP地址", "P地址", "IP", "地址", "主机IP",
                "hostIp", "host_ip", "ip", "资产IP"
            ])
            ds_col = find_column(asset_col_map, [
                "数据源", "数据来源", "dataSource", "data_source",
                "设备来源", "devSourceNames", "source"
            ])
            resp_col = find_column(asset_col_map, ["责任人", "负责人", "responsible", "负责人姓名"])

            if ip_col is not None:
                for row in asset_ws.iter_rows(min_row=2, values_only=True):
                    cell_ip = normalize(row[ip_col]) if len(row) > ip_col else ""
                    if not cell_ip:
                        continue

                    # 只缓存我们关心的IP
                    if cell_ip in seen_ips:
                        ds = normalize(row[ds_col]) if ds_col is not None and len(row) > ds_col else ""
                        if ds_col is not None:
                            ip_to_data_source[cell_ip] = ds

                        resp = normalize(row[resp_col]) if resp_col is not None and len(row) > resp_col else ""
                        if resp_col is not None:
                            ip_to_responsible[cell_ip] = resp

            asset_wb.close()
        except Exception as e:
            # 资产表读取失败时不阻断，只记个日志（stderr）
            print(f"资产表读取失败: {e}", file=sys.stderr)

    # ====== Phase 3: 计算结果 ======

    # 未被AES覆盖资产：在资产表中数据源列不含EDR的IP，最多2个
    non_aes_assets = []
    for ip in confirmed_ips_ordered:
        ds = ip_to_data_source.get(ip, "")
        # 如果在资产表中找不到该IP，视为未被覆盖
        if not ds or "EDR" not in ds.upper():
            non_aes_assets.append(ip)
            if len(non_aes_assets) >= 2:
                break

    # 未标注资产：在资产表中责任人列为空的IP，最多2个
    unlabeled_assets = []
    for ip in confirmed_ips_ordered:
        # 如果在资产表中找不到该IP，或责任人为空
        resp = ip_to_responsible.get(ip, "")
        if not resp:
            unlabeled_assets.append(ip)
            if len(unlabeled_assets) >= 2:
                break

    # ====== 输出结果 ======
    result = {
        "virusAttackAsset": first_virus_asset or "",
        "nonAesCoveredAssets": non_aes_assets,
        "unlabeledAssets": unlabeled_assets
    }
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
