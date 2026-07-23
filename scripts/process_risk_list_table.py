#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
资产表和事件表后处理脚本。

功能：
- asset: 删除 zdy、责任人电话、责任人(设备上报)、实时认证用户名 四列；保留原文件名复制到输出目录
- incident: 原样复制到输出目录，不修改列或数据

用法：
    python process_risk_list_table.py asset <input.xlsx> <output_dir>
    python process_risk_list_table.py incident <input.xlsx> <output_dir>

输出：
    JSON: {"filePath": "<output_dir>/<原文件名>.xlsx" }
"""
import json
import os
import shutil
import sys

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font as _Font

from _path_helper import decode_argv
decode_argv()


ASSET_COLUMNS_TO_REMOVE = [
    'zdy',
    '责任人电话',
    '责任人(设备上报)',
    '实时认证用户名',
]

BUSINESS_COLUMN_ALIASES = ['所属业务']


def normalize(value):
    return '' if value is None else str(value).strip()


def clean_business_column(value):
    """处理所属业务列：按逗号分隔，每个分段去掉 '/' 及其前面的部分，保留后面的内容。"""
    text = normalize(value)
    if not text:
        return text
    parts = text.split(',')
    cleaned = []
    for part in parts:
        part = part.strip()
        if '/' in part:
            part = part.rsplit('/', 1)[-1].strip()
        if part:
            cleaned.append(part)
    return ', '.join(cleaned) if cleaned else ''


def build_col_header(ws):
    """查找第一行非空行作为表头，返回 (col_name->index 字典, header_row_number)。"""
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        values = [normalize(cell) for cell in row]
        if any(values):
            return {name: i for i, name in enumerate(values) if name}, idx
    return {}, 1


def process_asset_table(input_path, output_dir, wait_approve_path=None):
    """处理资产表：删除指定列，清理所属业务列，可选合并待审核数据并新增"审核状态"列。

    - input_path: 资产台账（search_type=current）导出的 xlsx
    - wait_approve_path: 待审核资产（search_type=wait_approve）导出的 xlsx，如提供则合并并加"审核状态"列
    - 输出文件保留原资产台账的文件名
    """
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, os.path.basename(input_path))

    wb = load_workbook(input_path)
    ws = wb.active

    col_map, header_row = build_col_header(ws)

    # 1) 删除指定列
    to_remove_cols = []
    for col_name in ASSET_COLUMNS_TO_REMOVE:
        if col_name in col_map:
            to_remove_cols.append(col_map[col_name])

    if to_remove_cols:
        to_remove_cols.sort(reverse=True)
        for col_idx in to_remove_cols:
            ws.delete_cols(col_idx + 1, 1)

    # 删除列后重新读取表头（列索引已变）
    col_map, _ = build_col_header(ws)

    # 2) 清理所属业务列
    business_col_idx = None
    for alias in BUSINESS_COLUMN_ALIASES:
        if alias in col_map:
            business_col_idx = col_map[alias]
            break

    if business_col_idx is not None:
        for row in ws.iter_rows(min_row=header_row + 1):
            cell = row[business_col_idx]
            if cell.value is not None:
                raw = normalize(cell.value)
                cleaned = clean_business_column(raw)
                if cleaned != raw:
                    cell.value = cleaned

    # 3) 合并待审核资产 + 新增"审核状态"列
    if wait_approve_path and os.path.isfile(wait_approve_path):
        # 在现有表头追加"审核状态"列
        max_col = ws.max_column or 0
        approval_col_idx = max_col + 1
        approval_header_cell = ws.cell(row=header_row, column=approval_col_idx)
        approval_header_cell.value = '审核状态'

        # 设置"审核状态"表头底色为 #333333，字体颜色为白色，与 MSSW 导出表头风格一致
        try:
            approval_header_cell.fill = PatternFill(start_color='FF333333', end_color='FF333333', fill_type='solid')
            # 字体复制第一列表头的字体样式，但颜色改为白色
            first_header_cell = ws.cell(row=header_row, column=1)
            if first_header_cell.has_style:
                base_font = first_header_cell.font
                approval_header_cell.font = _Font(
                    name=base_font.name or '微软雅黑',
                    size=base_font.size,
                    bold=base_font.bold,
                    italic=base_font.italic,
                    underline=base_font.underline,
                    strike=base_font.strike,
                    vertAlign=base_font.vertAlign,
                    color='FFFFFFFF'
                )
                approval_header_cell.border = first_header_cell.border.copy()
                approval_header_cell.alignment = first_header_cell.alignment.copy()
                approval_header_cell.number_format = first_header_cell.number_format
                approval_header_cell.protection = first_header_cell.protection.copy()
        except Exception as e:
            sys.stderr.write(f'[WARN] 设置表头样式失败: {e}\n')

        # 给资产台账的所有数据行填"已审核"
        for row in ws.iter_rows(min_row=header_row + 1):
            cell = row[approval_col_idx - 1]  # row 是 0-indexed，cell column 用 1-indexed
            # row 列数可能不足，直接用 ws.cell 写入
            ws.cell(row=row[0].row, column=approval_col_idx, value='已审核')

        # 读待审核表，追加到资产台账末尾，填"待审核"
        try:
            wb_wait = load_workbook(wait_approve_path)
            ws_wait = wb_wait.active
            wait_col_map, wait_header_row = build_col_header(ws_wait)

            # 删除待审核表中与资产台账相同名称的多余列，使其列结构一致
            wait_remove_cols = []
            for col_name in ASSET_COLUMNS_TO_REMOVE:
                if col_name in wait_col_map:
                    wait_remove_cols.append(wait_col_map[col_name])
            if wait_remove_cols:
                wait_remove_cols.sort(reverse=True)
                for col_idx in wait_remove_cols:
                    ws_wait.delete_cols(col_idx + 1, 1)

            # 重新计算待审核表的列映射
            wait_col_map, _ = build_col_header(ws_wait)
            wait_business_col_idx = None
            for alias in BUSINESS_COLUMN_ALIASES:
                if alias in wait_col_map:
                    wait_business_col_idx = wait_col_map[alias]
                    break

            current_max_row = ws.max_row or header_row
            wait_max_col = ws_wait.max_column or 0
            # 确保列数一致：以资产台账的列数为准（追加的审核状态列除外）
            target_col_count = approval_col_idx - 1

            for row_idx, row in enumerate(ws_wait.iter_rows(min_row=wait_header_row + 1, values_only=True), start=current_max_row + 1):
                if not any(normalize(c) for c in row):
                    continue
                # 写每个单元格
                for c_idx, cell_value in enumerate(row, start=1):
                    if c_idx > target_col_count:
                        break
                    ws.cell(row=row_idx, column=c_idx, value=cell_value)
                # 清理所属业务列
                if wait_business_col_idx is not None and wait_business_col_idx < target_col_count:
                    cell = ws.cell(row=row_idx, column=wait_business_col_idx + 1)
                    if cell.value is not None:
                        raw = normalize(cell.value)
                        cleaned = clean_business_column(raw)
                        if cleaned != raw:
                            cell.value = cleaned
                # 填"待审核"
                ws.cell(row=row_idx, column=approval_col_idx, value='待审核')

            wb_wait.close()
        except Exception as e:
            # 合并失败不阻断主流程，但记录到 stderr
            sys.stderr.write(f'[WARN] 合并待审核资产失败: {e}\n')

    wb.save(output_path)
    wb.close()
    return output_path


def process_incident_table(input_path, output_dir):
    """原样复制事件表，不补列、不填充模拟数据。"""
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, os.path.basename(input_path))
    shutil.copy2(input_path, output_path)
    return output_path


def main():
    if len(sys.argv) < 4:
        raise SystemExit(
            'Usage: process_risk_list_table.py <asset|incident> <input.xlsx> <output_dir> [wait_approve.xlsx]'
        )

    table_type = sys.argv[1].lower()
    input_path = sys.argv[2]
    output_dir = sys.argv[3]
    wait_approve_path = sys.argv[4] if len(sys.argv) >= 5 else None

    if not os.path.isfile(input_path):
        raise SystemExit(f'输入文件不存在: {input_path}')

    if table_type == 'asset':
        output_path = process_asset_table(input_path, output_dir, wait_approve_path)
    elif table_type == 'incident':
        output_path = process_incident_table(input_path, output_dir)
    else:
        raise SystemExit(f'不支持的表类型: {table_type}（仅支持 asset / incident）')

    print(json.dumps({'filePath': output_path}, ensure_ascii=False))


if __name__ == '__main__':
    main()
