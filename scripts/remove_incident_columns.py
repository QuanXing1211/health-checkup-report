#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从事件表 Excel 中删除"外网IP地址"、"域名"、"文件"三列。
用法: python remove_incident_columns.py <incident.xlsx> <output_dir>
输出: JSON: {"filePath": "<output_dir>/<原文件名>.xlsx"}
"""
import json
import os
import shutil
import sys

from openpyxl import load_workbook

from _path_helper import decode_argv
decode_argv()


COLUMNS_TO_REMOVE = [
    '外网IP地址',
    '域名',
    '文件',
]


def normalize(value):
    return '' if value is None else str(value).strip()


def build_col_header(ws):
    """查找第一行非空行作为表头，返回 (col_name->index 字典, header_row_number)。"""
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        values = [normalize(cell) for cell in row]
        if any(values):
            return {name: i for i, name in enumerate(values) if name}, idx
    return {}, 1


def main():
    if len(sys.argv) < 3:
        raise SystemExit(
            'Usage: remove_incident_columns.py <incident.xlsx> <output_dir>'
        )

    input_path = sys.argv[1]
    output_dir = sys.argv[2]

    if not os.path.isfile(input_path):
        raise SystemExit(f'输入文件不存在: {input_path}')

    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, os.path.basename(input_path))

    wb = load_workbook(input_path)
    ws = wb.active

    col_map, header_row = build_col_header(ws)
    to_remove_cols = []
    for col_name in COLUMNS_TO_REMOVE:
        if col_name in col_map:
            to_remove_cols.append(col_map[col_name])

    if not to_remove_cols:
        # 没有需要删除的列，直接复制
        wb.close()
        shutil.copy2(input_path, output_path)
        print(json.dumps({'filePath': output_path}, ensure_ascii=False))
        return

    # 按列索引从大到小排序，从右往左删除（避免删除后索引偏移）
    to_remove_cols.sort(reverse=True)
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    ws.delete_rows(1, ws.max_row)
    for row in all_rows:
        row_list = list(row)
        for col_idx in to_remove_cols:
            if col_idx < len(row_list):
                del row_list[col_idx]
        ws.append(row_list)

    wb.save(output_path)
    wb.close()
    print(json.dumps({'filePath': output_path}, ensure_ascii=False))


if __name__ == '__main__':
    main()
