#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
读取事件表 Excel 的 GPT研判结论 和 GPT定性结论 列，输出 JSON 供 Node.js 使用。

输出格式:
{
  "hostCompromiseIds": ["incident-xxx", ...],  // GPT研判结论 = "主机失陷活动"
  "virusTrojanIds": ["incident-yyy", ...],      // GPT研判结论 = "病毒木马活动"
  "gptSubResultMap": {
    "incident-xxx": "勒索攻击",
    "incident-yyy": "银狐病毒"
  }
}
"""

import json
import sys

from openpyxl import load_workbook

from _path_helper import decode_argv
decode_argv()


HEADER_INCIDENT_ID = "事件ID"
HEADER_GPT_RESULT = "GPT研判结论"
HEADER_GPT_SUB_RESULT = "GPT定性结论"

GPT_RESULT_HOST_COMPROMISE = "主机失陷活动"
GPT_RESULT_VIRUS_TROJAN = "病毒木马活动"


def normalize(value):
    return "" if value is None else str(value).strip()


def build_col_map(ws):
    """读取表头行，返回 列名 → 列索引(0-based) 的映射"""
    header = [normalize(cell) for cell in next(ws.iter_rows(values_only=True))]
    return {name: i for i, name in enumerate(header) if name}


def main():
    if len(sys.argv) < 2:
        raise SystemExit("Usage: incident_gpt_stats.py <incident.xlsx>")

    workbook = load_workbook(sys.argv[1], data_only=True)
    sheet = workbook.active

    col_map = build_col_map(sheet)
    id_col = col_map.get(HEADER_INCIDENT_ID)
    gpt_col = col_map.get(HEADER_GPT_RESULT)
    sub_col = col_map.get(HEADER_GPT_SUB_RESULT)

    if id_col is None:
        raise SystemExit(f"Excel 缺少列: {HEADER_INCIDENT_ID}")
    if gpt_col is None:
        raise SystemExit(f"Excel 缺少列: {HEADER_GPT_RESULT}")

    host_compromise_ids = []
    virus_trojan_ids = []
    gpt_sub_result_map = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        incident_id = normalize(row[id_col]) if len(row) > id_col else ""
        if not incident_id:
            continue

        gpt_value = normalize(row[gpt_col]) if len(row) > gpt_col else ""

        if gpt_value == GPT_RESULT_HOST_COMPROMISE:
            host_compromise_ids.append(incident_id)
        elif gpt_value == GPT_RESULT_VIRUS_TROJAN:
            virus_trojan_ids.append(incident_id)

        if sub_col is not None and len(row) > sub_col:
            sub_value = normalize(row[sub_col])
            if sub_value and sub_value != "-":
                gpt_sub_result_map[incident_id] = sub_value

    print(json.dumps({
        "hostCompromiseIds": host_compromise_ids,
        "virusTrojanIds": virus_trojan_ids,
        "gptSubResultMap": gpt_sub_result_map
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
