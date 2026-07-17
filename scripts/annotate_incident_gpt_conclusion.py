#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""按统一规则给事件表的 GPT研判结论追加 C2/病毒木马分类后缀。"""

import json
import os
import sys

from openpyxl import load_workbook

from _incident_classification import (
    build_col_map,
    classify_event,
    format_gpt_conclusion,
    get_classification_columns,
    normalize,
)
from _path_helper import decode_argv, reset_read_only_dimensions
decode_argv()


def main():
    if len(sys.argv) < 3:
        raise SystemExit(
            "Usage: annotate_incident_gpt_conclusion.py <incident.xlsx> <output_dir>"
        )

    input_path = sys.argv[1]
    output_dir = sys.argv[2]
    if not os.path.isfile(input_path):
        raise SystemExit(f"输入文件不存在: {input_path}")

    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, os.path.basename(input_path))

    workbook = load_workbook(input_path)
    sheet = reset_read_only_dimensions(workbook.active)
    col_map = build_col_map(sheet)
    columns = get_classification_columns(col_map)
    gpt_col = columns["gpt"]
    if gpt_col is None:
        raise SystemExit("事件表缺少 GPT研判结论 列")

    classified = {"C2外联": 0, "病毒木马": 0}
    for row in sheet.iter_rows(min_row=2):
        values = tuple(cell.value for cell in row)
        raw_value = normalize(values[gpt_col]) if len(values) > gpt_col else ""
        classification = classify_event(raw_value, values, columns)
        formatted = format_gpt_conclusion(raw_value, classification)
        if len(row) > gpt_col and row[gpt_col].value != formatted:
            row[gpt_col].value = formatted
        if classification:
            classified[classification] += 1

    workbook.save(output_path)
    workbook.close()
    print(json.dumps({
        "filePath": output_path,
        "classified": classified,
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
