#!/usr/bin/env python3
"""
从事件 Excel 中移除指定 incident_id 的行。
用法: python remove_incident_rows.py <incident.xlsx> '<incident_ids_json>'
"""
import json
import sys

from openpyxl import load_workbook

from _path_helper import decode_argv
decode_argv()


def normalize(value):
    return "" if value is None else str(value).strip()


def find_column(sheet, aliases):
    """在表头中查找列名，支持多个别名"""
    header = [normalize(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    for index, name in enumerate(header):
        if not name:
            continue
        name_normalized = name.lower().replace(" ", "").replace("_", "").replace("-", "")
        for alias in aliases:
            if alias.lower().replace(" ", "").replace("_", "").replace("-", "") in name_normalized:
                return index
    return None


def main():
    if len(sys.argv) < 3:
        raise SystemExit("Usage: remove_incident_rows.py <incident.xlsx> '<incident_ids_json>'")

    excel_path = sys.argv[1]
    try:
        incident_ids = json.loads(sys.argv[2])
    except json.JSONDecodeError as e:
        raise SystemExit(f"无法解析 incident_ids JSON: {e}")

    if not incident_ids:
        print(json.dumps({"removed": 0, "total_before": 0, "total_after": 0, "message": "没有误报事件需要移除"}))
        return

    incident_id_set = set(str(item).strip() for item in incident_ids if item)

    workbook = load_workbook(excel_path)
    sheet = workbook.active

    # 查找事件ID列
    col_index = find_column(sheet, ["事件id", "事件编号", "incident_id", "uuid", "uu id"])
    if col_index is None:
        # 如果找不到事件ID列，尝试遍历所有列找包含 incident 的列
        header = [normalize(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        for i, name in enumerate(header):
            normalized = name.lower().replace(" ", "").replace("_", "").replace("-", "")
            if "incident" in normalized or "event" in normalized:
                col_index = i
                break

    if col_index is None:
        raise SystemExit(f"无法找到事件ID列 (表头: {[normalize(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]})")

    # 收集要保留的行（表头 + 不匹配的行）
    rows_to_keep = []
    removed_count = 0
    total_before = 0

    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=True)):
        if row_idx == 0:
            # 表头行，保留
            rows_to_keep.append(row)
            continue

        if not any(normalize(cell) for cell in row):
            continue

        total_before += 1

        cell_value = normalize(row[col_index]) if col_index < len(row) else ""
        if cell_value in incident_id_set:
            removed_count += 1
        else:
            rows_to_keep.append(row)

    # 清空原工作表并写入保留的行
    sheet.delete_rows(1, sheet.max_row)
    for row_data in rows_to_keep:
        sheet.append(row_data)

    workbook.save(excel_path)

    print(json.dumps({
        "removed": removed_count,
        "total_before": total_before,
        "total_after": total_before - removed_count,
        "message": f"已从事件表中移除 {removed_count} 条误报事件"
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
