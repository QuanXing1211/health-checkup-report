#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从事件表 Excel 中直接提取三类事件统计：

1. 候选池：GPT研判结论属于 "主机失陷活动" / "病毒木马活动"
2. C2 外联：候选事件的 外网IP地址 / 域名列中存在标记为"黑"的实体
3. 病毒木马：候选事件网络字段没有"黑"实体，但 文件列中存在标记为"黑"的实体
4. 网络字段优先于文件字段；三类字段都没有"黑"实体的候选事件跳过
5. 漏洞利用：安全事件一级分类列值为"漏洞利用"

用法:
  extract_incident_direct_stats.py <incident.xlsx>

输出 JSON:
{
  "hostCompromiseIds": ["incident-001"],
  "virusTrojanIds": ["incident-002"],
  "exploitIds": ["incident-003"]
}
"""

import json
import sys

from openpyxl import load_workbook

from _path_helper import decode_argv, reset_read_only_dimensions
from _incident_classification import (
    build_col_map,
    classify_event,
    get_classification_columns,
    normalize,
)
decode_argv()


def main():
    if len(sys.argv) < 2:
        raise SystemExit("Usage: extract_incident_direct_stats.py <incident.xlsx>")

    workbook = load_workbook(sys.argv[1], read_only=True, data_only=True)
    sheet = reset_read_only_dimensions(workbook.active)
    col_map = build_col_map(sheet)
    columns = get_classification_columns(col_map)
    id_col = columns["id"]
    gpt_col = columns["gpt"]
    class_col = next((col_map.get(alias) for alias in [
        "安全事件一级分类", "事件分类", "incident_threat_class",
        "incidentThreatClass", "威胁分类", "threatClass"
    ] if alias in col_map), None)

    if id_col is None:
        raise SystemExit(f"事件表缺少事件ID列，可用列: {list(col_map.keys())}")

    host_compromise_ids = []
    virus_trojan_ids = []
    exploit_ids = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        incident_id = normalize(row[id_col]) if len(row) > id_col else ""
        if not incident_id:
            continue

        gpt_value = normalize(row[gpt_col]) if gpt_col is not None and len(row) > gpt_col else ""
        classification = classify_event(gpt_value, row, columns)
        if classification == "C2外联":
            host_compromise_ids.append(incident_id)
        elif classification == "病毒木马":
            virus_trojan_ids.append(incident_id)

        class_value = normalize(row[class_col]) if class_col is not None and len(row) > class_col else ""
        if class_value == "漏洞利用":
            exploit_ids.append(incident_id)

    workbook.close()
    print(json.dumps({
        "hostCompromiseIds": host_compromise_ids,
        "virusTrojanIds": virus_trojan_ids,
        "exploitIds": exploit_ids
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
