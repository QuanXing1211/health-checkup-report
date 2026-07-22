# -*- coding: utf-8 -*-
"""
核心业务系统风险排行工具

规则：
1. 核心业务系统排序仅使用：漏洞表 + 事件表 + 弱口令表
2. 弱口令统一按“中危”处理
3. 业务归因使用资产表中所有具备“所属业务”的资产，映射字段为：IP地址 -> 所属业务
4. 暴露面不参与核心业务系统排序，仅参与总风险数统计
"""

import json
import os
import re
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from _path_helper import decode_argv, reset_read_only_dimensions
decode_argv()


SEVERITY_MAP = {
    '严重': 'critical',
    '超危': 'critical',
    '超危 改为  严重': 'critical',
    '高危': 'high',
    '中危': 'medium',
    '低危': 'low',
}

SEVERITY_CN = {
    'critical': '严重',
    'high': '高危',
    'medium': '中危',
    'low': '低危',
}


def log(msg='', **kwargs):
    kwargs.pop('file', None)
    print(msg, file=sys.stderr, **kwargs, flush=True)


def normalize(val):
    return '' if val is None else str(val).strip()


def extract_ip(raw):
    if not raw:
        return None
    match = re.search(r'(\d+\.\d+\.\d+\.\d+)', str(raw))
    return match.group(1) if match else None


def normalize_asset_key(raw):
    text = normalize(raw)
    if not text:
        return ''
    return extract_ip(text) or text


def open_sheet(filepath, preferred_sheet=None):
    wb = load_workbook(filepath, read_only=True, data_only=True)
    if preferred_sheet and preferred_sheet in wb.sheetnames:
        return reset_read_only_dimensions(wb[preferred_sheet])
    return reset_read_only_dimensions(wb.active)


def find_latest_matching_file(directory, keywords):
    if not os.path.isdir(directory):
        raise FileNotFoundError(f'目录不存在: {directory}')

    candidates = []
    for entry in os.scandir(directory):
        if not entry.is_file() or not entry.name.lower().endswith('.xlsx'):
            continue
        lowered = entry.name.lower()
        if any(keyword.lower() in lowered for keyword in keywords):
            candidates.append((entry.path, entry.stat().st_mtime))

    if not candidates:
        raise FileNotFoundError(f'未找到匹配文件: {directory} keywords={keywords}')

    candidates.sort(key=lambda item: item[1], reverse=True)
    return candidates[0][0]


def build_col_map(ws):
    header = None
    header_row = 1
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        values = [normalize(cell) for cell in row]
        if any(values):
            header = values
            header_row = idx
            break
    if header is None:
        return {}, 1
    return {name: i for i, name in enumerate(header) if name}, header_row


def parse_events(filepath):
    rows = []
    ws = open_sheet(filepath, '事件表')
    col_map, header_row = build_col_map(ws)
    sev_col = col_map.get('等级')
    ip_col = col_map.get('影响资产')
    name_col = col_map.get('事件名称')
    id_col = col_map.get('事件ID')

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        severity = normalize(row[sev_col] if sev_col is not None and len(row) > sev_col else None)
        raw_ip = normalize(row[ip_col] if ip_col is not None and len(row) > ip_col else None)
        asset_ip = extract_ip(raw_ip)
        if asset_ip and severity in SEVERITY_MAP:
            rows.append({
                'asset_ip': asset_ip,
                'risk_type': '事件',
                'severity': SEVERITY_MAP[severity],
                'name': normalize(row[name_col] if name_col is not None and len(row) > name_col else None),
                'source_id': normalize(row[id_col] if id_col is not None and len(row) > id_col else None),
            })
    return rows


def parse_vulns(filepath):
    rows = []
    ws = open_sheet(filepath, '漏洞')
    col_map, header_row = build_col_map(ws)
    sev_col = col_map.get('风险等级')
    ip_col = col_map.get('风险资产')
    name_col = col_map.get('漏洞名称')

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        severity = normalize(row[sev_col] if sev_col is not None and len(row) > sev_col else None)
        raw_ip = normalize(row[ip_col] if ip_col is not None and len(row) > ip_col else None)
        asset_ip = extract_ip(raw_ip)
        if asset_ip and severity in SEVERITY_MAP:
            rows.append({
                'asset_ip': asset_ip,
                'risk_type': '漏洞',
                'severity': SEVERITY_MAP[severity],
                'name': normalize(row[name_col] if name_col is not None and len(row) > name_col else None),
                'source_id': '',
            })
    return rows


def parse_weak_passwords(filepath):
    rows = []
    ws = open_sheet(filepath, '弱口令')
    col_map, header_row = build_col_map(ws)
    ip_col = col_map.get('风险资产')
    name_col = col_map.get('弱密码名称')

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        raw_ip = normalize(row[ip_col] if ip_col is not None and len(row) > ip_col else None)
        asset_ip = extract_ip(raw_ip)
        if asset_ip:
            rows.append({
                'asset_ip': asset_ip,
                'risk_type': '弱口令',
                'severity': 'medium',
                'name': normalize(row[name_col] if name_col is not None and len(row) > name_col else None),
                'source_id': '',
            })
    return rows


def count_exposure_rows(filepath):
    wb = load_workbook(filepath, read_only=True, data_only=True)
    total = 0
    for sheet_name in ('Web服务风险分布', '非Web服务风险分布'):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        _, header_row = build_col_map(ws)
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if any(normalize(cell) for cell in row):
                total += 1
    return total


def parse_managed_assets(filepath):
    managed_assets = set()
    ws = open_sheet(filepath)
    col_map, header_row = build_col_map(ws)
    ip_col = col_map.get('IP地址')
    managed_col = col_map.get('托管状态')

    if ip_col is None or managed_col is None:
        return managed_assets

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        asset_key = normalize_asset_key(row[ip_col] if len(row) > ip_col else '')
        managed_status = normalize(row[managed_col] if len(row) > managed_col else '')
        if asset_key and managed_status == '已托管':
            managed_assets.add(asset_key)
    return managed_assets


def parse_core_assets(filepath):
    asset_map = {}
    ws = open_sheet(filepath)
    col_map, header_row = build_col_map(ws)
    ip_col = col_map.get('IP地址')
    biz_col = col_map.get('所属业务')

    if ip_col is None or biz_col is None:
        return asset_map

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        asset_ip = normalize(row[ip_col] if len(row) > ip_col else '')
        system_name = normalize(row[biz_col] if len(row) > biz_col else '')
        if asset_ip and system_name:
            asset_map[asset_ip] = system_name
    return asset_map


def count_managed_risk_rows(risk_rows, managed_assets):
    return sum(1 for row in risk_rows if normalize_asset_key(row['asset_ip']) in managed_assets)


def count_managed_high_and_above_rows(risk_rows, managed_assets):
    return sum(
        1 for row in risk_rows
        if normalize_asset_key(row['asset_ip']) in managed_assets
        and row['severity'] in ('critical', 'high')
    )


def count_managed_exposure_rows(filepath, managed_assets):
    wb = load_workbook(filepath, read_only=True, data_only=True)
    path_to_host = {}

    if '端口表' in wb.sheetnames:
        port_ws = wb['端口表']
        col_map, header_row = build_col_map(port_ws)
        path_col = col_map.get('访问路径')
        host_col = col_map.get('Host')
        if path_col is not None and host_col is not None:
            for row in port_ws.iter_rows(min_row=header_row + 1, values_only=True):
                access_path = normalize(row[path_col] if len(row) > path_col else '')
                host = normalize_asset_key(row[host_col] if len(row) > host_col else '')
                if access_path and host:
                    path_to_host[access_path] = host

    web_count = 0
    if 'Web服务风险分布' in wb.sheetnames:
        web_ws = wb['Web服务风险分布']
        col_map, header_row = build_col_map(web_ws)
        path_col = col_map.get('访问路径')
        if path_col is not None:
            for row in web_ws.iter_rows(min_row=header_row + 1, values_only=True):
                access_path = normalize(row[path_col] if len(row) > path_col else '')
                host = path_to_host.get(access_path, '')
                if host and host in managed_assets:
                    web_count += 1

    non_web_count = 0
    if '非Web服务风险分布' in wb.sheetnames:
        non_web_ws = wb['非Web服务风险分布']
        col_map, header_row = build_col_map(non_web_ws)
        host_col = col_map.get('IP地址/子域名')
        if host_col is not None:
            for row in non_web_ws.iter_rows(min_row=header_row + 1, values_only=True):
                host = normalize_asset_key(row[host_col] if len(row) > host_col else '')
                if host and host in managed_assets:
                    non_web_count += 1

    return {
        'web': web_count,
        'nonWeb': non_web_count,
        'total': web_count + non_web_count,
    }


def aggregate_core_risks(risk_rows, asset_map):
    from collections import defaultdict

    system_counts = defaultdict(lambda: {
        'critical': 0,
        'high': 0,
        'medium': 0,
        'low': 0,
        'vulnCount': 0,
        'eventCount': 0,
        'weakPwdCount': 0,
    })

    matched_risks = []
    unmatched_count = 0
    for risk in risk_rows:
        system_name = asset_map.get(risk['asset_ip'])
        if not system_name:
            unmatched_count += 1
            continue
        system_counts[system_name][risk['severity']] += 1
        if risk['risk_type'] == '漏洞':
            system_counts[system_name]['vulnCount'] += 1
        elif risk['risk_type'] == '事件':
            system_counts[system_name]['eventCount'] += 1
        elif risk['risk_type'] == '弱口令':
            system_counts[system_name]['weakPwdCount'] += 1

        risk_copy = dict(risk)
        risk_copy['system'] = system_name
        matched_risks.append(risk_copy)

    ranking = []
    for system_name, counts in system_counts.items():
        ranking.append({
            'system': system_name,
            'critical': counts['critical'],
            'high': counts['high'],
            'medium': counts['medium'],
            'low': counts['low'],
            'vulnCount': counts['vulnCount'],
            'eventCount': counts['eventCount'],
            'weakPwdCount': counts['weakPwdCount'],
            'total': counts['vulnCount'] + counts['eventCount'] + counts['weakPwdCount'],
        })

    ranking.sort(key=lambda item: (
        item['critical'],
        item['high'],
        item['medium'],
        item['low'],
        item['total'],
    ), reverse=True)
    return ranking, matched_risks, unmatched_count


def export_comparison_excel(matched_risks, ranking, output_path):
    from collections import defaultdict

    wb = Workbook()
    ws = wb.active
    ws.title = '核心业务系统风险对照'

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    sub_header_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    summary_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    white_font = Font(bold=True, size=12, color='FFFFFF')
    bold_font = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    system_order = [item['system'] for item in ranking]
    system_risks = defaultdict(list)
    for risk in matched_risks:
        system_risks[risk['system']].append((
            risk['asset_ip'],
            SEVERITY_CN[risk['severity']],
            risk['risk_type'],
            risk['name'],
        ))

    max_rows = max((len(system_risks[system_name]) for system_name in system_order), default=0)

    def group_start_col(group_idx):
        return 1 + group_idx * 5

    for idx, system_name in enumerate(system_order):
        col = group_start_col(idx)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 3)
        for current_col in range(col, col + 4):
            ws.cell(row=1, column=current_col).fill = header_fill
            ws.cell(row=1, column=current_col).border = thin_border
        title_cell = ws.cell(row=1, column=col, value=system_name)
        title_cell.font = white_font
        title_cell.fill = header_fill
        title_cell.alignment = Alignment(horizontal='center')

    for idx in range(len(system_order)):
        col = group_start_col(idx)
        for offset, label in enumerate(['IP', '等级', '类型', '名称']):
            cell = ws.cell(row=2, column=col + offset, value=label)
            cell.font = bold_font
            cell.fill = sub_header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    for row_offset in range(max_rows):
        excel_row = 3 + row_offset
        for idx, system_name in enumerate(system_order):
            col = group_start_col(idx)
            risks = system_risks[system_name]
            if row_offset >= len(risks):
                continue
            for offset, value in enumerate(risks[row_offset]):
                cell = ws.cell(row=excel_row, column=col + offset, value=value)
                cell.border = thin_border
                if offset in (1, 2):
                    cell.alignment = Alignment(horizontal='center')

    summary_items = [
        ('总数', 'total'),
        ('严重', 'critical'),
        ('高危', 'high'),
        ('中危', 'medium'),
        ('低危', 'low'),
        ('漏洞', 'vulnCount'),
        ('事件', 'eventCount'),
        ('弱口令', 'weakPwdCount'),
    ]
    summary_start_row = 4 + max_rows
    for summary_idx, (label, key) in enumerate(summary_items):
        row = summary_start_row + summary_idx
        for idx, _system_name in enumerate(system_order):
            col = group_start_col(idx)
            label_cell = ws.cell(row=row, column=col, value=label)
            value_cell = ws.cell(row=row, column=col + 1, value=ranking[idx][key])
            label_cell.font = bold_font
            value_cell.font = bold_font
            label_cell.fill = summary_fill
            value_cell.fill = summary_fill
            label_cell.border = thin_border
            value_cell.border = thin_border
            value_cell.alignment = Alignment(horizontal='center')

    for idx in range(len(system_order)):
        col = group_start_col(idx)
        ws.column_dimensions[get_column_letter(col)].width = 15
        ws.column_dimensions[get_column_letter(col + 1)].width = 8
        ws.column_dimensions[get_column_letter(col + 2)].width = 10
        ws.column_dimensions[get_column_letter(col + 3)].width = 32

    wb.save(output_path)
    log(f'中间表格已导出: {output_path}')
    return output_path


def main():
    download_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'tmp', 'exports')
    if len(sys.argv) >= 6:
        events_path, weakpwd_path, vuln_path, exposure_path, asset_path = sys.argv[1:6]
    else:
        events_path = find_latest_matching_file(download_dir, ['incident', '事件'])
        weakpwd_path = find_latest_matching_file(download_dir, ['weakpwd', '弱口令'])
        vuln_path = find_latest_matching_file(download_dir, ['vuln', '漏洞'])
        exposure_path = find_latest_matching_file(download_dir, ['exposure', '暴露面'])
        asset_path = find_latest_matching_file(download_dir, ['asset', '资产'])

    event_rows = parse_events(events_path)
    weakpwd_rows = parse_weak_passwords(weakpwd_path)
    vuln_rows = parse_vulns(vuln_path)
    exposure_total = count_exposure_rows(exposure_path)
    core_risk_rows = vuln_rows + event_rows + weakpwd_rows
    total_risk_count = len(core_risk_rows) + exposure_total

    log(
        f'漏洞: {len(vuln_rows)} 事件: {len(event_rows)} '
        f'弱口令: {len(weakpwd_rows)} 暴露面: {exposure_total} 总风险数: {total_risk_count}'
    )

    asset_map = parse_core_assets(asset_path)
    managed_assets = parse_managed_assets(asset_path)
    log(f'核心资产映射: {len(asset_map)} 条')
    log(f'已托管资产: {len(managed_assets)} 条')

    managed_vuln_count = count_managed_risk_rows(vuln_rows, managed_assets)
    managed_event_count = count_managed_risk_rows(event_rows, managed_assets)
    managed_weakpwd_count = count_managed_risk_rows(weakpwd_rows, managed_assets)
    managed_exposure = count_managed_exposure_rows(exposure_path, managed_assets)
    managed_asset_risk_count = (
        managed_vuln_count +
        managed_event_count +
        managed_weakpwd_count +
        managed_exposure['total']
    )
    managed_high_and_above_vuln_count = count_managed_high_and_above_rows(vuln_rows, managed_assets)
    managed_high_and_above_event_count = count_managed_high_and_above_rows(event_rows, managed_assets)
    managed_high_and_above_weakpwd_count = 0
    managed_high_and_above_exposure = {
        'web': 0,
        'nonWeb': 0,
        'total': 0,
    }
    managed_high_and_above_risk_count = (
        managed_high_and_above_vuln_count +
        managed_high_and_above_event_count +
        managed_high_and_above_weakpwd_count +
        managed_high_and_above_exposure['total']
    )
    log(
        f'已托管资产风险: 漏洞={managed_vuln_count} 事件={managed_event_count} '
        f'弱口令={managed_weakpwd_count} 暴露面Web={managed_exposure["web"]} '
        f'暴露面非Web={managed_exposure["nonWeb"]} 总计={managed_asset_risk_count}'
    )
    log(
        f'已托管资产高危及以上风险: 漏洞={managed_high_and_above_vuln_count} '
        f'事件={managed_high_and_above_event_count} 弱口令=0 暴露面=0 '
        f'总计={managed_high_and_above_risk_count}'
    )

    ranking, matched_risks, unmatched_count = aggregate_core_risks(core_risk_rows, asset_map)
    top5 = ranking[:5]
    ranking_names = [item['system'] for item in top5]
    max_risk_system = ranking[0]['system'] if ranking else None
    for idx, item in enumerate(top5, 1):
        log(
            f"  #{idx} {item['system']} "
            f"critical={item['critical']} high={item['high']} "
            f"medium={item['medium']} low={item['low']} total={item['total']}"
        )

    output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'output')
    os.makedirs(output_dir, exist_ok=True)
    excel_path = os.path.join(output_dir, '核心业务系统风险对照.xlsx')
    try:
        export_comparison_excel(matched_risks, ranking, excel_path)
    except PermissionError as exc:
        log(f'导出对照文件失败（文件可能被占用）: {exc}')
    except OSError as exc:
        log(f'导出对照文件失败: {exc}')

    print(json.dumps({
        'coreBusinessSystemRanking': ranking_names,
        'maxRiskSystem': max_risk_system,
        'securityRiskTotal': managed_asset_risk_count,
        'highAndAboveRiskCount': managed_high_and_above_risk_count,
    }, ensure_ascii=True, indent=2))


if __name__ == '__main__':
    main()
