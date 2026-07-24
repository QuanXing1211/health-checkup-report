#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
风险资产数计算脚本。

当前口径：
1. 事件表：取“影响资产”列；排除“处置状态”=“处置完成”的行
2. 弱口令表：取“风险资产”列；排除“处置状态”=“处置完成”的行（若该列不存在则不过滤）
3. 漏洞表：取“风险资产”列
4. 暴露面表：
   - Web服务风险分布：取“访问路径”，关联“端口表”sheet 的“访问路径”，再取对应“Host”
   - 非Web服务风险分布：取“IP地址/子域名”

输出：
- affectedAssetCount: 综合去重后的风险资产数
- riskBusinessCount: 聚合风险资产命中的业务系统数
- riskAssetTop5: 按风险数排序的风险资产 Top5
- top1BusinessSystem: 参考业务系统风险排序口径得到的第一名业务系统

说明：
- 这里只做独立能力，不接主流程。
- 暴露面/弱口令/漏洞后续接入下载时，建议统一从 tmp/exports 读取复制件。
"""
import json
import re
import sys

from openpyxl import load_workbook
from _path_helper import decode_argv, reset_read_only_dimensions
decode_argv()


IP_PATTERN = re.compile(r'(?<!\d)(?:\d{1,3}\.){3}\d{1,3}(?!\d)')
BIZ_SPLIT_RE = re.compile(r'[,，、]')
SEVERITY_MAP = {
    '严重': 'critical',
    '超危': 'critical',
    '超危 改为  严重': 'critical',
    '高危': 'high',
    '中危': 'medium',
    '低危': 'low',
}


def normalize(value):
    return '' if value is None else str(value).strip()


def split_biz(raw):
    """将"核心交易系统, 风控系统"等复合业务字段拆分为独立业务名列表"""
    if not raw:
        return []
    return [b.strip() for b in BIZ_SPLIT_RE.split(raw) if b.strip()]


def build_col_map(ws):
    header = None
    header_row = 1
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        values = [normalize(cell) for cell in row]
        if any(values):
            header = values
            header_row = idx
            break
    if header is None:
        return {}, 1
    return {name: i for i, name in enumerate(header) if name}, header_row


def open_sheet(filepath, preferred_sheet=None):
    wb = load_workbook(filepath, read_only=True, data_only=True)
    if preferred_sheet and preferred_sheet in wb.sheetnames:
        return reset_read_only_dimensions(wb[preferred_sheet])
    return reset_read_only_dimensions(wb.active)


def find_column(col_map, aliases):
    for alias in aliases:
        if alias in col_map:
            return col_map[alias]
    return None


def normalize_asset_key(raw):
    text = normalize(raw)
    if not text:
        return ''
    match = IP_PATTERN.search(text)
    return match.group(0) if match else text


def collect_assets_from_column(filepath, sheet_name, asset_col_name, status_col_name=None, skip_status=None):
    ws = open_sheet(filepath, sheet_name)
    col_map, header_row = build_col_map(ws)
    asset_col = col_map.get(asset_col_name)
    status_col = col_map.get(status_col_name) if status_col_name else None
    if asset_col is None:
        return set()

    assets = set()
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(normalize(cell) for cell in row):
            continue
        if status_col is not None and len(row) > status_col:
            status = normalize(row[status_col])
            if status == skip_status:
                continue
        asset = normalize_asset_key(row[asset_col] if len(row) > asset_col else '')
        if asset:
            assets.add(asset)
    return assets


def collect_risk_records_from_column(
    filepath,
    sheet_name,
    asset_col_name,
    source_name,
    status_col_name=None,
    skip_status=None,
    include_ids=None
):
    ws = open_sheet(filepath, sheet_name)
    col_map, header_row = build_col_map(ws)
    asset_col = col_map.get(asset_col_name)
    status_col = col_map.get(status_col_name) if status_col_name else None
    id_col = find_column(col_map, ['事件ID', '事件Id', 'incident_id', 'incidentId', 'id', 'ID']) if include_ids is not None else None
    if asset_col is None:
        return []

    include_id_set = set(normalize(item) for item in include_ids) if include_ids is not None else None
    records = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(normalize(cell) for cell in row):
            continue
        if include_id_set is not None:
            incident_id = normalize(row[id_col]) if id_col is not None and len(row) > id_col else ''
            if incident_id not in include_id_set:
                continue
        if status_col is not None and len(row) > status_col:
            status = normalize(row[status_col])
            if status == skip_status:
                continue
        asset = normalize_asset_key(row[asset_col] if len(row) > asset_col else '')
        if asset:
            records.append({
                'asset': asset,
                'source': source_name
            })
    return records


def collect_typed_risk_records_from_column(
    filepath,
    sheet_name,
    asset_col_name,
    source_name,
    severity_col_name=None,
    fixed_severity=None,
    status_col_name=None,
    skip_status=None,
    include_ids=None
):
    ws = open_sheet(filepath, sheet_name)
    col_map, header_row = build_col_map(ws)
    asset_col = col_map.get(asset_col_name)
    status_col = col_map.get(status_col_name) if status_col_name else None
    severity_col = col_map.get(severity_col_name) if severity_col_name else None
    id_col = find_column(col_map, ['事件ID', '事件Id', 'incident_id', 'incidentId', 'id', 'ID']) if include_ids is not None else None
    if asset_col is None:
        return []

    include_id_set = set(normalize(item) for item in include_ids) if include_ids is not None else None
    records = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(normalize(cell) for cell in row):
            continue
        if include_id_set is not None:
            incident_id = normalize(row[id_col]) if id_col is not None and len(row) > id_col else ''
            if incident_id not in include_id_set:
                continue
        if status_col is not None and len(row) > status_col:
            status = normalize(row[status_col])
            if status == skip_status:
                continue
        asset = normalize_asset_key(row[asset_col] if len(row) > asset_col else '')
        if not asset:
            continue

        if fixed_severity:
            severity = fixed_severity
        else:
            raw_severity = normalize(row[severity_col] if severity_col is not None and len(row) > severity_col else '')
            severity = SEVERITY_MAP.get(raw_severity)

        if not severity:
            continue

        records.append({
            'asset': asset,
            'source': source_name,
            'severity': severity
        })
    return records


def collect_exposure_assets(filepath):
    wb = load_workbook(filepath, read_only=True, data_only=True)
    assets = set()
    path_to_host = {}

    if '端口表' in wb.sheetnames:
        port_ws = wb['端口表']
        col_map, header_row = build_col_map(port_ws)
        path_col = col_map.get('访问路径')
        host_col = col_map.get('Host')
        if path_col is not None and host_col is not None:
            for row in port_ws.iter_rows(min_row=header_row + 1, values_only=True):
                if not any(normalize(cell) for cell in row):
                    continue
                access_path = normalize(row[path_col] if len(row) > path_col else '')
                host = normalize_asset_key(row[host_col] if len(row) > host_col else '')
                if access_path and host:
                    path_to_host[access_path] = host

    if 'Web服务风险分布' in wb.sheetnames:
        web_ws = wb['Web服务风险分布']
        col_map, header_row = build_col_map(web_ws)
        path_col = col_map.get('访问路径')
        if path_col is not None:
            for row in web_ws.iter_rows(min_row=header_row + 1, values_only=True):
                if not any(normalize(cell) for cell in row):
                    continue
                access_path = normalize(row[path_col] if len(row) > path_col else '')
                host = path_to_host.get(access_path, '')
                if host:
                    assets.add(host)

    if '非Web服务风险分布' in wb.sheetnames:
        non_web_ws = wb['非Web服务风险分布']
        col_map, header_row = build_col_map(non_web_ws)
        asset_col = col_map.get('IP地址/子域名')
        if asset_col is not None:
            for row in non_web_ws.iter_rows(min_row=header_row + 1, values_only=True):
                if not any(normalize(cell) for cell in row):
                    continue
                asset = normalize_asset_key(row[asset_col] if len(row) > asset_col else '')
                if asset:
                    assets.add(asset)

    return assets


def collect_exposure_risk_records(filepath):
    wb = load_workbook(filepath, read_only=True, data_only=True)
    records = []
    path_to_host = {}

    if '端口表' in wb.sheetnames:
        port_ws = wb['端口表']
        col_map, header_row = build_col_map(port_ws)
        path_col = col_map.get('访问路径')
        host_col = col_map.get('Host')
        if path_col is not None and host_col is not None:
            for row in port_ws.iter_rows(min_row=header_row + 1, values_only=True):
                if not any(normalize(cell) for cell in row):
                    continue
                access_path = normalize(row[path_col] if len(row) > path_col else '')
                host = normalize_asset_key(row[host_col] if len(row) > host_col else '')
                if access_path and host:
                    path_to_host[access_path] = host

    if 'Web服务风险分布' in wb.sheetnames:
        web_ws = wb['Web服务风险分布']
        col_map, header_row = build_col_map(web_ws)
        path_col = col_map.get('访问路径')
        if path_col is not None:
            for row in web_ws.iter_rows(min_row=header_row + 1, values_only=True):
                if not any(normalize(cell) for cell in row):
                    continue
                access_path = normalize(row[path_col] if len(row) > path_col else '')
                host = path_to_host.get(access_path, '')
                if host:
                    records.append({
                        'asset': host,
                        'source': 'exposureWeb',
                        'severity': 'medium'
                    })

    if '非Web服务风险分布' in wb.sheetnames:
        non_web_ws = wb['非Web服务风险分布']
        col_map, header_row = build_col_map(non_web_ws)
        asset_col = col_map.get('IP地址/子域名')
        if asset_col is not None:
            for row in non_web_ws.iter_rows(min_row=header_row + 1, values_only=True):
                if not any(normalize(cell) for cell in row):
                    continue
                asset = normalize_asset_key(row[asset_col] if len(row) > asset_col else '')
                if asset:
                    records.append({
                        'asset': asset,
                        'source': 'exposureNonWeb',
                        'severity': 'medium'
                    })

    return records


def build_asset_business_map(filepath):
    ws = open_sheet(filepath)
    col_map, header_row = build_col_map(ws)
    ip_col = col_map.get('IP地址')
    business_col = col_map.get('所属业务')
    if ip_col is None or business_col is None:
        return {}

    asset_business_map = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(normalize(cell) for cell in row):
            continue
        asset = normalize_asset_key(row[ip_col] if len(row) > ip_col else '')
        business = normalize(row[business_col] if len(row) > business_col else '')
        if asset and business:
            asset_business_map[asset] = business
    return asset_business_map


def build_asset_detail_map(filepath):
    ws = open_sheet(filepath)
    col_map, header_row = build_col_map(ws)
    ip_col = col_map.get('IP地址')
    business_col = col_map.get('所属业务')
    if ip_col is None:
        return {}

    asset_detail_map = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(normalize(cell) for cell in row):
            continue
        asset = normalize_asset_key(row[ip_col] if len(row) > ip_col else '')
        if not asset:
            continue
        asset_detail_map[asset] = {
            'ip': asset,
            'businessSystem': normalize(row[business_col] if business_col is not None and len(row) > business_col else '')
        }
    return asset_detail_map


def build_asset_all_business_map(filepath):
    ws = open_sheet(filepath)
    col_map, header_row = build_col_map(ws)
    ip_col = col_map.get('IP地址')
    business_col = col_map.get('所属业务')
    if ip_col is None or business_col is None:
        return {}

    asset_business_map = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(normalize(cell) for cell in row):
            continue
        asset = normalize_asset_key(row[ip_col] if len(row) > ip_col else '')
        business = normalize(row[business_col] if len(row) > business_col else '')
        if asset and business:
            asset_business_map[asset] = business
    return asset_business_map


def rank_risk_assets(typed_risk_records, risk_records, asset_detail_map, limit=5):
    asset_counts = {}

    for record in typed_risk_records:
        asset = record.get('asset')
        if not asset:
            continue
        bucket = asset_counts.setdefault(asset, {
            'critical': 0,
            'high': 0,
            'medium': 0,
            'low': 0,
            'total': 0
        })
        severity = record.get('severity')
        if severity in ('critical', 'high', 'medium', 'low'):
            bucket[severity] += 1

    for record in risk_records:
        asset = record.get('asset')
        if not asset:
            continue
        bucket = asset_counts.setdefault(asset, {
            'critical': 0,
            'high': 0,
            'medium': 0,
            'low': 0,
            'total': 0
        })
        bucket['total'] += 1

    ranking = []
    for asset, counts in asset_counts.items():
        detail = asset_detail_map.get(asset, {})
        ranking.append({
            'ip': asset,
            'businessSystem': detail.get('businessSystem', ''),
            'riskCount': counts['total'],
            '_critical': counts['critical'],
            '_high': counts['high'],
            '_medium': counts['medium'],
            '_low': counts['low']
        })

    ranking.sort(key=lambda item: (
        -item['_critical'],
        -item['_high'],
        -item['_medium'],
        -item['_low'],
        -item['riskCount'],
        item['ip']
    ))

    return [
        {
            'ip': item['ip'],
            'businessSystem': item['businessSystem'],
            'riskCount': item['riskCount']
        }
        for item in ranking[:limit]
    ]


def resolve_top1_business_system(risk_records, asset_business_map):
    business_counts = {}
    for record in risk_records:
        business_raw = asset_business_map.get(record['asset'])
        if not business_raw:
            continue
        for business in split_biz(business_raw):
            if business not in business_counts:
                business_counts[business] = {
                    'critical': 0,
                    'high': 0,
                    'medium': 0,
                    'low': 0,
                    'total': 0
                }
            severity = record.get('severity')
            if severity in business_counts[business]:
                business_counts[business][severity] += 1
            business_counts[business]['total'] += 1

    if not business_counts:
        return ''

    ranking = sorted(
        business_counts.items(),
        key=lambda item: (
            item[1]['critical'],
            item[1]['high'],
            item[1]['medium'],
            item[1]['low'],
            item[1]['total']
        ),
        reverse=True
    )
    return ranking[0][0]


def main():
    if len(sys.argv) < 6:
        raise SystemExit('Usage: risk_asset_count.py <event.xlsx> <weakpwd.xlsx> <vuln.xlsx> <exposure.xlsx> <asset.xlsx>')

    event_path, weakpwd_path, vuln_path, exposure_path, asset_path = sys.argv[1:6]
    top_risk_incident_ids = []
    if len(sys.argv) >= 7:
        try:
            parsed_ids = json.loads(sys.argv[6])
            if isinstance(parsed_ids, list):
                top_risk_incident_ids = [normalize(item) for item in parsed_ids if normalize(item)]
        except Exception:
            top_risk_incident_ids = []

    event_assets = collect_assets_from_column(event_path, '事件表', '影响资产', '处置状态', '处置完成')
    weakpwd_assets = collect_assets_from_column(weakpwd_path, '弱口令', '风险资产', '处置状态', '处置完成')
    vuln_assets = collect_assets_from_column(vuln_path, '漏洞', '风险资产')
    exposure_assets = collect_exposure_assets(exposure_path)
    risk_records = (
        collect_risk_records_from_column(event_path, '事件表', '影响资产', 'events', '处置状态', '处置完成')
        + collect_risk_records_from_column(weakpwd_path, '弱口令', '风险资产', 'weakPasswords', '处置状态', '处置完成')
        + collect_risk_records_from_column(vuln_path, '漏洞', '风险资产', 'vulnerabilities')
        + collect_exposure_risk_records(exposure_path)
    )
    typed_risk_records = (
        collect_typed_risk_records_from_column(event_path, '事件表', '影响资产', 'events', '等级', None, '处置状态', '处置完成')
        + collect_typed_risk_records_from_column(weakpwd_path, '弱口令', '风险资产', 'weakPasswords', None, 'medium', '处置状态', '处置完成')
        + collect_typed_risk_records_from_column(vuln_path, '漏洞', '风险资产', 'vulnerabilities', '风险等级')
        + collect_exposure_risk_records(exposure_path)
    )
    top_risk_records = (
        collect_risk_records_from_column(event_path, '事件表', '影响资产', 'events', '处置状态', '处置完成', top_risk_incident_ids)
        + collect_risk_records_from_column(weakpwd_path, '弱口令', '风险资产', 'weakPasswords', '处置状态', '处置完成')
        + collect_risk_records_from_column(vuln_path, '漏洞', '风险资产', 'vulnerabilities')
        + collect_exposure_risk_records(exposure_path)
    )
    top_typed_risk_records = (
        collect_typed_risk_records_from_column(event_path, '事件表', '影响资产', 'events', '等级', None, '处置状态', '处置完成', top_risk_incident_ids)
        + collect_typed_risk_records_from_column(weakpwd_path, '弱口令', '风险资产', 'weakPasswords', None, 'medium', '处置状态', '处置完成')
        + collect_typed_risk_records_from_column(vuln_path, '漏洞', '风险资产', 'vulnerabilities', '风险等级')
        + collect_exposure_risk_records(exposure_path)
    )

    all_assets = event_assets | weakpwd_assets | vuln_assets | exposure_assets
    asset_business_map = build_asset_business_map(asset_path)
    asset_all_business_map = build_asset_all_business_map(asset_path)
    asset_detail_map = build_asset_detail_map(asset_path)
    business_systems = sorted({
        biz
        for asset in all_assets
        if asset in asset_business_map
        for biz in split_biz(asset_business_map[asset])
    })
    top1_business_system = resolve_top1_business_system(typed_risk_records, asset_all_business_map)
    risk_asset_top5 = rank_risk_assets(top_typed_risk_records, top_risk_records, asset_detail_map, 5)

    print(json.dumps({
        'affectedAssetCount': len(all_assets),
        'riskBusinessCount': len(business_systems),
        'riskAssetTop5': risk_asset_top5,
        'top1BusinessSystem': top1_business_system
    }, ensure_ascii=False))


if __name__ == '__main__':
    main()
