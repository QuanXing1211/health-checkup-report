#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MSSW 资产分批导出脚本（方案 B：通过 ids 分批走 export 接口）。

当默认导出接口（不带 ids / is_all=true）因数据量过大超时时，本脚本通过：
1) count 接口拿 total
2) list 接口分页只取每条记录的 id 字段，凑齐 ids 全集（带本地缓存，1 小时内复用）
3) 把 ids 切分批次（默认每批 1000 个），多线程并发调 export 接口（带 ids 参数）
4) 下载每批生成的 xlsx 到临时目录
5) 失败批次进 dead letter，主流程结束后串行重试 1 次
6) 把所有批次 xlsx 的数据行（去掉空行 + 非首批表头）拼接到一个总 xlsx
7) 删除中间临时批次 xlsx

接口：
- POST /apps/asset/view/asset/asset_view/count?_method=GET -> { success, total }
- POST /apps/asset/view/asset/asset_view/list?_method=GET  -> { success, data: [...] }
- POST /apps/asset/view/asset/export                        -> { success, data: <filename> }
- GET  /apps/asset/view/asset/download_file?file=<filename>  -> xlsx binary

用法：
    python mssw_asset_paged_export.py \
        --cookie-path <cookie.txt> \
        --base-url sitmssw.soar.sangfor.com \
        --company-id 88602882 \
        --output-dir tmp/exports \
        [--batch-size 1000] \
        [--list-page-size 1000] \
        [--concurrency 5] \
        [--search-type current|wait_approve|both] \
        [--keep-temp] \
        [--no-cache]
"""
import argparse
import json
import os
import queue
import shutil
import signal
import ssl
import sys
import tempfile
import threading
import time
import urllib.request
import urllib.error
from concurrent.futures import ThreadPoolExecutor, as_completed

from openpyxl import load_workbook, Workbook

from _path_helper import decode_argv
decode_argv()

# 强制 stdout/stderr 使用 UTF-8，避免 Windows 默认 GBK 编码导致中文进度日志乱码
# 即使父进程已设置 PYTHONIOENCODING=utf-8，也作为防御性兜底
try:
    sys.stdout.reconfigure(encoding='utf-8', line_buffering=True)
    sys.stderr.reconfigure(encoding='utf-8', line_buffering=True)
except Exception:
    pass


# 全局终止标志：收到 SIGTERM/SIGINT 时设为 True，主流程在合适的位置检查并优雅退出
# 触发 finally 块清理临时目录，避免残留
_TERMINATING = threading.Event()


def _signal_handler(signum, frame):
    import sys as _sys
    _sys.stderr.write(f'[paged-export] 收到信号 {signum}，准备优雅退出（清理临时目录）...\n')
    _sys.stderr.flush()
    _TERMINATING.set()


def _install_signal_handlers():
    try:
        signal.signal(signal.SIGTERM, _signal_handler)
    except (ValueError, AttributeError):
        pass  # Windows 上 SIGTERM 行为有限，忽略
    try:
        signal.signal(signal.SIGINT, _signal_handler)
    except (ValueError, AttributeError):
        pass


_install_signal_handlers()


# 与 MSSW 前端抓包一致的 export_fields 结构
DEFAULT_EXPORT_FIELDS = {
    'asset_info': [
        {'disabled': True, 'key': 'ip', 'label': 'IP地址', 'selected': True},
        {'disabled': False, 'key': 'branch_name', 'label': '资产组名', 'selected': True},
        {'disabled': False, 'key': 'group_name', 'label': '所属业务', 'selected': True},
        {'disabled': False, 'key': 'org_name', 'label': '组织架构', 'selected': True},
        {'disabled': False, 'key': 'mac', 'label': 'MAC地址', 'selected': True},
        {'disabled': False, 'key': 'internet_ip', 'label': '公网IP', 'selected': True},
        {'disabled': False, 'key': 'elastic_ip', 'label': 'EIP', 'selected': True},
        {'disabled': False, 'key': 'system', 'label': '操作系统', 'selected': True},
        {'disabled': False, 'key': 'host_name', 'label': '主机名', 'selected': True},
        {'disabled': False, 'key': 'classify1', 'label': '资产类型(一级)', 'selected': True},
        {'disabled': False, 'key': 'classify2', 'label': '资产类型(二级)', 'selected': True},
        {'disabled': False, 'key': 'connect_status', 'label': 'agent状态', 'selected': True},
        {'disabled': False, 'key': 'exposure', 'label': '互联网暴露', 'selected': True},
        {'disabled': False, 'key': 'magnitude', 'label': '重要级别', 'selected': True},
        {'disabled': False, 'key': 'name', 'label': '资产名称', 'selected': True},
        {'disabled': False, 'key': 'location_detail', 'label': '资产位置', 'selected': True},
        {'disabled': False, 'key': 'tag', 'label': '资产标签', 'selected': True},
        {'disabled': False, 'key': 'data_source', 'label': '数据源', 'selected': True},
        {'disabled': False, 'key': 'inference', 'label': '资产定位推理', 'selected': True},
    ],
    'user_info': [
        {'disabled': False, 'key': 'user', 'label': '责任人', 'selected': True},
        {'disabled': False, 'key': 'mobile', 'label': '责任人电话', 'selected': True},
        {'disabled': False, 'key': 'adapter_user', 'label': '责任人(设备上报)', 'selected': True},
        {'disabled': False, 'key': 'auth_account', 'label': '实时认证用户名', 'selected': True},
    ],
    'custom_attribute': [],
    'mss_service_info': [
        {'disabled': False, 'key': 'managed_level', 'label': '托管状态', 'selected': True},
        {'disabled': False, 'key': 'mss_service_id', 'label': '增值服务', 'selected': True},
    ],
    'finger_print_info': [
        {'disabled': False, 'key': 'port_service', 'label': '端口服务', 'selected': True},
        {'disabled': False, 'key': 'software_name', 'label': '软件名称', 'selected': True},
        {'disabled': False, 'key': 'database_type', 'label': '数据库类型', 'selected': True},
        {'disabled': False, 'key': 'web_service_name', 'label': 'web服务名', 'selected': True},
        {'disabled': False, 'key': 'framework_name', 'label': '框架名', 'selected': True},
        {'disabled': False, 'key': 'app_name', 'label': '应用名', 'selected': True},
        {'disabled': False, 'key': 'domain_name', 'label': '域名', 'selected': True},
    ],
}


# ids 缓存有效期：1 小时（避免 ids 过期导致漏数据）
IDS_CACHE_TTL_SECONDS = 3600


def build_headers(cookie_str, base_url, company_id):
    csrf_token = ''
    for pair in cookie_str.split(';'):
        pair = pair.strip()
        if pair.startswith('csrf_token='):
            csrf_token = pair[len('csrf_token='):]
            break
    return {
        'Content-Type': 'application/json',
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Cookie': cookie_str,
        'Origin': f'https://{base_url}',
        'Referer': f'https://{base_url}/index.html',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36',
        'X-Requested-With': 'XMLHttpRequest',
        'x-mssw-company-id': str(company_id or ''),
        'x-csrf-token': csrf_token,
    }


_SSL_CTX = ssl.create_default_context()
_SSL_CTX.check_hostname = False
_SSL_CTX.verify_mode = ssl.CERT_NONE


def post_json(url, body, headers, timeout=60):
    data = json.dumps(body, ensure_ascii=False).encode('utf-8')
    req = urllib.request.Request(url, data=data, headers=headers, method='POST')
    try:
        with urllib.request.urlopen(req, timeout=timeout, context=_SSL_CTX) as resp:
            raw = resp.read().decode('utf-8')
            return json.loads(raw)
    except urllib.error.HTTPError as e:
        body_text = ''
        try:
            body_text = e.read().decode('utf-8', errors='ignore')
        except Exception:
            pass
        raise RuntimeError(f'HTTP {e.code} {e.reason}: {body_text[:500]}')
    except urllib.error.URLError as e:
        raise RuntimeError(f'URLError: {e.reason}')


def download_binary(url, headers, timeout=120):
    req = urllib.request.Request(url, headers=headers, method='GET')
    try:
        with urllib.request.urlopen(req, timeout=timeout, context=_SSL_CTX) as resp:
            return resp.read()
    except urllib.error.HTTPError as e:
        body_text = ''
        try:
            body_text = e.read().decode('utf-8', errors='ignore')
        except Exception:
            pass
        raise RuntimeError(f'下载失败 HTTP {e.code} {e.reason}: {body_text[:500]}')


def fetch_total(base_url, headers, search_type):
    url = f'https://{base_url}/apps/asset/view/asset/asset_view/count?_method=GET'
    body = {'branch_id': 'all', 'search_type': search_type, 'start': 0, 'limit': 20}
    resp = post_json(url, body, headers, timeout=30)
    if not resp or resp.get('success') is not True:
        raise RuntimeError(f'count 接口返回异常: {json.dumps(resp, ensure_ascii=False)[:500]}')
    return int(resp.get('total') or 0)


def fetch_ids_page(base_url, headers, search_type, start, limit):
    url = f'https://{base_url}/apps/asset/view/asset/asset_view/list?_method=GET'
    body = {'branch_id': 'all', 'search_type': search_type, 'start': start, 'limit': limit}
    resp = post_json(url, body, headers, timeout=120)
    if not resp or resp.get('success') is not True:
        raise RuntimeError(f'list 接口返回异常 (start={start}): {json.dumps(resp, ensure_ascii=False)[:500]}')
    data = resp.get('data') or []
    return data if isinstance(data, list) else []


# ----------------- ids 缓存 -----------------

def cache_dir(output_dir):
    p = os.path.join(output_dir, '.ids_cache')
    os.makedirs(p, exist_ok=True)
    return p


def cache_path(output_dir, company_id, search_type):
    return os.path.join(cache_dir(output_dir), f'{company_id}_{search_type}.json')


def load_ids_cache(path, expected_total):
    if not os.path.isfile(path):
        return None
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception:
        return None
    if not isinstance(data, dict):
        return None
    if int(data.get('total') or -1) != int(expected_total):
        return None
    cached_at = float(data.get('cached_at') or 0)
    if cached_at <= 0:
        return None
    age = time.time() - cached_at
    if age > IDS_CACHE_TTL_SECONDS:
        return None
    ids = data.get('ids') or []
    if not isinstance(ids, list) or not ids:
        return None
    # 历史缓存可能因 MSSW list 接口分页不稳定含重复 id，这里兜底去重（保持首次出现顺序）
    seen = {}
    deduped = []
    for i in ids:
        try:
            i_int = int(i)
        except (TypeError, ValueError):
            continue
        if i_int not in seen:
            seen[i_int] = True
            deduped.append(i_int)
    if len(deduped) > int(expected_total):
        # 去重后仍超过 total，缓存异常，丢弃
        return None
    return deduped


def save_ids_cache(path, ids, total):
    try:
        with open(path, 'w', encoding='utf-8') as f:
            json.dump({'total': int(total), 'cached_at': time.time(), 'ids': [int(i) for i in ids]}, f)
    except Exception:
        pass


# ----------------- ids 收集 -----------------

def collect_all_ids(base_url, headers, search_type, list_page_size, list_concurrency, log_progress, output_dir, company_id, use_cache):
    total = fetch_total(base_url, headers, search_type)
    log_progress(f'[ids] {search_type} total={total}, list_page_size={list_page_size}, list_concurrency={list_concurrency}')

    cache_file = cache_path(output_dir, company_id, search_type)
    if use_cache and total > 0:
        cached = load_ids_cache(cache_file, total)
        if cached is not None:
            log_progress(f'[ids] {search_type} 命中缓存: {len(cached)} 个 ids (total={total})，跳过 list 阶段')
            return cached

    if total <= 0:
        return []

    # 预计算所有需要拉的 (page_no, start)
    pages = []
    start = 0
    page_no = 0
    while start < total:
        page_no += 1
        pages.append({'no': page_no, 'start': start})
        start += list_page_size

    log_progress(f'[ids] {search_type} 需拉取 {len(pages)} 页, 并发 {list_concurrency}')

    page_ids_map = {}  # page_no -> list[int]
    page_ids_lock = threading.Lock()

    def fetch_one_page(page):
        no = page['no']
        start = page['start']
        retry = 0
        last_err = None
        while retry < 3:
            try:
                page_data = fetch_ids_page(base_url, headers, search_type, start, list_page_size)
                ids = [int(r['id']) for r in page_data if isinstance(r, dict) and r.get('id') is not None]
                with page_ids_lock:
                    page_ids_map[no] = ids
                return no, len(ids), None
            except Exception as e:
                last_err = e
                retry += 1
                msg = str(e)
                if 'HTTP 401' in msg or 'HTTP 403' in msg:
                    log_progress(f'[ids] {search_type} page={no} start={start} 鉴权失败，跳过重试: {msg[:200]}')
                    break
                log_progress(f'[ids] {search_type} page={no} start={start} retry={retry} err={msg[:200]}')
                time.sleep(2)
        return no, 0, last_err

    failed_pages = []
    with ThreadPoolExecutor(max_workers=list_concurrency) as executor:
        futures = [executor.submit(fetch_one_page, p) for p in pages]
        completed = 0
        for fut in as_completed(futures):
            try:
                no, cnt, err = fut.result()
                completed += 1
                if err is not None:
                    failed_pages.append((no, err))
                else:
                    log_progress(f'[ids] {search_type} page={no} fetched ids={cnt} completed={completed}/{len(pages)}')
            except Exception as e:
                log_progress(f'[ids] {search_type} 页拉取异常: {e}')

    if failed_pages:
        err_summary = '; '.join(f'page={n}: {str(e)[:100]}' for n, e in failed_pages[:3])
        raise RuntimeError(f'{search_type} 有 {len(failed_pages)} 页 ids 拉取失败: {err_summary}')

    # 按 page_no 顺序拼接，保证 ids 顺序稳定
    # 注意：MSSW list 接口并发分页拉取时，后端排序不稳定，
    # 不同 page 边界的 ids 可能跨页重复（实测 201703 中有 141 个重复）。
    # 这里用 dict 去重，保留首次出现的顺序，避免后续 export/merge 阶段产生重复数据。
    seen_ids = {}
    for p in pages:
        ids = page_ids_map.get(p['no'])
        if ids:
            for i in ids:
                if i not in seen_ids:
                    seen_ids[i] = True
    all_ids = list(seen_ids.keys())
    raw_count = sum(len(page_ids_map.get(p['no'], [])) for p in pages)
    dup_count = raw_count - len(all_ids)
    log_progress(f'[ids] {search_type} 完成: 共 {len(all_ids)} 个 ids (接口 total={total}, 去重前 {raw_count}, 重复 {dup_count})')

    if use_cache:
        save_ids_cache(cache_file, all_ids, total)
        log_progress(f'[ids] {search_type} 缓存已写入: {cache_file}')

    return all_ids


# ----------------- 单批次导出 + 下载 -----------------

def export_batch(base_url, headers, search_type, ids_chunk, export_fields):
    url = f'https://{base_url}/apps/asset/view/asset/export'
    body = {
        'branch_id': 'all',
        'search_type': search_type,
        'is_all': False,
        'ids': ids_chunk,
        'exclude_ids': [],
        'export_fields': export_fields,
    }
    resp = post_json(url, body, headers, timeout=300)
    if not resp or resp.get('success') is not True:
        raise RuntimeError(f'export 接口返回异常 (batch_size={len(ids_chunk)}): {json.dumps(resp, ensure_ascii=False)[:500]}')
    filename = resp.get('data')
    if not filename:
        raise RuntimeError(f'export 接口未返回文件名: {json.dumps(resp, ensure_ascii=False)[:500]}')
    return filename


def download_batch(base_url, headers, filename, dest_path):
    url = f'https://{base_url}/apps/asset/view/asset/download_file?file={urllib.parse.quote(filename)}'
    data = download_binary(url, headers, timeout=120)
    with open(dest_path, 'wb') as f:
        f.write(data)
    return dest_path


def export_and_download_batch(base_url, headers, search_type, ids_chunk, export_fields, dest_dir, batch_no, total_batches, log_progress):
    t0 = time.time()
    filename = export_batch(base_url, headers, search_type, ids_chunk, export_fields)
    batch_path = os.path.join(dest_dir, f'batch_{batch_no:04d}_{filename}')
    download_batch(base_url, headers, filename, batch_path)
    elapsed = time.time() - t0
    log_progress(
        f'[export] {search_type} batch={batch_no}/{total_batches} '
        f'ids={len(ids_chunk)} elapsed={elapsed:.1f}s'
    )
    return batch_path


# ----------------- 合并 xlsx -----------------

def _copy_cell_style(src_cell, dst_cell):
    """从源单元格复制样式（font/fill/alignment/border/number_format）到目标单元格。"""
    if src_cell.has_style:
        dst_cell.font = src_cell.font.copy()
        dst_cell.fill = src_cell.fill.copy()
        dst_cell.alignment = src_cell.alignment.copy()
        dst_cell.border = src_cell.border.copy()
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = src_cell.protection.copy()


def _copy_column_widths(src_ws, dst_ws):
    """复制源 sheet 的列宽到目标 sheet。"""
    for col_letter, col_dim in src_ws.column_dimensions.items():
        if col_dim.width is not None:
            dst_ws.column_dimensions[col_letter].width = col_dim.width
        if col_dim.hidden:
            dst_ws.column_dimensions[col_letter].hidden = True


def _copy_row_dimensions(src_ws, dst_ws, row_idx):
    """复制源 sheet 指定行的行高到目标 sheet。"""
    src_row_dim = src_ws.row_dimensions.get(row_idx)
    if src_row_dim and src_row_dim.height is not None:
        dst_ws.row_dimensions[row_idx].height = src_row_dim.height


def merge_xlsx_batches(batch_files, output_path, log_progress):
    """合并多个 xlsx 到一个总 xlsx，保留首批批次的表头样式 + 列宽。

    MSSW export xlsx 固定结构：第 1 行空 + 第 2 行表头 + 第 3 行起数据。
    合并规则（输出 xlsx 第 1 行直接是表头，无空行）：
    - 首批：跳过第 1 行空行，第 2 行表头作为输出 xlsx 的第 1 行（连同样式 + 列宽）
    - 后续批次：跳过前 2 行（空行 + 表头），只追加数据（不复制样式以节省内存）

    使用普通模式（非 write_only）以支持表头样式保留。数据行仅 append 值，
    20w+ 行的内存压力通过 read_only 加载批次 + 即时关闭控制。
    """
    if not batch_files:
        raise RuntimeError('merge_xlsx_batches: 没有批次文件可合并')

    batch_files_sorted = sorted(batch_files, key=lambda p: os.path.basename(p))

    wb_out = Workbook(write_only=False)
    ws_out = wb_out.active
    ws_out.title = '资产清单'

    header_written = False
    total_rows = 0

    for idx, batch_path in enumerate(batch_files_sorted):
        try:
            # 首批需要读样式 + 列宽，用普通模式；后续批次只需要值，用 read_only 提速
            wb = load_workbook(batch_path, read_only=header_written)
        except Exception as e:
            log_progress(f'[merge] 读取批次 {idx + 1} 失败: {batch_path} -> {e}')
            raise
        ws = wb.active

        if not header_written:
            # 首批：跳过第 1 行空行，第 2 行表头作为输出 xlsx 的第 1 行（保留样式 + 列宽）
            _copy_column_widths(ws, ws_out)
            out_row_idx = 0
            for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
                if row_idx == 1:
                    # 跳过空行
                    continue
                out_row_idx += 1
                for col_idx, cell in enumerate(row, start=1):
                    out_cell = ws_out.cell(row=out_row_idx, column=col_idx, value=cell.value)
                    _copy_cell_style(cell, out_cell)
                if row_idx == 2:
                    # 表头行额外复制行高
                    _copy_row_dimensions(ws, ws_out, row_idx)
                total_rows += 1
                if row_idx >= 2:
                    # 首批已写入表头 + 数据，继续读完所有数据
                    pass
            header_written = True
            total_rows -= 1  # 减去表头
            log_progress(f'[merge] 批次 {idx + 1}/{len(batch_files_sorted)}: 写入表头 + 数据, 累计 {total_rows} 行')
        else:
            # 后续批次：只 append 值，跳过前 2 行（空 + 表头）
            skipped = 0
            for row in ws.iter_rows(values_only=True):
                if skipped < 2:
                    skipped += 1
                    continue
                ws_out.append(list(row))
                total_rows += 1
            log_progress(f'[merge] 批次 {idx + 1}/{len(batch_files_sorted)}: 追加数据, 累计 {total_rows} 行')
        wb.close()

    wb_out.save(output_path)
    log_progress(f'[merge] 完成: {output_path} ({total_rows} 行数据)')
    return output_path


# ----------------- 单 search_type 完整流程（生产者-消费者） -----------------

def export_search_type(base_url, headers, search_type, output_dir, batch_size, list_page_size, concurrency, list_concurrency, keep_temp, use_cache, log_progress, company_id):
    ids = collect_all_ids(base_url, headers, search_type, list_page_size, list_concurrency, log_progress, output_dir, company_id, use_cache)
    if not ids:
        ts = time.strftime('%Y%m%d%H%M%S')
        empty_path = os.path.join(output_dir, f'Asset_Paged_{ts}_{search_type}.xlsx')
        wb = Workbook()
        ws = wb.active
        ws.title = '资产清单'
        header_row = []
        for group in ('asset_info', 'user_info', 'custom_attribute', 'mss_service_info', 'finger_print_info'):
            for field in DEFAULT_EXPORT_FIELDS.get(group, []):
                if field.get('selected'):
                    header_row.append(field.get('label', ''))
        ws.append(header_row)
        wb.save(empty_path)
        wb.close()
        log_progress(f'[export] {search_type} 无数据，生成仅含表头的 xlsx: {empty_path}')
        return empty_path

    total_batches = (len(ids) + batch_size - 1) // batch_size
    log_progress(f'[export] {search_type} 开始分批导出: {len(ids)} ids -> {total_batches} 批 (batch_size={batch_size}, export_concurrency={concurrency}, download_concurrency={concurrency})')

    temp_dir = tempfile.mkdtemp(prefix=f'mssw_paged_{search_type}_')
    log_progress(f'[export] 临时目录: {temp_dir}')

    batches = []
    for i, start in enumerate(range(0, len(ids), batch_size)):
        batches.append({
            'no': i + 1,
            'ids': ids[start:start + batch_size],
        })

    # 共享状态
    batch_files = {b['no']: None for b in batches}  # batch_no -> local path
    export_dead = []    # export 失败: (no, ids, err)
    download_dead = []  # download 失败: (no, ids, filename, err)
    state_lock = threading.Lock()
    completed_count = [0]
    download_queue = queue.Queue()  # 传递 ('ready', no, ids, filename) 或 ('stop', None, None, None)

    def producer(batch):
        """生产者：调 export 接口拿 filename，推入 download_queue。"""
        if _TERMINATING.is_set():
            return
        no = batch['no']
        ids_chunk = batch['ids']
        t0 = time.time()
        last_err = None
        filename = None
        for retry in range(3):
            if _TERMINATING.is_set():
                return
            try:
                filename = export_batch(base_url, headers, search_type, ids_chunk, DEFAULT_EXPORT_FIELDS)
                last_err = None
                break
            except Exception as e:
                last_err = e
                msg = str(e)
                if 'HTTP 401' in msg or 'HTTP 403' in msg:
                    log_progress(f'[export] {search_type} batch={no} 鉴权失败，跳过重试: {msg[:200]}')
                    break
                log_progress(f'[export] {search_type} batch={no} export retry={retry} err={msg[:200]}')
                time.sleep(2)
        elapsed = time.time() - t0
        if last_err is not None or filename is None:
            with state_lock:
                export_dead.append((no, ids_chunk, last_err or RuntimeError('unknown')))
            log_progress(f'[export] {search_type} batch={no} export 失败 ({elapsed:.1f}s): {str(last_err)[:200]}')
            return
        log_progress(f'[export] {search_type} batch={no} export 完成 ({elapsed:.1f}s)，入下载队列')
        download_queue.put(('ready', no, ids_chunk, filename))

    def consumer():
        """消费者：从 download_queue 拉 (no, filename)，下载到本地。"""
        while True:
            try:
                item = download_queue.get(timeout=1)
            except queue.Empty:
                if _TERMINATING.is_set():
                    return
                continue
            if item is None:
                download_queue.task_done()
                return
            kind, no, ids_chunk, filename = item
            if kind != 'ready':
                download_queue.task_done()
                continue
            t0 = time.time()
            last_err = None
            for retry in range(3):
                if _TERMINATING.is_set():
                    return
                try:
                    batch_path = os.path.join(temp_dir, f'batch_{no:04d}_{filename}')
                    download_batch(base_url, headers, filename, batch_path)
                    with state_lock:
                        batch_files[no] = batch_path
                        completed_count[0] += 1
                    last_err = None
                    break
                except Exception as e:
                    last_err = e
                    msg = str(e)
                    if 'HTTP 401' in msg or 'HTTP 403' in msg:
                        log_progress(f'[export] {search_type} batch={no} download 鉴权失败，跳过重试: {msg[:200]}')
                        break
                    log_progress(f'[export] {search_type} batch={no} download retry={retry} err={msg[:200]}')
                    time.sleep(2)
            elapsed = time.time() - t0
            if last_err is not None:
                with state_lock:
                    download_dead.append((no, ids_chunk, filename, last_err))
                log_progress(f'[export] {search_type} batch={no} download 失败 ({elapsed:.1f}s): {str(last_err)[:200]}')
            else:
                log_progress(f'[export] {search_type} batch={no}/{total_batches} download 完成 ({elapsed:.1f}s), 累计 {completed_count[0]}')
            download_queue.task_done()

    try:
        # 启动 download 消费者（固定 concurrency 路线程）
        download_threads = []
        for _ in range(concurrency):
            t = threading.Thread(target=consumer, daemon=True)
            t.start()
            download_threads.append(t)

        # 启动 export 生产者（concurrency 路）
        with ThreadPoolExecutor(max_workers=concurrency) as export_executor:
            futures = [export_executor.submit(producer, b) for b in batches]
            for fut in as_completed(futures):
                try:
                    fut.result()
                except Exception as e:
                    log_progress(f'[export] {search_type} producer 异常: {e}')

        # 等所有 download 完成
        download_queue.join()
        # 通知 download worker 退出
        for _ in range(concurrency):
            download_queue.put(None)
        for t in download_threads:
            t.join(timeout=10)

        if _TERMINATING.is_set():
            log_progress(f'[export] {search_type} 检测到终止信号，跳过重试和合并，直接清理临时目录')
            raise RuntimeError(f'{search_type} 被终止')

        log_progress(f'[export] {search_type} 主流程完成: 成功 {completed_count[0]}/{total_batches}, export失败 {len(export_dead)}, download失败 {len(download_dead)}')

        # 失败批次串行重试 1 次
        all_failed = [(no, ids, err) for (no, ids, err) in export_dead]
        all_failed.extend((no, ids, err) for (no, ids, _, err) in download_dead)
        if all_failed:
            log_progress(f'[export] {search_type} 开始重试 {len(all_failed)} 个失败批次')
            retry_failed = []
            for no, ids_chunk, _ in all_failed:
                try:
                    t0 = time.time()
                    filename = export_batch(base_url, headers, search_type, ids_chunk, DEFAULT_EXPORT_FIELDS)
                    batch_path = os.path.join(temp_dir, f'batch_{no:04d}_{filename}')
                    download_batch(base_url, headers, filename, batch_path)
                    with state_lock:
                        batch_files[no] = batch_path
                        completed_count[0] += 1
                    log_progress(f'[export] {search_type} batch={no} 重试成功 ({time.time() - t0:.1f}s)')
                except Exception as e:
                    retry_failed.append((no, ids_chunk, e))
                    log_progress(f'[export] {search_type} batch={no} 重试仍失败: {e}')

            if retry_failed:
                err_summary = '; '.join(f'batch={n}: {str(e)[:100]}' for n, _, e in retry_failed[:3])
                raise RuntimeError(f'{search_type} {len(retry_failed)} 个批次重试后仍失败: {err_summary}')

        ordered_paths = [batch_files[b['no']] for b in batches if batch_files.get(b['no'])]
        if not ordered_paths:
            raise RuntimeError(f'{search_type} 无任何批次成功，无法合并')

        ts = time.strftime('%Y%m%d%H%M%S')
        output_path = os.path.join(output_dir, f'Asset_Paged_{ts}_{search_type}.xlsx')
        merge_xlsx_batches(ordered_paths, output_path, log_progress)
        return output_path
    finally:
        if not keep_temp:
            try:
                shutil.rmtree(temp_dir, ignore_errors=True)
                log_progress(f'[export] {search_type} 已清理临时目录')
            except Exception as e:
                log_progress(f'[export] {search_type} 清理临时目录失败（不影响主流程）: {e}')
        else:
            log_progress(f'[export] {search_type} 保留临时目录: {temp_dir}')


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument('--cookie-path', required=True)
    parser.add_argument('--base-url', default='sitmssw.soar.sangfor.com')
    parser.add_argument('--company-id', required=True)
    parser.add_argument('--output-dir', default='tmp/exports')
    parser.add_argument('--batch-size', type=int, default=1000)
    parser.add_argument('--list-page-size', type=int, default=1000)
    parser.add_argument('--concurrency', type=int, default=20, help='export 阶段并发数')
    parser.add_argument('--list-concurrency', type=int, default=10, help='list 阶段并发数')
    parser.add_argument('--search-type', default='both', choices=['current', 'wait_approve', 'both'])
    parser.add_argument('--keep-temp', action='store_true')
    parser.add_argument('--no-cache', action='store_true')
    return parser.parse_args()


def main():
    args = parse_args()

    if not os.path.isfile(args.cookie_path):
        raise SystemExit(f'cookie 文件不存在: {args.cookie_path}')
    with open(args.cookie_path, 'r', encoding='utf-8') as f:
        cookie_str = f.read().strip()
    if not cookie_str:
        raise SystemExit(f'cookie 文件为空: {args.cookie_path}')

    base_url = args.base_url.strip().strip('/')
    os.makedirs(args.output_dir, exist_ok=True)
    headers = build_headers(cookie_str, base_url, args.company_id)

    def log_progress(msg):
        print(msg, flush=True)

    result = {}
    if args.search_type in ('current', 'both'):
        result['currentFilePath'] = export_search_type(
            base_url, headers, 'current', args.output_dir,
            args.batch_size, args.list_page_size, args.concurrency, args.list_concurrency,
            args.keep_temp, not args.no_cache, log_progress, args.company_id
        )
    if args.search_type in ('wait_approve', 'both'):
        try:
            result['waitApproveFilePath'] = export_search_type(
                base_url, headers, 'wait_approve', args.output_dir,
                args.batch_size, args.list_page_size, args.concurrency, args.list_concurrency,
                args.keep_temp, not args.no_cache, log_progress, args.company_id
            )
        except Exception as e:
            log_progress(f'[export] wait_approve 失败（不阻断）: {e}')
            result['waitApproveFilePath'] = ''

    print('###JSON###' + json.dumps(result, ensure_ascii=False))


if __name__ == '__main__':
    main()
