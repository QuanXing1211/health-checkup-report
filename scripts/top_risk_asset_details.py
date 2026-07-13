#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
按风险资产 TOP5 统计风险详情。

用法:
  top_risk_asset_details.py <incident.xlsx> <asset.xlsx> <weakpwd.xlsx> <vuln.xlsx> <exposure.xlsx> <top_assets_json> <c2_ids_json> <virus_ids_json> <exploit_ids_json>
"""

import json
import re
import sys

from openpyxl import load_workbook

from _path_helper import decode_argv
decode_argv()


IP_PATTERN = re.compile(r'(?<!\d)(?:\d{1,3}\.){3}\d{1,3}(?!\d)')


def normalize(value):
    return "" if value is None else str(value).strip()


def normalize_asset_key(raw):
    text = normalize(raw)
    if not text:
        return ""
    match = IP_PATTERN.search(text)
    return match.group(0) if match else text


def build_col_map(ws):
    header = None
    header_row = 1
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        values = [normalize(cell) for cell in row]
        if len([value for value in values if value]) >= 2:
            header = values
            header_row = idx
            break
    if header is None:
        return {}, 1
    return {name: i for i, name in enumerate(header) if name}, header_row


def build_col_map_allow_single(ws):
    header = None
    header_row = 1
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        values = [normalize(cell) for cell in row]
        if any(values):
            header = values
            header_row = idx
            break
    if header is None:
        return {}, 1
    return {name: i for i, name in enumerate(header) if name}, header_row


def find_column(col_map, aliases):
    for alias in aliases:
        if alias in col_map:
            return col_map[alias]
    return None


def get_cell_value(ws, row_index, col_index):
    cell = ws.cell(row=row_index, column=col_index + 1)
    value = normalize(cell.value)
    if value:
        return value

    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return normalize(ws.cell(row=merged_range.min_row, column=merged_range.min_col).value)

    return ""


def add_example(examples, value, limit=2):
    text = normalize(value)
    if not text or text in examples:
        return
    examples.append(text)
    if len(examples) > limit:
        del examples[limit:]


def collect_incident_counts(incident_path, target_assets, c2_ids, virus_ids, exploit_ids):
    counts = {
        asset: {
            "totalEvents": 0,
            "c2Events": 0,
            "virusTrojanEvents": 0,
            "vulnExploitEvents": 0,
        }
        for asset in target_assets
    }
    if not incident_path or not target_assets:
        return counts

    workbook = load_workbook(incident_path, read_only=True, data_only=True)
    sheet = workbook.active
    col_map, header_row = build_col_map(sheet)
    id_col = find_column(col_map, ["事件ID", "事件Id", "incident_id", "incidentId", "id", "ID"])
    asset_col = find_column(col_map, ["影响资产", "受影响资产", "host_ip", "hostIp", "主机IP", "ip"])
    status_col = find_column(col_map, ["处置状态", "disposalStatus", "status"])
    if asset_col is None:
        workbook.close()
        return counts

    target_set = set(target_assets)
    c2_id_set = set(c2_ids)
    virus_id_set = set(virus_ids)
    exploit_id_set = set(exploit_ids)
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(normalize(cell) for cell in row):
            continue
        status = normalize(row[status_col]) if status_col is not None and len(row) > status_col else ""
        if status == "处置完成":
            continue
        asset = normalize_asset_key(row[asset_col] if len(row) > asset_col else "")
        if asset not in target_set:
            continue

        incident_id = normalize(row[id_col]) if id_col is not None and len(row) > id_col else ""

        bucket = counts[asset]
        bucket["totalEvents"] += 1
        if incident_id in c2_id_set:
            bucket["c2Events"] += 1
        if incident_id in virus_id_set:
            bucket["virusTrojanEvents"] += 1
        if incident_id in exploit_id_set:
            bucket["vulnExploitEvents"] += 1

    workbook.close()
    return counts


def collect_aes_install_status(asset_path, target_assets):
    status = {asset: False for asset in target_assets}
    if not asset_path or not target_assets:
        return status

    try:
        workbook = load_workbook(asset_path, read_only=True, data_only=True)
    except Exception:
        return status

    sheet = workbook.active
    col_map, header_row = build_col_map(sheet)
    ip_col = find_column(col_map, [
        "IP地址", "P地址", "IP", "地址", "主机IP", "hostIp", "host_ip", "ip", "资产IP"
    ])
    data_source_col = find_column(col_map, [
        "数据源", "数据来源", "dataSource", "data_source", "设备来源", "devSourceNames", "source"
    ])
    if ip_col is None:
        workbook.close()
        return status

    target_set = set(target_assets)
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(normalize(cell) for cell in row):
            continue
        asset = normalize_asset_key(row[ip_col] if len(row) > ip_col else "")
        if asset not in target_set:
            continue
        data_source = normalize(row[data_source_col]) if data_source_col is not None and len(row) > data_source_col else ""
        status[asset] = "EDR" in data_source.upper()

    workbook.close()
    return status


def collect_vulnerability_counts(vuln_path, target_assets):
    counts = {
        asset: {
            "totalVulnerabilities": 0,
            "highAndAboveVulnerabilities": 0,
        }
        for asset in target_assets
    }
    if not vuln_path or not target_assets:
        return counts

    workbook = load_workbook(vuln_path, read_only=True, data_only=True)
    sheet = workbook["漏洞"] if "漏洞" in workbook.sheetnames else workbook.active
    col_map, header_row = build_col_map(sheet)
    asset_col = find_column(col_map, ["风险资产", "影响资产", "资产IP", "IP地址", "ip"])
    severity_col = find_column(col_map, ["风险等级", "等级", "severity"])
    if asset_col is None:
        workbook.close()
        return counts

    target_set = set(target_assets)
    high_and_above = {"严重", "超危", "高危", "超危 改为  严重"}
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(normalize(cell) for cell in row):
            continue
        asset = normalize_asset_key(row[asset_col] if len(row) > asset_col else "")
        if asset not in target_set:
            continue
        counts[asset]["totalVulnerabilities"] += 1
        severity = normalize(row[severity_col]) if severity_col is not None and len(row) > severity_col else ""
        if severity in high_and_above:
            counts[asset]["highAndAboveVulnerabilities"] += 1

    workbook.close()
    return counts


def collect_weak_password_counts(weakpwd_path, target_assets):
    counts = {asset: {"weakPasswords": 0} for asset in target_assets}
    if not weakpwd_path or not target_assets:
        return counts

    workbook = load_workbook(weakpwd_path, read_only=True, data_only=True)
    sheet = workbook["弱口令"] if "弱口令" in workbook.sheetnames else workbook.active
    col_map, header_row = build_col_map(sheet)
    asset_col = find_column(col_map, ["风险资产", "影响资产", "资产IP", "IP地址", "ip"])
    status_col = find_column(col_map, ["处置状态", "disposalStatus", "status"])
    if asset_col is None:
        workbook.close()
        return counts

    target_set = set(target_assets)
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(normalize(cell) for cell in row):
            continue
        status = normalize(row[status_col]) if status_col is not None and len(row) > status_col else ""
        if status == "处置完成":
            continue
        asset = normalize_asset_key(row[asset_col] if len(row) > asset_col else "")
        if asset in target_set:
            counts[asset]["weakPasswords"] += 1

    workbook.close()
    return counts


def collect_exposure_counts(exposure_path, target_assets):
    counts = {
        asset: {
            "totalExposures": 0,
            "webExposures": 0,
            "nonWebExposures": 0,
            "webComponentExamples": [],
            "nonWebServiceExamples": [],
        }
        for asset in target_assets
    }
    if not exposure_path or not target_assets:
        return counts

    workbook = load_workbook(exposure_path, read_only=False, data_only=True)
    target_set = set(target_assets)
    path_to_host = {}

    if "端口表" in workbook.sheetnames:
        port_sheet = workbook["端口表"]
        col_map, header_row = build_col_map(port_sheet)
        path_col = find_column(col_map, ["访问路径", "url", "URL"])
        host_col = find_column(col_map, ["Host", "host", "IP地址", "资产IP"])
        if path_col is not None and host_col is not None:
            for row in port_sheet.iter_rows(min_row=header_row + 1, values_only=True):
                if not any(normalize(cell) for cell in row):
                    continue
                access_path = normalize(row[path_col] if len(row) > path_col else "")
                host = normalize_asset_key(row[host_col] if len(row) > host_col else "")
                if access_path and host:
                    path_to_host[access_path] = host

    if "Web服务风险分布" in workbook.sheetnames:
        web_sheet = workbook["Web服务风险分布"]
        col_map, header_row = build_col_map_allow_single(web_sheet)
        path_col = find_column(col_map, ["访问路径", "url", "URL"])
        component_col = find_column(col_map, ["组件名称", "组件", "服务", "服务名称", "应用组件", "component"])
        if path_col is not None:
            for row_index in range(header_row + 1, web_sheet.max_row + 1):
                row_values = [web_sheet.cell(row=row_index, column=col).value for col in range(1, web_sheet.max_column + 1)]
                if not any(normalize(cell) for cell in row_values):
                    continue
                access_path = get_cell_value(web_sheet, row_index, path_col)
                asset = path_to_host.get(access_path, "")
                if asset in target_set:
                    counts[asset]["webExposures"] += 1
                    counts[asset]["totalExposures"] += 1
                    if component_col is not None:
                        add_example(counts[asset]["webComponentExamples"], get_cell_value(web_sheet, row_index, component_col))

    if "非Web服务风险分布" in workbook.sheetnames:
        non_web_sheet = workbook["非Web服务风险分布"]
        col_map, header_row = build_col_map_allow_single(non_web_sheet)
        asset_col = find_column(col_map, ["IP地址/子域名", "IP地址", "资产IP", "Host", "host"])
        service_col = find_column(col_map, ["服务", "服务名称", "组件名称", "组件", "service"])
        if asset_col is not None:
            for row_index in range(header_row + 1, non_web_sheet.max_row + 1):
                row_values = [non_web_sheet.cell(row=row_index, column=col).value for col in range(1, non_web_sheet.max_column + 1)]
                if not any(normalize(cell) for cell in row_values):
                    continue
                asset = normalize_asset_key(get_cell_value(non_web_sheet, row_index, asset_col))
                if asset in target_set:
                    counts[asset]["nonWebExposures"] += 1
                    counts[asset]["totalExposures"] += 1
                    if service_col is not None:
                        add_example(counts[asset]["nonWebServiceExamples"], get_cell_value(non_web_sheet, row_index, service_col))

    workbook.close()
    return counts


def build_detail_lines(asset, incident_counts, vuln_counts, weakpwd_counts, exposure_counts, has_aes):
    total_events = int(incident_counts.get("totalEvents", 0))
    c2_events = int(incident_counts.get("c2Events", 0))
    virus_events = int(incident_counts.get("virusTrojanEvents", 0))
    exploit_events = int(incident_counts.get("vulnExploitEvents", 0))
    malware_and_c2_events = virus_events + c2_events
    other_events = max(total_events - malware_and_c2_events - exploit_events, 0)

    high_vulns = int(vuln_counts.get("highAndAboveVulnerabilities", 0))
    weak_passwords = int(weakpwd_counts.get("weakPasswords", 0))
    total_exposures = int(exposure_counts.get("totalExposures", 0))
    web_exposures = int(exposure_counts.get("webExposures", 0))
    non_web_exposures = int(exposure_counts.get("nonWebExposures", 0))
    non_web_examples = exposure_counts.get("nonWebServiceExamples") or []
    web_examples = exposure_counts.get("webComponentExamples") or []
    non_web_example_text = "、".join(non_web_examples[:2])
    web_example_text = "、".join(web_examples[:2])

    lines = []
    if total_events > 0:
        lines.append(
            f"总计安全事件{total_events}个，其中，"
            f"病毒木马&C2外联{malware_and_c2_events}起，"
            f"网站攻击&漏洞攻击{exploit_events}起，"
            f"其他事件{other_events}起"
        )
    if high_vulns > 0:
        lines.append(f"该资产共发现在{high_vulns}个高危及以上漏洞")
    if weak_passwords > 0:
        lines.append(f"该资产共发现{weak_passwords}个弱口令")
    if total_exposures > 0:
        lines.append(
            f"该资产共发现{total_exposures}个风险暴露面。"
            f"含{non_web_exposures}个非Web服务（如{non_web_example_text}等）"
            f"与{web_exposures}个Web服务（如{web_example_text}）"
        )
    lines.append(f"该资产{'已安装' if has_aes else '尚未安装'}aES")
    return lines


def main():
    if len(sys.argv) < 10:
        raise SystemExit(
            "Usage: top_risk_asset_details.py <incident.xlsx> <asset.xlsx> <weakpwd.xlsx> <vuln.xlsx> <exposure.xlsx> "
            "<top_assets_json> <c2_ids_json> <virus_ids_json> <exploit_ids_json>"
        )

    incident_path = sys.argv[1]
    asset_path = sys.argv[2]
    weakpwd_path = sys.argv[3]
    vuln_path = sys.argv[4]
    exposure_path = sys.argv[5]
    top_assets = json.loads(sys.argv[6])
    c2_ids = json.loads(sys.argv[7])
    virus_ids = json.loads(sys.argv[8])
    exploit_ids = json.loads(sys.argv[9])

    ordered_assets = []
    for item in top_assets:
        asset = normalize_asset_key(item.get("ip") if isinstance(item, dict) else item)
        if asset and asset not in ordered_assets:
            ordered_assets.append(asset)

    incident_counts = collect_incident_counts(incident_path, ordered_assets, c2_ids, virus_ids, exploit_ids)
    aes_status = collect_aes_install_status(asset_path, ordered_assets)
    vuln_counts = collect_vulnerability_counts(vuln_path, ordered_assets)
    weakpwd_counts = collect_weak_password_counts(weakpwd_path, ordered_assets)
    exposure_counts = collect_exposure_counts(exposure_path, ordered_assets)

    details = {}
    for asset in ordered_assets:
        counts = incident_counts.get(asset, {})
        asset_vuln_counts = vuln_counts.get(asset, {})
        asset_weakpwd_counts = weakpwd_counts.get(asset, {})
        asset_exposure_counts = exposure_counts.get(asset, {})
        has_aes = bool(aes_status.get(asset))
        details[asset] = {
            **counts,
            **asset_vuln_counts,
            **asset_weakpwd_counts,
            **asset_exposure_counts,
            "malwareAndC2Events": int(counts.get("virusTrojanEvents", 0)) + int(counts.get("c2Events", 0)),
            "otherEvents": max(
                int(counts.get("totalEvents", 0))
                - int(counts.get("virusTrojanEvents", 0))
                - int(counts.get("c2Events", 0))
                - int(counts.get("vulnExploitEvents", 0)),
                0,
            ),
            "hasAes": has_aes,
            "detailLines": build_detail_lines(asset, counts, asset_vuln_counts, asset_weakpwd_counts, asset_exposure_counts, has_aes),
        }

    print(json.dumps({"assets": details}, ensure_ascii=False))


if __name__ == "__main__":
    main()
