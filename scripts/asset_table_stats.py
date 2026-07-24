import json
import re
import sys

from openpyxl import load_workbook

from _path_helper import decode_argv
decode_argv()


def normalize(value):
    return "" if value is None else str(value).strip()


def normalize_header(value):
    return normalize(value).lower().replace(" ", "").replace("_", "")


def find_column(headers, aliases):
    normalized_aliases = [normalize_header(alias) for alias in aliases]
    for index, header in enumerate(headers):
        normalized = normalize_header(header)
        if not normalized:
            continue
        for alias in normalized_aliases:
            if alias in normalized:
                return index
    return None


def classify_asset_type(value):
    text = normalize(value)
    if "服务器" in text:
        return "server"
    if "终端" in text:
        return "terminal"
    return "other"


def parse_dev_types(datasource_text):
    """从数据源列文本提取设备类型集合（小写化、按顿号分隔、去括号后内容）。

    与 分支1/report/scoring.py 中 _parse_dev_types_from_datasource 保持一致。
    """
    if not datasource_text or not isinstance(datasource_text, str):
        return set()
    types = set()
    for seg in datasource_text.split("、"):
        seg = seg.strip()
        if not seg:
            continue
        match = re.match(r"([^(（]+)", seg)
        if match:
            name = match.group(1).strip().lower()
            if name:
                types.add(name)
        else:
            types.add(seg.lower())
    return types


# 未防护判定：数据源为空 OR dev_types 集合 ⊆ {"人工", "云镜-服务版"} 类别
# 即"人工"或"云镜-服务版"或两者组合，不含其他任何设备代号才算未防护。
# 注意："云镜"和"云镜-服务版"是两种不同的独立设备，不能混用子串匹配。
UNPROTECTED_TARGET = frozenset({"云镜-服务版"})
MANUAL_KEYWORD = "人工"
CLOUD_MIRROR_EXACT = "云镜-服务版"


def is_unprotected(datasource_text):
    if not datasource_text or not str(datasource_text).strip():
        return True
    types = parse_dev_types(str(datasource_text))
    if not types:
        return True
    # 每一个 type 必须命中"人工"关键字 或 等于"云镜-服务版"，否则算防护
    for t in types:
        if MANUAL_KEYWORD in t:
            continue
        if t == CLOUD_MIRROR_EXACT:
            continue
        return False
    return True


def classify_unprotected(datasource_text):
    """细分未防护资产的 4 类来源：
    - empty               : 数据源为空
    - manual              : 只含"人工"字样 token（人工录入 / 人工专家等）
    - cloud_mirror        : 只含"云镜-服务版"这一个 token
    - manual_and_cloud    : 同时含"人工"和"云镜-服务版"

    其它情况（含 AF/EDR/STA/云镜(非服务版) 等设备代号）返回 None（算防护）。

    注意：数据源里"云镜"和"云镜-服务版"是两种不同的独立设备，
    这里只把"云镜-服务版"算未防护，"云镜"独立设备按已纳管算防护。
    """
    if not datasource_text or not str(datasource_text).strip():
        return "empty"
    types = parse_dev_types(str(datasource_text))
    if not types:
        return "empty"
    # 任一 type 既不含"人工"也不等于"云镜-服务版" → 算防护
    for t in types:
        if MANUAL_KEYWORD not in t and t != CLOUD_MIRROR_EXACT:
            return None
    has_manual = any(MANUAL_KEYWORD in t for t in types)
    has_cloud = any(t == CLOUD_MIRROR_EXACT for t in types)
    if has_manual and has_cloud:
        return "manual_and_cloud"
    if has_manual:
        return "manual"
    if has_cloud:
        return "cloud_mirror"
    return None


CN_PROTECTION_STATUSES = {"在线", "离线", "已禁用", "已降级"}
CN_UNPROTECTED_STATUSES = {"未授权", "未安装", "已卸载", "已移除"}


def main():
    if len(sys.argv) < 2:
        raise SystemExit("Usage: asset_table_stats.py <asset.xlsx>")

    workbook = load_workbook(sys.argv[1], read_only=True, data_only=True)
    sheet = workbook.active
    # 自适应表头行：原始批次 xlsx 第 1 行空 + 第 2 行表头；
    # merge_xlsx_batches 合并后的 xlsx 第 1 行直接是表头。
    # 取前 2 行扫描，含 "IP 地址" 或 "审核状态" 等已知列名的就是表头行。
    candidate_rows = list(sheet.iter_rows(min_row=1, max_row=2, values_only=True))
    header_row = ()
    header_row_idx = 1
    known_aliases = ("ip地址", "资产类型", "审核状态", "agent状态", "托管状态", "数据源")
    for idx, row in enumerate(candidate_rows, start=1):
        normalized = [normalize_header(c) for c in row if c is not None]
        if any(any(alias in h for h in normalized) for alias in known_aliases):
            header_row = row
            header_row_idx = idx
            break
    if not header_row:
        # 兜底：默认 row 1
        header_row = candidate_rows[0] if candidate_rows else ()
        header_row_idx = 1
    headers = list(header_row)

    # 列匹配（仅指定的列名）
    type_column = find_column(headers, ("资产类型(一级)",))
    protection_column = find_column(headers, ("agent状态",))
    exposure_column = find_column(headers, ("互联网暴露",))
    importance_column = find_column(headers, ("重要级别",))
    managed_column = find_column(headers, ("托管状态",))
    datasource_column = find_column(headers, ("数据源",))
    approval_column = find_column(headers, ("审核状态",))

    asset_total = 0
    manage_asset = 0
    core_asset = 0
    core_managed_asset = 0
    type_counts = {"server": 0, "terminal": 0}
    protection_counts = {}  # 直接记录中文值
    protection_protected = 0
    protection_unprotected = 0
    # 未防护细分 4 子类（生成过程统计用，饼图仍展示两类）
    unprotected_empty = 0
    unprotected_manual = 0
    unprotected_cloud_mirror = 0
    unprotected_manual_and_cloud = 0
    exposure_total = 0
    exposure_type_counts = {"server": 0, "terminal": 0}
    wait_approve_count = 0  # 待审核资产数量
    current_asset_count = 0  # 资产台账（已审核）数量

    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not any(normalize(cell) for cell in row):
            continue

        asset_total += 1

        # ---- 资产类型（一级） ----
        if type_column is not None and type_column < len(row):
            asset_type = classify_asset_type(row[type_column])
            if asset_type in type_counts:
                type_counts[asset_type] += 1

        # ---- agent 状态：直接记录中文值（兼容旧逻辑，仍输出 protectionDistribution） ----
        if protection_column is not None and protection_column < len(row):
            raw = normalize(row[protection_column])
            if raw in CN_PROTECTION_STATUSES:
                protection_counts[raw] = protection_counts.get(raw, 0) + 1
            elif raw in CN_UNPROTECTED_STATUSES:
                protection_counts["未防护"] = protection_counts.get("未防护", 0) + 1

        # ---- 资产防护统计（按"数据源"列）：防护/未防护两类 + 未防护 4 子类 ----
        if datasource_column is not None and datasource_column < len(row):
            sub = classify_unprotected(row[datasource_column])
            if sub == "empty":
                protection_unprotected += 1
                unprotected_empty += 1
            elif sub == "manual":
                protection_unprotected += 1
                unprotected_manual += 1
            elif sub == "cloud_mirror":
                protection_unprotected += 1
                unprotected_cloud_mirror += 1
            elif sub == "manual_and_cloud":
                protection_unprotected += 1
                unprotected_manual_and_cloud += 1
            else:
                protection_protected += 1
        else:
            protection_protected += 1

        # ---- 互联网暴露：仅"暴露"算暴露 ----
        exposed = False
        if exposure_column is not None and exposure_column < len(row):
            raw = normalize(row[exposure_column])
            if raw == "暴露":
                exposed = True

        if exposed:
            exposure_total += 1
            if type_column is not None and type_column < len(row):
                asset_type = classify_asset_type(row[type_column])
                if asset_type in exposure_type_counts:
                    exposure_type_counts[asset_type] += 1

        # ---- 核心资产：根据"重要级别"列统计 ----
        is_core = False
        if importance_column is not None and importance_column < len(row):
            raw = normalize(row[importance_column])
            if raw and "核心" in raw:
                is_core = True
                core_asset += 1

        # ---- 核心已托管：核心资产中"托管状态"为"已托管" ----
        if is_core and managed_column is not None and managed_column < len(row):
            raw = normalize(row[managed_column])
            if "已托管" in raw:
                core_managed_asset += 1

        # ---- 已托管资产：托管状态为"已托管" ----
        if managed_column is not None and managed_column < len(row):
            raw = normalize(row[managed_column])
            if "已托管" in raw:
                manage_asset += 1

        # ---- 待审核资产：审核状态为"待审核" ----
        if approval_column is not None and approval_column < len(row):
            raw = normalize(row[approval_column])
            if raw == "待审核":
                wait_approve_count += 1
            elif raw == "已审核":
                current_asset_count += 1
        else:
            # 没有审核状态列时，所有资产都视为资产台账
            current_asset_count += 1

    # 未防护 4 子类过程日志（生成过程统计）
    print(
        f"[asset-table-stats] 未防护细分: 人工录入={unprotected_manual} "
        f"云镜-服务版={unprotected_cloud_mirror} "
        f"人工+云镜-服务版={unprotected_manual_and_cloud} "
        f"空数据={unprotected_empty} 合计={protection_unprotected}",
        file=sys.stderr
    )

    print(json.dumps({
        "assetTotal": asset_total,
        "currentAssetCount": current_asset_count,
        "manage_asset": manage_asset,
        "core_asset": core_asset,
        "core_managed_asset": core_managed_asset,
        "waitApproveAssetCount": wait_approve_count,
        "typeDistribution": {
            "server": type_counts["server"],
            "terminal": type_counts["terminal"],
            "other": max(asset_total - type_counts["server"] - type_counts["terminal"], 0)
        },
        "protectionDistribution": protection_counts,
        "protectionStats": {
            "protected": protection_protected,
            "unprotected": protection_unprotected,
            "unprotected_breakdown": {
                "manual": unprotected_manual,
                "cloud_mirror": unprotected_cloud_mirror,
                "manual_and_cloud_mirror": unprotected_manual_and_cloud,
                "empty": unprotected_empty
            }
        },
        "internetExposureTotal": exposure_total,
        "internetExposureDistribution": {
            "server": exposure_type_counts["server"],
            "terminal": exposure_type_counts["terminal"],
            "other": max(exposure_total - exposure_type_counts["server"] - exposure_type_counts["terminal"], 0)
        }
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
