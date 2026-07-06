#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从事件表 Excel 中提取已确认 C2 外联事件举例。

用法:
  extract_c2_connection_examples.py <incident.xlsx> <confirmed_ids_json>

输出 JSON:
{
  "c2Connections": [
    {
      "incidentId": "incident-001",
      "ioc": "1.1.1.2、www.bad.com",
      "affectedAsset": "10.10.10.8",
      "lastOccurredAt": "2026-06-28 10:22:31",
      "disposalStatus": "处置中"
    }
  ]
}
"""

import json
import re
import sys

from openpyxl import load_workbook

from _path_helper import decode_argv
decode_argv()


SEVERE_LABEL = "严重"
LEVEL_PRIORITY = {
    "严重": 4,
    "高危": 3,
    "中危": 2,
    "低危": 1,
}


def normalize(value):
    return "" if value is None else str(value).strip()


def build_col_map(ws):
    header = [normalize(cell) for cell in next(ws.iter_rows(values_only=True))]
    return {name: i for i, name in enumerate(header) if name}


def find_column(col_map, aliases):
    for alias in aliases:
        if alias in col_map:
            return col_map[alias]
    return None


def extract_severe_entities(raw):
    text = normalize(raw)
    if not text:
        return []

    matches = re.findall(r'([^，,、()（）]+?)\s*[（(]\s*([^()（）]+?)\s*[）)]', text)
    entities = []
    for entity, severity in matches:
        if normalize(severity) == SEVERE_LABEL:
            entities.append(normalize(entity))
    return entities


def main():
    if len(sys.argv) < 3:
        raise SystemExit("Usage: extract_c2_connection_examples.py <incident.xlsx> <confirmed_ids_json>")

    incident_path = sys.argv[1]
    confirmed_ids = json.loads(sys.argv[2])
    confirmed_set = set(str(item).strip() for item in confirmed_ids if item)

    workbook = load_workbook(incident_path, read_only=True, data_only=True)
    sheet = workbook.active
    col_map = build_col_map(sheet)

    id_col = find_column(col_map, ["事件ID", "incident_id"])
    ext_ip_col = find_column(col_map, ["外网IP", "外联IP"])
    domain_col = find_column(col_map, ["域名", "外联域名"])
    asset_col = find_column(col_map, ["受影响资产", "影响资产", "host_ip", "hostIp", "主机IP", "ip"])
    time_col = find_column(col_map, ["最近发生时间", "最近发现时间", "endTime", "结束时间"])
    status_col = find_column(col_map, ["处置状态", "dealStatus"])
    level_col = find_column(col_map, ["等级", "severity"])

    if id_col is None:
        raise SystemExit(f"事件表缺少事件ID列，可用列: {list(col_map.keys())}")

    rows = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        incident_id = normalize(row[id_col]) if len(row) > id_col else ""
        if not incident_id or incident_id not in confirmed_set:
            continue

        severe_entities = []
        if ext_ip_col is not None and len(row) > ext_ip_col:
            severe_entities.extend(extract_severe_entities(row[ext_ip_col]))
        if domain_col is not None and len(row) > domain_col:
            severe_entities.extend(extract_severe_entities(row[domain_col]))

        if not severe_entities:
            continue

        rows.append({
            "_levelPriority": LEVEL_PRIORITY.get(
                normalize(row[level_col]) if level_col is not None and len(row) > level_col else "",
                0
            ),
            "_rowIndex": row_index,
            "incidentId": incident_id,
            "ioc": "、".join(severe_entities),
            "affectedAsset": normalize(row[asset_col]) if asset_col is not None and len(row) > asset_col else "",
            "lastOccurredAt": normalize(row[time_col]) if time_col is not None and len(row) > time_col else "",
            "disposalStatus": normalize(row[status_col]) if status_col is not None and len(row) > status_col else ""
        })

    rows.sort(key=lambda item: (-item["_levelPriority"], item["_rowIndex"]))
    rows = rows[:5]
    for item in rows:
        item.pop("_levelPriority", None)
        item.pop("_rowIndex", None)

    workbook.close()
    print(json.dumps({"c2Connections": rows}, ensure_ascii=False))


if __name__ == "__main__":
    main()
