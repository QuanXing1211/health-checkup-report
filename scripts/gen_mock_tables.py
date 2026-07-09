#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
在原始 Excel 副本基础上，只替换数据行，保持列结构完全不变。
"""
import os
import uuid
from datetime import datetime, timedelta
from openpyxl import load_workbook

OUTPUT_DIR = r"C:\Users\User\Desktop\health-checkup-report\tmp"

# ======================== 数据模板 ========================
ASSET_IPS = [
    "10.248.38.136", "192.168.213.156", "192.168.76.20",
    "192.168.200.33", "192.168.200.22", "192.168.100.205",
    "192.168.100.200", "172.17.75.124", "168.196.23.21",
    "192.168.24.241", "10.245.224.255", "192.168.118.232",
    "192.168.254.90", "10.16.11.65", "10.18.20.115",
    "172.20.64.95", "192.168.50.10", "10.100.1.50",
    "172.16.88.30", "192.168.1.100",
]

MALICIOUS_IPS = [
    "45.33.32.156", "103.235.46.91", "185.220.101.34",
    "198.54.117.210", "91.121.87.45", "23.227.38.65",
    "5.188.206.78", "94.102.61.23", "185.153.196.30",
    "146.185.239.120", "195.123.237.91", "45.155.205.99",
]

MALICIOUS_DOMAINS = [
    "evil-c2.xyz", "malware-update.top", "phishing-panel.cc",
    "backdoor-gate.net", "trojan-cnc.info", "stealer-api.org",
    "ransom-bot.biz", "cobaltstrike-c2.xyz", "miner-pool.cc",
    "ddos-agent.net", "rat-controller.top", "keylogger-svr.cc",
]

MALICIOUS_FILES = [
    "svchost.exe", "updater.dll", "system32.dll", "winlogon.exe",
    "explorer.dll", "runtime.exe", "helper.sys", "servicehost.exe",
    "lsass.dll", "security.dll", "kernel32.exe", "browser.exe",
]

GPT_SUB_HOST = ["勒索攻击", "银狐攻击", "海莲花APT攻击", "响尾蛇APT攻击", "cobaltstrike远控", "摩诃草APT"]
GPT_SUB_VIRUS = ["银狐病毒", "勒索病毒", "挖矿病毒", "后门木马", "下载者木马", "蠕虫病毒"]

EVENT_CLASSES = [
    ("漏洞利用",    ["Web应用漏洞", "系统漏洞攻击", "远程代码执行", "文件上传漏洞", "SQL注入漏洞", "命令注入"]),
    ("恶意程序事件", ["病毒木马行为", "勒索软件活动", "挖矿木马", "后门程序", "蠕虫传播"]),
    ("渗透攻击事件", ["CobaltStrike", "WebShell上传", "端口扫描", "暴力破解", "横向移动"]),
    ("异常行为事件", ["异常外联行为", "异常登录行为", "数据窃取", "DNS隧道"]),
]

CUSTOMER = "测试001"
SEVERITY = ["严重", "高危", "中危", "低危"]
MONITOR_STATUS = ["已忽略", "已处理", "处理中"]
ATTACK_RESULTS = ["失陷", "攻击", "成功", "未成功"]
DISPOSAL_STATUS = ["待处理", "已遏制", "处置中", "已处置"]


def gen_id(): return f"incident-{uuid.uuid4()}"
def mal(label, val): return f"{val}（{label}）"
def rand_dt():
    s = datetime(2026, 6, 1) + timedelta(days=(hash(str(uuid.uuid4())) % 30))
    return s


def build_event_rows():
    rows = []

    # ====== 主机失陷活动 12条 ======
    for i in range(12):
        l1, l2_list = EVENT_CLASSES[i % len(EVENT_CLASSES)]
        l2 = l2_list[i % len(l2_list)]
        ip = MALICIOUS_IPS[i]
        domain = MALICIOUS_DOMAINS[i]
        gpt_sub = GPT_SUB_HOST[i % len(GPT_SUB_HOST)]
        asset_ip = ASSET_IPS[i]
        dt = rand_dt()

        ext_ip = mal("恶意", ip)
        domain_val = mal("恶意", domain) if i % 3 == 0 else "-"
        file_val = "-"   # 主机失陷不靠文件识别

        rows.append([
            gen_id(),
            f"检测到{gpt_sub}攻击活动—{l2}",
            asset_ip,
            "-",
            "主机失陷活动",
            gpt_sub,
            SEVERITY[i % 4],
            MONITOR_STATUS[i % 3],
            l1,
            l2,
            ATTACK_RESULTS[i % 4],
            dt.strftime("%Y-%m-%d %H:%M:%S"),
            (dt + timedelta(days=2)).strftime("%Y-%m-%d %H:%M:%S"),
            CUSTOMER,
            "-",
            "-",
            ip,
            f"STA (B{1000+i:05X}),EDR (3433000000{i+1:08d})",
            ext_ip,
            domain_val,
            file_val,
            DISPOSAL_STATUS[i % 4],
            (dt - timedelta(hours=2)).strftime("%Y-%m-%d %H:%M:%S"),
        ])

    # ====== 病毒木马活动 12条 ======
    for i in range(12):
        l1, l2_list = EVENT_CLASSES[(i + 1) % len(EVENT_CLASSES)]
        l2 = l2_list[i % len(l2_list)]
        fname = MALICIOUS_FILES[i]
        gpt_sub = GPT_SUB_VIRUS[i % len(GPT_SUB_VIRUS)]
        asset_ip = ASSET_IPS[(i + 6) % 20]
        dt = rand_dt()

        # 文件必须包含"恶意"标签 — 病毒木马靠文件识别
        file_val = mal("恶意", fname)
        # 外网IP/域名不应标"恶意"（病毒木马不靠这些识别）
        ext_ip = mal("其他", MALICIOUS_IPS[(i + 3) % 12]) if i % 4 == 0 else "-"
        domain_val = mal("其他", MALICIOUS_DOMAINS[(i + 5) % 12]) if i % 5 == 0 else "-"

        rows.append([
            gen_id(),
            f"检测到{gpt_sub}活动—{l2}",
            asset_ip,
            "-",
            "病毒木马活动",
            gpt_sub,
            SEVERITY[(i + 1) % 4],
            MONITOR_STATUS[(i + 1) % 3],
            l1,
            l2,
            ATTACK_RESULTS[(i + 2) % 4],
            dt.strftime("%Y-%m-%d %H:%M:%S"),
            (dt + timedelta(days=1)).strftime("%Y-%m-%d %H:%M:%S"),
            CUSTOMER,
            "-",
            "-",
            MALICIOUS_IPS[(i + 3) % 12] if i % 3 == 0 else "-",
            f"EDR (3433000000{i+100:08d}),STA (A{2000+i:04X})",
            ext_ip,
            domain_val,
            file_val,
            DISPOSAL_STATUS[(i + 2) % 4],
            (dt - timedelta(hours=1)).strftime("%Y-%m-%d %H:%M:%S"),
        ])

    return rows


def patch_event_table():
    """生成 mock 事件表。优先使用原始模板文件，不存在则从零创建。"""
    dst = os.path.join(OUTPUT_DIR, "mock_incident_table.xlsx")
    dst_v3 = os.path.join(OUTPUT_DIR, "mock_incident_table_v3.xlsx")

    # 检查目标文件是否已存在——存在则跳过，保持固定不变
    if os.path.exists(dst_v3):
        print(f"事件表已存在，跳过生成: {dst_v3}")
        return

    src = r"C:\Users\User\Downloads\测试001_事件跟踪表2026-07-09 10_32_02.xlsx"

    if os.path.exists(src):
        # 从原始模板读取结构
        wb = load_workbook(src)
        ws = wb.active
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
    else:
        # 原始模板不存在，从零创建工作簿
        print(f"原始模板不存在 ({src})，从零创建事件表")
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append([
            "事件ID", "事件名称", "资产IP", "资产名称", "事件类型",
            "事件子类型(GPT)", "严重级别", "监控状态", "事件分类(一级)",
            "事件分类(二级)", "攻击结果", "发现时间", "最近发生时间",
            "客户名称", "内网IP地址", "MAC地址", "外网IP地址",
            "发现源", "外网IP地址(标签)", "域名(标签)", "文件(标签)",
            "处置状态", "首次发现时间"
        ])

    for row_data in build_event_rows():
        ws.append(row_data)

    wb.save(dst_v3)
    print(f"事件表已生成: {dst_v3} ({ws.max_row - 1} 条)")
    # 同时保存无版本号副本
    wb.save(dst)
    print(f"事件表已生成: {dst} ({ws.max_row - 1} 条)")


def build_asset_rows():
    rows = []
    for i, ip in enumerate(ASSET_IPS):
        os_types = ["Windows Server 2019", "Linux CentOS 7", "Windows 10"]
        names = ["WEB", "API", "DB", "APP", "FS", "MAIL"]
        locations = ["机房A", "机房B", "机房C"]
        rows.append([
            ip,
            f"办公网络",
            f"业务系统{(i % 4) + 1}",
            "未知",
            f"{i:02X}:{3*i:02X}:{7*i:02X}:{65+i:02X}:{3*i:02X}:{99-i:02X}",
            f"203.{100+i}.50.{i+1}",                            # 公网IP — 每条都有
            os_types[i % 3],
            f"SRV-{names[i % 6]}-{i+1:02d}",
            "服务器" if i % 2 == 0 else "终端",
            "核心" if i % 6 == 0 else "重要",
            "在线" if i % 5 != 0 else "离线",
            "暴露" if i % 4 == 2 else "未暴露",
            "核心" if i % 6 == 0 else "重要",
            f"{['WEB服务器','API服务器','数据库服务器','应用服务器','文件服务器','邮件服务器','终端PC','测试服务器','备份服务器','DNS服务器','监控服务器','日志服务器','开发PC','运维PC','堡垒机','VPN设备','防火墙','交换机','路由器','打印机'][i]}",
            f"{locations[i % 3]}-{i+1:02d}号机柜",
            f"标签{i+1}",                                       # 资产标签 — 每条都有
            "EDR" if i % 3 != 2 else "STA",                     # 发现源 — 每条都有
            f"负责人{(i % 5) + 1}",                              # 负责人 — 每条都有
            f"1380000{i:04d}",                                   # 负责人电话 — 每条都有
            f"host-srv-{i:03d}",
            f"user_{i:03d}",                                     # 实时认证用户名 — 每条都有
            "已托管" if i % 3 != 0 else "未托管",                  # 托管状态
        ])
    return rows


def patch_asset_table():
    """生成 mock 资产表。已生成过的文件不会再覆盖，保持固定不变。"""
    dst = os.path.join(OUTPUT_DIR, "mock_asset_table.xlsx")

    # 检查目标文件是否已存在——存在则跳过，保持固定不变
    if os.path.exists(dst):
        print(f"资产表已存在，跳过生成: {dst}")
        return

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    # 表头放第1行（business_system_ranking.py 的 build_col_map 取第一个非空行当表头，
    # 第1行放占位会让它误把“资产清单”当表头，导致 IP地址/所属业务/重要级别/托管状态 全部找不到）
    ws.append([
        "IP地址", "网络区域", "所属业务", "资产名称", "MAC地址",
        "公网IP地址", "操作系统", "主机名", "资产类型(一级)", "重要级别",
        "agent状态", "互联网暴露", "资产等级", "资产类型(二级)",
        "物理位置", "资产标签", "发现源", "负责人", "负责人电话",
        "主机名(FQDN)", "实时认证用户名", "托管状态"
    ])

    for row_data in build_asset_rows():
        ws.append(row_data)

    wb.save(dst)
    print(f"资产表已生成: {dst} ({ws.max_row - 1} 条)")


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    patch_event_table()
    patch_asset_table()

    # 验证（使用 v3 文件名）
    incident_path = os.path.join(OUTPUT_DIR, "mock_incident_table_v3.xlsx")
    if os.path.exists(incident_path):
        wb = load_workbook(incident_path)
        ws = wb.active
        host = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if str(r[4] or '') == "主机失陷活动")
        virus = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if str(r[4] or '') == "病毒木马活动")
        print(f"统计: 主机失陷 {host}, 病毒木马 {virus}, 总计 {ws.max_row - 1}")
    else:
        print(f"事件表未找到: {incident_path}")


if __name__ == "__main__":
    main()
