import json
import sys
from collections import Counter

from openpyxl import load_workbook

from _path_helper import decode_argv
decode_argv()

# 确保 stdout 使用 UTF-8 编码（Windows 兼容）
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')


def normalize(value):
    return "" if value is None else str(value).strip()


def normalize_header(value):
    return normalize(value).lower().replace(" ", "").replace("_", "")


def find_column(headers, aliases):
    for index, header in enumerate(headers):
        normalized = normalize_header(header)
        if not normalized:
            continue
        for alias in aliases:
            if alias in normalized:
                return index
    return None


def top1_name(counter):
    """取 Counter 第1项的名称, 无可返回空字符串"""
    items = counter.most_common(1)
    return items[0][0] if items else ""


def top5_list(counter):
    """取 Counter 前5项, 返回 [{name, value}] 格式"""
    return [{"name": k, "value": v} for k, v in counter.most_common(5)]


def main():
    if len(sys.argv) < 2:
        raise SystemExit("Usage: incident_type_stats.py <incident.xlsx>")

    incident_path = sys.argv[1]

    try:
        wb = load_workbook(incident_path, data_only=True)
        ws = wb.active

        header_row = next(ws.iter_rows(values_only=True))
        headers = [normalize(h) for h in header_row]

        type_col = find_column(headers, ["安全事件一级分类"])

        counter = Counter()

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(normalize(cell) for cell in row):
                continue
            if type_col is not None and type_col < len(row):
                val = normalize(row[type_col])
                if val:
                    counter[val] += 1

        wb.close()

        result = {
            "topEventType": top1_name(counter),
            "eventTypeDistribution": top5_list(counter)
        }
        print(json.dumps(result, ensure_ascii=False))

    except Exception as e:
        print(json.dumps({
            "topEventType": "",
            "eventTypeDistribution": [],
            "error": f"读取事件表失败: {str(e)}"
        }, ensure_ascii=False))


if __name__ == "__main__":
    main()
