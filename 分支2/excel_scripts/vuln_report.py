#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
漏洞清单生成工具
从 EASM 平台（外网）和 MSSW 平台（内网）分别获取漏洞数据，
合并生成统一的「漏洞清单.xlsx」。

用法: python vuln_report.py <客户ID或客户名称关键词>
示例: python vuln_report.py 深圳市口袋网络
      python vuln_report.py 35690473
"""

import os
import sys
import time
import threading
import argparse
import requests
from copy import copy
from openpyxl.styles import Font, PatternFill, Border, Alignment
from datetime import datetime, timezone, timedelta
from typing import Dict, List, Optional, Tuple
import openpyxl
from openpyxl import load_workbook


# ==================== 配置（按需修改） ====================

# --- 通用 ---
TEMP_DIR     = r"C:\Users\User\Downloads\temp_report"
OUTPUT_FILE  = r"C:\Users\User\Downloads\漏洞清单.xlsx"

POLL_INTERVAL   = 5     # 轮询间隔（秒）
SCRIPT_TIMEOUT  = 3600  # 全局超时：1小时
MAX_RETRIES     = 3     # 最大重试次数
RETRY_DELAY     = 3     # 重试等待时间（秒）
PAGE_LIMIT      = 100   # 列表接口每次查询数量

# --- EASM 平台（内网） ---
EASM_BASE_URL    = "https://soar59.sangfor.com.cn"
EASM_COOKIES_FILE = r"C:\Users\User\Downloads\cookies.txt"
EASM_VULN_STATUSES = [5, 1, 12, 9]                  # 过滤跟进状态：[未审核=5, 处置中=1, 待复测=12, 修复失败=9]

# --- MSSW 平台（内网） ---
MSSW_BASE_URL      = "https://pre.soar.sangfor.com"
MSSW_COOKIES_FILE   = r"C:\Users\User\Downloads\mssw_cookies.txt"

MSSW_FIXED_STATUSES = [0, 1]                        # 过滤处置状态：[待处置=0, 处置中=1]，排除已闭环和误报

# ==========================================================


# ---------- 日志 ----------

def log(msg: str, level: str = "INFO") -> None:
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}][{level}] {msg}")


# ---------- 工具函数 ----------

def _get_system_timezone() -> str:
    offset = datetime.now().astimezone().utcoffset()
    if offset is None:
        return "+00:00"
    total_seconds = int(offset.total_seconds())
    sign = "+" if total_seconds >= 0 else "-"
    total_seconds = abs(total_seconds)
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    return f"{sign}{hours:02d}:{minutes:02d}"


def extract_cookie_value(cookie_str: str, name: str) -> Optional[str]:
    for part in cookie_str.split(';'):
        part = part.strip()
        if '=' in part:
            k, _, v = part.partition('=')
            if k.strip() == name:
                return v.strip()
    return None


def _fmt_minute(val) -> str:
    """
    将时间值统一格式化到分钟精度（YYYY-MM-DD HH:MM）。
    支持 datetime 对象和字符串，空值返回空字符串。
    """
    if val is None or val == "":
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M")
    s = str(val).strip()
    return s[:16] if len(s) >= 16 else s


# ---------- Cookie 读取 ----------

def read_cookies_as_string(filepath: str) -> str:
    """
    读取 cookies.txt，返回 Cookie 请求头字符串（name=value; name2=value2 格式）。
    支持两种输入格式：
      1. 浏览器原始 Cookie 字符串
      2. Netscape 格式（EditThisCookie 等工具导出，tab 分隔）
    """
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read().strip()

    lines = [ln.strip() for ln in content.splitlines() if ln.strip()]
    is_netscape = any('\t' in ln for ln in lines if not ln.startswith('#'))

    if is_netscape:
        pairs = []
        for ln in lines:
            if ln.startswith('#'):
                continue
            parts = ln.split('\t')
            if len(parts) >= 7:
                pairs.append(f"{parts[5]}={parts[6]}")
        return '; '.join(pairs)
    else:
        return '; '.join(lines)


# ---------- 请求封装 ----------

def _build_default_headers(base_url: str) -> Dict[str, str]:
    return {
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Content-Type": "application/json",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": f"{base_url}/index.html",
        "timezone": _get_system_timezone(),
    }


def request_with_retry(method: str, url: str, base_url: str,
                       cookie_str: str = "",
                       timeout: int = 60,
                       extra_headers: Optional[Dict] = None,
                       **kwargs) -> Optional[requests.Response]:
    headers = _build_default_headers(base_url)
    if extra_headers:
        headers.update(extra_headers)
    if cookie_str:
        headers["Cookie"] = cookie_str
        csrf_token = extract_cookie_value(cookie_str, "csrf_token")
        if csrf_token:
            headers["X-Csrftoken"] = csrf_token
        # MSSW 平台的 CSRF token（cookie名和header名都是 x-csrf-token）
        mssw_csrf = extract_cookie_value(cookie_str, "x-csrf-token")
        if mssw_csrf:
            headers["x-csrf-token"] = mssw_csrf

    session = requests.Session()
    adapter = requests.adapters.HTTPAdapter(
        pool_connections=10, pool_maxsize=10, max_retries=0
    )
    session.mount('https://', adapter)
    session.mount('http://', adapter)

    for attempt in range(MAX_RETRIES):
        try:
            if method.upper() == "POST":
                resp = session.post(url, headers=headers, timeout=timeout,
                                    verify=False, **kwargs)
            else:
                resp = session.get(url, headers=headers, timeout=timeout,
                                   verify=False, **kwargs)
            return resp
        except (requests.exceptions.Timeout, requests.exceptions.ConnectionError,
                requests.exceptions.SSLError, requests.exceptions.RequestException) as e:
            attempt_num = attempt + 1
            log(f"请求失败 (尝试 {attempt_num}/{MAX_RETRIES}) - {e}", "WARNING")
            if attempt < MAX_RETRIES - 1:
                time.sleep(RETRY_DELAY)
            else:
                log(f"重试 {MAX_RETRIES} 次后仍然失败，返回 None", "ERROR")
                return None
    return None


def _parse_json(resp: requests.Response, api_name: str) -> dict:
    if resp is None:
        raise RuntimeError(f"{api_name}：请求失败（已重试 {MAX_RETRIES} 次）")
    try:
        return resp.json()
    except Exception:
        raise RuntimeError(
            f"{api_name}：响应不是JSON（状态码={resp.status_code}，可能Cookie已过期）"
        )


# ---------- 风险等级归一化 ----------

RISK_NORMALIZE = {
    "超危": "严重",
    "高危": "高",
    "中危": "中",
    "低危": "低",
}


def normalize_risk(level: str) -> str:
    """将各平台的风险等级统一为映射表使用的格式"""
    return RISK_NORMALIZE.get(level, level)


# ---------- 修复优先级计算 ----------

def calc_priority(risk_norm: str, is_internal: bool, is_exploitable: bool) -> str:
    """
    根据映射表计算修复优先级。
    risk_norm: 归一化后的风险等级（严重/高/中/低）
    is_internal: True=内网, False=外网
    is_exploitable: True=可利用, False=不可利用
    """
    inout = "内网" if is_internal else "外网"
    exp   = "可利用" if is_exploitable else "不可利用"

    TABLE: Dict[Tuple[str, str, str], str] = {
        ("严重", "内网", "可利用"):    "急需修复",
        ("严重", "内网", "不可利用"):  "尽快修复",
        ("严重", "外网", "可利用"):    "急需修复",
        ("严重", "外网", "不可利用"):  "急需修复",
        ("高",   "内网", "可利用"):    "急需修复",
        ("高",   "内网", "不可利用"):  "尽快修复",
        ("高",   "外网", "可利用"):    "急需修复",
        ("高",   "外网", "不可利用"):  "急需修复",
        ("中",   "内网", "可利用"):    "急需修复",
        ("中",   "内网", "不可利用"):  "尽快修复",
        ("中",   "外网", "可利用"):    "急需修复",
        ("中",   "外网", "不可利用"):  "尽快修复",
        ("低",   "内网", "可利用"):    "建议修复",
        ("低",   "内网", "不可利用"):  "建议修复",
        ("低",   "外网", "可利用"):    "建议修复",
        ("低",   "外网", "不可利用"):  "建议修复",
    }
    return TABLE.get((risk_norm, inout, exp), "建议修复")


# ---------- 通用：搜索客户结果匹配 ----------

def _pick_exact_match(customers: list, keyword: str):
    """从多个模糊搜索结果中优先选精确匹配。返回匹配项或 None"""
    if len(customers) == 1:
        return customers[0]
    exact = [c for c in customers
             if (c.get('company_name', '') or '').strip() == keyword.strip()
             or (c.get('pms_customer_name', '') or '').strip() == keyword.strip()
             or str(c.get('company_id', '')).strip() == keyword.strip()]
    return exact[0] if exact else None


def parse_date_to_ms(date_str: str, is_end: bool = False) -> int:
    """
    将日期字符串转为13位 UTC+8 毫秒时间戳。
    支持格式: '2016-01-01' 或 '2016年1月1日'
    is_end=False → 当天 00:00:00.000
    is_end=True  → 当天 23:59:59.999
    """
    date_str = date_str.replace('年', '-').replace('月', '-').replace('日', '').strip()
    dt = datetime.strptime(date_str[:10], "%Y-%m-%d")
    dt = dt.replace(tzinfo=timezone(timedelta(hours=8)))
    if is_end:
        return int(dt.timestamp() * 1000) + 86399999
    return int(dt.timestamp() * 1000)


# ---------- 通用：接口0 搜索客户 ----------

def api0_search_customer(cookie_str: str, keyword: str) -> list:
    """根据关键词模糊搜索客户，返回列表"""
    url = f"{EASM_BASE_URL}/gateway/customer-mgr-service/order/v1/user?_method=GET"
    # 正确入参
    payload = {
        "order": "asc", "offset": 0, "limit": 20, "keyword": keyword,
        "share_ids": [], "delivery_channel_id": [], "service_code": [],
        "industry": [], "industry_segmentation": [], "customer_type": [],
        "customer_stratification": [], "protection_type": [], "service_group": [],
        "delivery_method": [], "platform_type": [], "service_status": 0, "my_customer": 0,
    }
    # 35环境
    # payload = {
    #     "order": "asc", "offset": 0, "limit": 10, "keyword": keyword,
    #     "share_ids": [], "delivery_channel_id": [], "phase": [],
    #     "industry": [], "industry_segmentation": [], "auth_status": 2,
    #     "customer_type": [], "customer_stratification": [],
    #     "protection_type": [], "delivery_method": [], "platform_type": [],
    #     "operate_status": [], "service_status": 0, "my_customer": 0,
    # }
    resp = request_with_retry("POST", url, EASM_BASE_URL, cookie_str, json=payload, timeout=120)
    data = _parse_json(resp, "接口0")
    if data.get('code') != 0:
        raise RuntimeError(f"接口0失败: {data.get('msg')}")
    return data['data']['list']


# ---------- MSSW：接口8-0 搜索客户 ----------

def mssw_api0_search_customer(cookie_str: str, keyword: str) -> list:
    """接口8-0：MSSW 平台客户搜索，根据关键词（名称或ID）模糊搜索"""
    url = f"{MSSW_BASE_URL}/gateway/customer-mgr-service/order/v1/user?_method=GET"
    key_field = "company_id_keyword" if keyword.isdigit() else "company_name_keyword"
    payload = {
        "my_customer": 0,
        key_field: keyword,
        "offset": 0, "limit": 20,
    }
    resp = request_with_retry("POST", url, MSSW_BASE_URL, cookie_str, json=payload, timeout=120)
    data = _parse_json(resp, "接口8-0（MSSW客户搜索）")
    if data.get('code') != 0:
        raise RuntimeError(f"接口8-0失败: {data.get('msg')}")
    return data['data']['list']


# ====================================================================
#  EASM 平台（外网）接口调用
# ====================================================================

def easm_api5_export_vuln(cookie_str: str, company_id: str,
                           scan_method: int,
                           last_time: list) -> str:
    """接口5：触发漏洞导出，返回 task_id"""
    url = f"{EASM_BASE_URL}/gateway/vuln-manager/vm/order/v1/vulnmgr/exposed_surface/report"
    payload = {
        "params": {
            "order": {},
            "offset": 0,
            "limit": 10,
            "keyword": "",
            "target_company_id": [],
            "task_id": [],
            "vt_type": [],
            "scan_method": scan_method,
            "found_time": [],
            "last_time": last_time,
            "verified_time": [],
            "closed_loop_time": [],
            "vulnerability_level": -1,
            "vulnerability_status": EASM_VULN_STATUSES,
            "vuln_class": [],
            "company_id": company_id,
            "data_id_list": [],
        },
        "need_split": False,
        "data_type": 13,
    }
    resp = request_with_retry("POST", url, EASM_BASE_URL, cookie_str, json=payload, timeout=120)
    data = _parse_json(resp, "接口5（EASM漏洞导出）")
    if data.get('code') != 0:
        raise RuntimeError(f"接口5失败: {data.get('msg')}")
    task_id = data['data']['task_id']
    log(f"EASM 导出任务 task_id={task_id}")
    return task_id


def easm_api6_poll_export(cookie_str: str, task_id: str) -> str:
    """接口6：轮询报告生成状态，返回下载 url（相对路径）"""
    url = f"{EASM_BASE_URL}/gateway/vuln-manager/vm/order/v1/vulnmgr/exposed_surface/report_async_task"
    payload = {"task_id_list": [task_id]}

    attempt = 0
    while True:
        attempt += 1
        resp = request_with_retry("POST", url, EASM_BASE_URL, cookie_str, json=payload, timeout=120)
        data = _parse_json(resp, "接口6（EASM轮询）")
        if data.get('code') != 0:
            raise RuntimeError(f"接口6失败: {data.get('msg')}")

        rotation_status = data.get('data', {}).get('rotation_status')
        attachment_list = data.get('data', {}).get('attachment_list', [])
        log(f"  attachment_list={attachment_list}")
        log(f"  attachment_list status={attachment_list[0].get('status')}")
        log(f"  rotation_status={rotation_status}")

        log(f"  第{attempt}次轮询: rotation_status={rotation_status}")

        if rotation_status == 2:
            if attachment_list and attachment_list[0].get('status') == 'success':
                download_path = attachment_list[0].get('url', '')
                log(f"  导出完成，下载路径: {download_path}")
                return download_path
            else:
                raise RuntimeError(f"接口6：导出任务状态为结束但附件为空或状态异常: {attachment_list}")

        # rotation_status=1 或其它 → 继续等待
        time.sleep(POLL_INTERVAL)


def easm_download_file(cookie_str: str, download_path: str, save_dir: str) -> str:
    """下载 EASM 导出的漏洞文件，返回本地文件路径"""
    url = f"{EASM_BASE_URL}{download_path}"
    resp = request_with_retry("GET", url, EASM_BASE_URL, cookie_str, stream=True)
    if resp is None:
        raise RuntimeError("EASM 漏洞文件下载失败")

    content_disp = resp.headers.get('Content-Disposition', '')
    if 'filename=' in content_disp:
        filename = content_disp.split('filename=')[-1].strip().strip('"\'')
    else:
        filename = f"easm_vuln_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

    filepath = os.path.join(save_dir, filename)
    total = 0
    with open(filepath, 'wb') as f:
        for chunk in resp.iter_content(chunk_size=8192):
            if chunk:
                f.write(chunk)
                total += len(chunk)

    log(f"EASM 漏洞文件已保存: {os.path.basename(filepath)} ({total / 1024:.1f} KB)")
    return filepath


# ====================================================================
#  MSSW 平台（内网）接口调用
# ====================================================================

def mssw_api8_export_vuln(cookie_str: str, company_id: str, latest_time_range: list) -> str:
    """接口8：触发 MSSW 漏洞导出，返回 file_name"""
    url = f"{MSSW_BASE_URL}/order/v1/vul_manage/vul_risk_export"
    payload = {
        "latest_time_range": latest_time_range,
        "asset_ip": {"op": "=", "val": ""},
        "asset_manager": {"op": "=", "val": ""},
        "asset_status": [],
        "asset_tags": [],
        "asset_type": "all",
        "attack_state": [],
        "branch_ids": [],
        "disposal_tag": [],
        "exposure": [],
        "fix_priority": [],
        "fixed_status": MSSW_FIXED_STATUSES,
        "group_ids": [],
        "keyword": "",
        "keyword_all": "",
        "magnitude": [],
        "name": {"op": "=", "val": ""},
        "order_status": [],
        "platform_ids": [],
        "platform_filter": [],
        "retest_status": [],
        "source_device": [],
        "attack_type": "",
        "cve": {"op": "=", "val": ""},
        "data_type": ["loophole"],
        "risk_level": [],
        "scan_type": [],
        "threat_tag": [],
        "custom_headers": {
            "asset_info": [
                {"disabled": True,  "key": "asset",             "label": "风险资产",      "selected": True},
                {"disabled": False, "key": "asset_type",        "label": "资产类型",      "selected": True},
                {"disabled": False, "key": "business_name",     "label": "所属资产组",    "selected": True},
                {"disabled": False, "key": "group_name",        "label": "所属业务",      "selected": True},
                {"disabled": False, "key": "manager",           "label": "资产责任人",    "selected": True},
                {"disabled": False, "key": "magnitude",         "label": "资产重要性",    "selected": True},
                {"disabled": True,  "key": "port",              "label": "端口",          "selected": True},
                {"disabled": True,  "key": "url",               "label": "url",           "selected": True},
                {"disabled": False, "key": "exposure",          "label": "互联网暴露",    "selected": True},
                {"disabled": True,  "key": "evidence_information", "label": "举证信息",  "selected": True},
                {"disabled": False, "key": "asset_status",      "label": "资产管理状态",  "selected": True},
                {"disabled": False, "key": "platform_name",     "label": "来源平台",      "selected": True},
                {"disabled": False, "key": "managed_level",   "label": "托管状态",      "selected": True},
            ],
            "base_info": [
                {"disabled": True,  "key": "name",               "label": "漏洞名称",      "selected": True},
                {"disabled": True,  "key": "fix_priority_level", "label": "修复优先级",    "selected": True},
                {"disabled": True,  "key": "risk_Level",         "label": "风险等级",      "selected": True},
                {"disabled": False, "key": "attack_type",        "label": "漏洞类型",      "selected": True},
                {"disabled": False, "key": "fix_advise",         "label": "修复建议",      "selected": True},
                {"disabled": False, "key": "risk_description",   "label": "风险描述",      "selected": True},
                {"disabled": False, "key": "threat_tags",        "label": "威胁标签",      "selected": True},
                {"disabled": False, "key": "src_type",           "label": "数据源",        "selected": True},
                {"disabled": False, "key": "scan_type",          "label": "检测方式",      "selected": True},
                {"disabled": False, "key": "last_time",          "label": "最近发现时间",  "selected": True},
                {"disabled": False, "key": "found_time",         "label": "首次发现时间",  "selected": True},
                {"disabled": False, "key": "cve",                "label": "CVE 编号",      "selected": True},
                {"disabled": False, "key": "attack_state",       "label": "攻击结果",      "selected": True},
                {"disabled": False, "key": "is_gpt",             "label": "GPT检测",       "selected": True},
            ],
            "disposal_info": [
                {"disabled": False, "key": "fixed_status",  "label": "处置状态",        "selected": True},
                {"disabled": False, "key": "fixed_tag",     "label": "处置标签",        "selected": True},
                {"disabled": False, "key": "order_progress","label": "最新工单进展",    "selected": True},
                {"disabled": False, "key": "retest_status", "label": "验证状态",        "selected": True},
            ],
        },
        "header_id": "loophole_1",
        "is_all": False,
        "multiple_choice": [],
        "exclude_multiple_choice": [],
    }
    extra_hdrs = {"X-MSSW-Company-Id": company_id} if company_id else None
    resp = request_with_retry("POST", url, MSSW_BASE_URL, cookie_str,
                               json=payload, extra_headers=extra_hdrs, timeout=660)
    data = _parse_json(resp, "接口8（MSSW漏洞导出）")
    if data.get('code') != 0:
        raise RuntimeError(f"接口8失败: {data.get('msg')}")
    file_name = data['data']['file_name']
    log(f"MSSW 导出文件名: {file_name}")
    return file_name


def mssw_api10_download(cookie_str: str, company_id: str, file_name: str, save_dir: str) -> str:
    """接口10：下载 MSSW 导出的漏洞文件，返回本地文件路径"""
    url = f"{MSSW_BASE_URL}/order/v1/vul_manage/download_file?file={file_name}"
    extra_hdrs = {"X-MSSW-Company-Id": company_id} if company_id else None
    resp = request_with_retry("GET", url, MSSW_BASE_URL, cookie_str, stream=True, extra_headers=extra_hdrs)
    if resp is None:
        raise RuntimeError("MSSW 漏洞文件下载失败")

    filepath = os.path.join(save_dir, file_name)
    total = 0
    with open(filepath, 'wb') as f:
        for chunk in resp.iter_content(chunk_size=8192):
            if chunk:
                f.write(chunk)
                total += len(chunk)

    # 检查是否为有效xlsx（xlsx本质是zip，魔数为PK）
    with open(filepath, 'rb') as f:
        magic = f.read(4)
    if magic != b'PK\x03\x04':
        # 不是zip/xlsx，尝试读取内容打印
        with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
            content = f.read(500)
        raise RuntimeError(f"MSSW 下载文件不是有效xlsx（{total}字节），服务器响应: {content}")

    log(f"MSSW 漏洞文件已保存: {os.path.basename(filepath)} ({total / 1024:.1f} KB)")
    return filepath


# ====================================================================
#  Excel 处理
# ====================================================================


# ---------- EASM 文件A 列名映射 ----------

# 文件A 的预期列名
EASM_COLUMNS = [
    "序号", "IP/子域名", "端口", "漏洞名称", "漏洞URL", "漏洞类型",
    "风险等级", "扫描方式", "引擎来源", "CVE ID", "组件",
    "高可利用漏洞", "关联核查项", "举证信息（请求）", "举证信息（响应）",
    "目标单位", "责任单位/部门", "关联任务", "首次发现时间", "最近发现时间",
    "审核时间", "闭环时间", "备注", "跟进状态",
    "漏洞介绍",  # EASM 导出文件已包含此列
]

# 文件C 需要的字段 ← EASM 文件A 取值规则
# (文件C列名, 来源A列名 或 固定值/计算逻辑)
C_FROM_EASM: List[Tuple[str, str]] = [
    ("漏洞名称",      "漏洞名称"),
    ("修复优先级",    None),          # 由 calc_priority 计算
    ("风险等级",      "风险等级"),
    ("漏洞类型",      "漏洞类型"),
    ("修复建议",      "__EMPTY__"),   # 固定为空
    ("风险描述",      "漏洞介绍"),     # 直接来自文件A的【漏洞介绍】列
    ("威胁标签",      "__EMPTY__"),   # 固定为空
    ("数据源",        "__FIXED_bjx"), # 固定为 "bjx"
    ("检测方式",      "扫描方式"),
    ("CVE 编号",      "__FIXED_CVE_ID"), # 固定为 "CVE ID"
    ("最近发现时间",  "最近发现时间"),
    ("首次发现时间",  "首次发现时间"),
    ("风险资产",      "IP/子域名"),
    ("资产类型",      "__EMPTY__"),   # 固定为空
    ("所属资产组",    "__EMPTY__"),
    ("所属业务",      "__EMPTY__"),
    ("资产责任人",    "__EMPTY__"),
    ("资产重要性",    "__EMPTY__"),
    ("资产管理状态",  "__EMPTY__"),
    ("端口",          "端口"),
    ("url",           "漏洞URL"),
    ("互联网暴露",    "__FIXED_暴露"), # 固定为 "暴露"
    ("举证信息",      "__PROOF__"),   # 举证信息（请求）+ \n + 举证信息（响应）
    ("处置状态",      "跟进状态"),
    ("数据来源",      "__FIXED_外网"), # 固定为 "外网"
]


def process_easm_file(file_a_path: str) -> List[dict]:
    """
    处理 EASM 导出的漏洞文件 A：
    - 添加/计算缺失字段（修复优先级、举证信息等）
    - 映射为文件C 的统一格式
    返回 list[dict]，每个 dict 是文件C 的一行
    """
    wb = load_workbook(file_a_path, data_only=True)
    ws = wb.active

    # 构建列名 → 列索引 映射
    col_map: Dict[str, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value is not None:
            col_map[str(cell.value).strip()] = col_idx

    log(f"EASM 文件A 列数: {len(col_map)}, 数据行数: {ws.max_row - 1}")

    # 检查预期列是否存在，缺失时仅打警告、不中断（取值时会赋空）
    expected = ["漏洞名称", "IP/子域名", "端口", "风险等级", "高可利用漏洞",
                "漏洞介绍", "漏洞类型", "扫描方式", "CVE ID",
                "最近发现时间", "首次发现时间", "漏洞URL", "跟进状态",
                "举证信息（请求）", "举证信息（响应）"]
    missing = [k for k in expected if k not in col_map]
    if missing:
        log(f"  文件A 缺少以下列，对应字段将填空: {missing}", "WARNING")

    rows_out: List[dict] = []

    for row_idx in range(2, ws.max_row + 1):
        row_data = {}

        for c_col, source in C_FROM_EASM:
            if source is None:
                # 修复优先级：需计算
                risk_raw = ws.cell(row=row_idx, column=col_map["风险等级"]).value
                risk_norm = normalize_risk(str(risk_raw or ""))
                high_exploit = str(ws.cell(row=row_idx, column=col_map["高可利用漏洞"]).value or "")
                is_exploitable = (high_exploit == "是")
                # EASM 固定为外网
                row_data[c_col] = calc_priority(risk_norm, is_internal=False, is_exploitable=is_exploitable)

            elif source == "__EMPTY__":
                row_data[c_col] = ""

            elif source == "__FIXED_bjx":
                row_data[c_col] = "bjx"

            elif source == "__FIXED_暴露":
                row_data[c_col] = "暴露"

            elif source == "__FIXED_外网":
                row_data[c_col] = "外网"

            elif source == "__FIXED_CVE_ID":
                row_data[c_col] = "CVE ID"

            elif source == "__PROOF__":
                # 举证信息：请求 + 响应 拼接
                req_col = col_map.get("举证信息（请求）")
                rsp_col = col_map.get("举证信息（响应）")
                req = (ws.cell(row=row_idx, column=req_col).value or ""
                       if req_col else "")
                rsp = (ws.cell(row=row_idx, column=rsp_col).value or ""
                       if rsp_col else "")
                proof_parts = [p for p in [str(req), str(rsp)] if p.strip()]
                row_data[c_col] = "\n".join(proof_parts)

            else:
                # 直接从 Excel 对应列取值
                src_col = col_map.get(source)
                if src_col is not None:
                    val = ws.cell(row=row_idx, column=src_col).value
                    val = val if val is not None else ""
                    if c_col == "风险等级" and str(val) == "超危":
                        val = "严重"
                    row_data[c_col] = val
                else:
                    row_data[c_col] = ""

        rows_out.append(row_data)

    wb.close()
    log(f"EASM 处理后共 {len(rows_out)} 行")
    return rows_out


# ---------- MSSW 文件B 处理 ----------

C_FROM_MSSW: List[Tuple[str, str]] = [
    ("漏洞名称",      "漏洞名称"),
    ("修复优先级",    None),          # 由 calc_priority 计算
    ("风险等级",      "风险等级"),
    ("漏洞类型",      "漏洞类型"),
    ("修复建议",      "修复建议"),
    ("风险描述",      "风险描述"),
    ("威胁标签",      "威胁标签"),
    ("数据源",        "数据源"),
    ("检测方式",      "检测方式"),
    ("CVE 编号",      "CVE 编号"),
    ("最近发现时间",  "最近发现时间"),
    ("首次发现时间",  "首次发现时间"),
    ("风险资产",      "风险资产"),
    ("资产类型",      "资产类型"),
    ("所属资产组",    "所属资产组"),
    ("所属业务",      "所属业务"),
    ("资产责任人",    "资产责任人"),
    ("资产重要性",    "资产重要性"),
    ("资产管理状态",  "资产管理状态"),
    ("托管状态",      "托管状态"),
    ("端口",          "端口"),
    ("url",           "url"),
    ("互联网暴露",    "互联网暴露"),
    ("举证信息",      "举证信息"),
    ("处置状态",      "处置状态"),
    ("数据来源",      "__FIXED_内网"), # 固定为 "内网"
]


def process_mssw_file(file_b_path: str) -> List[dict]:
    """
    处理 MSSW 导出的漏洞文件 B：
    - 重新计算修复优先级
    - 映射为文件C 的统一格式
    返回 list[dict]
    """
    wb = load_workbook(file_b_path, data_only=True)
    ws = wb.active

    col_map: Dict[str, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value is not None:
            col_map[str(cell.value).strip()] = col_idx

    log(f"MSSW 文件B 列数: {len(col_map)}, 数据行数: {ws.max_row - 1}")

    # 确认必要列存在
    needed = ["漏洞名称", "风险等级", "威胁标签"]
    for key in needed:
        if key not in col_map:
            log(f"  MSSW 文件缺少必要列: {key}，实际列: {list(col_map.keys())}", "WARNING")
            raise RuntimeError(f"MSSW 文件B 缺少必要列「{key}」")

    rows_out: List[dict] = []

    for row_idx in range(2, ws.max_row + 1):
        row_data = {}

        for c_col, source in C_FROM_MSSW:
            if source is None:
                # 修复优先级：需计算
                risk_raw = ws.cell(row=row_idx, column=col_map["风险等级"]).value
                risk_norm = normalize_risk(str(risk_raw or ""))

                # 威胁标签：是否包含"高可利用"
                threat_tag = str(ws.cell(row=row_idx, column=col_map["威胁标签"]).value or "")
                is_exploitable = "高可利用" in threat_tag

                # MSSW 固定为内网
                row_data[c_col] = calc_priority(risk_norm, is_internal=True, is_exploitable=is_exploitable)

            elif source.startswith("__FIXED_"):
                # 固定值，如 __FIXED_内网 → "内网"
                row_data[c_col] = source.replace("__FIXED_", "")

            else:
                src_col = col_map.get(source)
                if src_col is not None:
                    val = ws.cell(row=row_idx, column=src_col).value
                    val = val if val is not None else ""
                    if c_col in ("最近发现时间", "首次发现时间"):
                        val = _fmt_minute(val)
                    row_data[c_col] = val
                else:
                    log(f"  第{row_idx}行: 列「{source}」不存在，赋空值", "WARNING")
                    row_data[c_col] = ""

        rows_out.append(row_data)

    wb.close()
    log(f"MSSW 处理后共 {len(rows_out)} 行")
    return rows_out


def extract_mssw_header_styles(file_b_path: str) -> Tuple[Dict[str, object], object]:
    """
    从 MSSW 导出文件第 1 行表头提取样式信息。
    返回:
      header_styles: {列名 → {'font','fill','border','alignment','number_format'}}
      ref_style:    参考样式（取第 1 个有样式的表头单元格），用于额外字段列回退
    异常或空表头时返回 ({}, None)，不抛错。

    注意：不保存 cell._style（那是工作簿内部索引，不可跨簿复制），
    而是保存 Font/Fill/Border/Alignment 等可移植的样式属性对象。
    """
    try:
        wb = load_workbook(file_b_path, data_only=True)
    except Exception as e:
        log(f"  提取MSSW表头样式失败（文件读取失败）: {e}", "WARNING")
        return {}, None

    try:
        ws = wb.active
        header_styles: Dict[str, object] = {}
        ref_style: object = None

        for cell in ws[1]:
            if cell.value is None:
                continue
            if not cell.has_style:          # 跳过无显式样式的列，避免默认样式污染 ref_style
                continue
            col_name = str(cell.value).strip()
            # 保存样式属性对象（可跨工作簿使用），而非 _style（工作簿内部索引）
            style_props = {
                "font":          copy(cell.font),
                "fill":          copy(cell.fill),
                "border":        copy(cell.border),
                "alignment":     copy(cell.alignment),
                "number_format": cell.number_format,  # 字符串不可变，无需 copy
            }
            header_styles[col_name] = style_props
            if ref_style is None:
                ref_style = dict(style_props)       # 浅拷贝 dict，避免与第一列共享同一对象

        if not header_styles:
            log("  MSSW 表头为空或无样式，最终 Excel 表头将使用默认样式", "WARNING")

        return header_styles, ref_style
    except Exception as e:
        log(f"  提取MSSW表头样式失败（遍历异常）: {e}", "WARNING")
        return {}, None
    finally:
        wb.close()


# ====================================================================
#  主流程
# ====================================================================

def _on_script_timeout():
    log(f"错误：脚本执行超时（超过 {SCRIPT_TIMEOUT // 60} 分钟），强制退出", "ERROR")
    os._exit(1)


def main():
    _timer = threading.Timer(SCRIPT_TIMEOUT, _on_script_timeout)
    _timer.daemon = True
    _timer.start()

    parser = argparse.ArgumentParser(
        description='漏洞清单生成工具',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            '示例:\n'
            '  python vuln_report.py 深圳市口袋网络\n'
            '  python vuln_report.py 35690473\n'
            '  python vuln_report.py 客户名 --skip-easm\n'
            '  python vuln_report.py 客户名 --skip-mssw\n'
            '  python vuln_report.py 客户名 --start-time 178231680000 --end-time 1782853199000'
        )
    )
    parser.add_argument('keyword', nargs='?', default=None,
                        help='客户ID或客户名称（支持模糊匹配）')
    parser.add_argument('--skip-mssw', action='store_true',
                        help='跳过MSSW平台（内网）')
    parser.add_argument('--skip-easm', action='store_true',
                        help='跳过EASM平台（外网）')
    parser.add_argument('--start-time', default=None,
                        help='（可选）过滤开始日期，如 2016-01-01 或 2016年1月1日')
    parser.add_argument('--end-time', default=None,
                        help='（可选）过滤结束日期，如 2026-06-30 或 2026年6月30日')
    args = parser.parse_args()

    if not args.keyword:
        sys.exit("错误：请提供客户ID或客户名称关键词\n示例: python vuln_report.py 35690473")

    os.makedirs(TEMP_DIR, exist_ok=True)

    # 时间过滤：传入则过滤，不传则 []（全部数据）
    if args.start_time and args.end_time:
        start_ms = parse_date_to_ms(args.start_time, is_end=False)
        end_ms   = parse_date_to_ms(args.end_time,   is_end=True)
        time_range = [start_ms, end_ms]
        log(f"时间过滤: {args.start_time} → {start_ms}  /  {args.end_time} → {end_ms}")
    else:
        time_range = []




    # ==================== 1. 加载 Cookie ====================
    log("=" * 50)
    log("步骤1：加载 Cookie")

    # EASM
    if not args.skip_easm:
        if not os.path.exists(EASM_COOKIES_FILE):
            sys.exit(f"错误：EASM cookies文件不存在 → {EASM_COOKIES_FILE}")
        easm_cookie = read_cookies_as_string(EASM_COOKIES_FILE)
        if not easm_cookie:
            sys.exit(f"错误：EASM cookies解析结果为空 → {EASM_COOKIES_FILE}")
        log(f"  EASM Cookie: {EASM_COOKIES_FILE}")
    else:
        easm_cookie = ""
        log(f"  EASM Cookie: 已跳过（--skip-easm）")

    # MSSW
    mssw_cookie = ""  # 可能通过 --skip-mssw 跳过
    if not args.skip_mssw:
        if not os.path.exists(MSSW_COOKIES_FILE):
            sys.exit(f"错误：MSSW cookies文件不存在 → {MSSW_COOKIES_FILE}")
        mssw_cookie = read_cookies_as_string(MSSW_COOKIES_FILE)
        if not mssw_cookie:
            sys.exit(f"错误：MSSW cookies解析结果为空 → {MSSW_COOKIES_FILE}")
        log(f"  MSSW Cookie: {MSSW_COOKIES_FILE}")
    else:
        log(f"  MSSW Cookie: 已跳过（--skip-mssw）")

    # ==================== 2. 搜索客户（分平台独立搜索） ====================
    log("=" * 50)
    log("步骤2：搜索客户")

    # 2a. EASM 客户搜索
    easm_company_id = ""
    easm_company_name = args.keyword
    if args.skip_easm:
        log("  步骤2a：EASM — 已跳过（--skip-easm）")
    else:
        easm_keyword = args.keyword + "[影子]"
        log(f"  步骤2a：EASM 搜索客户「{easm_keyword}」...")
        easm_customers = api0_search_customer(easm_cookie, easm_keyword)
        if not easm_customers:
            sys.exit(f"错误：EASM 未找到匹配客户，请检查关键词「{easm_keyword}」")
        if len(easm_customers) > 1:
            exact = _pick_exact_match(easm_customers, easm_keyword)
            if exact:
                easm_customers = [exact]
            else:
                log(f"错误：EASM 找到 {len(easm_customers)} 个匹配客户，请使用更精确的关键词：", "ERROR")
                for c in easm_customers:
                    name = c.get('company_name') or c.get('pms_customer_name', '未知')
                    print(f"  ID={c['company_id']}  名称={name}")
                sys.exit(1)
        c = easm_customers[0]
        easm_company_id   = c['company_id']
        easm_company_name = c.get('company_name') or c.get('pms_customer_name', '未知')
        log(f"  确认 EASM 客户: {easm_company_name}（ID={easm_company_id}）")

    # 2b. MSSW 客户搜索
    mssw_company_id = ""
    mssw_company_name = args.keyword
    if args.skip_mssw:
        log("  步骤2b：MSSW — 已跳过（--skip-mssw）")
    else:
        log(f"  步骤2b：MSSW 搜索客户「{args.keyword}」...")
        mssw_customers = mssw_api0_search_customer(mssw_cookie, args.keyword)
        if not mssw_customers:
            sys.exit(f"错误：MSSW 未找到匹配客户，请检查关键词「{args.keyword}」")
        if len(mssw_customers) > 1:
            exact = _pick_exact_match(mssw_customers, args.keyword)
            if exact:
                mssw_customers = [exact]
            else:
                log(f"错误：MSSW 找到 {len(mssw_customers)} 个匹配客户，请使用更精确的关键词：", "ERROR")
                for c in mssw_customers:
                    name = c.get('company_name') or c.get('pms_customer_name', '未知')
                    print(f"  ID={c['company_id']}  名称={name}")
                sys.exit(1)
        c = mssw_customers[0]
        mssw_company_id   = c['company_id']
        mssw_company_name = c.get('company_name') or c.get('pms_customer_name', '未知')
        log(f"  确认 MSSW 客户: {mssw_company_name}（ID={mssw_company_id}）")

    # ==================== 3. EASM 平台处理 ====================
    log("=" * 50)
    if args.skip_easm:
        log("步骤3：EASM 平台（外网）已跳过（--skip-easm）")
        easm_data = []
    else:
        log("步骤3：EASM 平台（外网）漏洞数据获取")
        export_time = datetime.now().strftime("%Y%m%d%H%M%S")
        easm_data: List[dict] = []
        scan_method_labels = {0: "原理扫描", 2: "人工录入"}

        for sm in [0, 2]:
            label = scan_method_labels.get(sm, f"scan_method={sm}")
            log(f"  3a. 触发漏洞导出（scan_method={sm}，{label}）...")
            easm_task_id = easm_api5_export_vuln(easm_cookie, easm_company_id,
                                                  scan_method=sm,
                                                  last_time=time_range)

            # 3b. 接口6：轮询状态，获取下载 URL
            log(f"  3b. 轮询导出状态（{label}）...")
            easm_download_path = easm_api6_poll_export(easm_cookie, easm_task_id)

            # 3c. 下载文件
            log(f"  3c. 下载漏洞文件（{label}）...")
            file_path = easm_download_file(easm_cookie, easm_download_path, TEMP_DIR)
            easm_file_name = f"EASM-漏洞表{easm_company_name}_{label}_{export_time}.xlsx"
            renamed = os.path.join(TEMP_DIR, easm_file_name)
            if os.path.exists(renamed):
                os.remove(renamed)
            os.rename(file_path, renamed)
            file_path = renamed
            log(f"  文件: {os.path.basename(file_path)}")

            # 处理并累积数据
            part_data = process_easm_file(file_path)
            easm_data.extend(part_data)
            log(f"  {label} 数据: {len(part_data)} 行（累计 {len(easm_data)} 行）")

    log(f"  EASM 数据准备完毕: {len(easm_data)} 行")

    # ==================== 4. MSSW 平台处理 ====================
    log("=" * 50)
    file_b_path = ""  # 初始化，--skip-mssw 时保持为空
    if args.skip_mssw:
        log("步骤4：MSSW 平台（内网）已跳过（--skip-mssw）")
        mssw_data = []
    else:
        log("步骤4：MSSW 平台（内网）漏洞数据获取")

        # 4a. 接口8：触发导出
        log("  4a. 触发MSSW漏洞导出...")
        mssw_file_name = mssw_api8_export_vuln(mssw_cookie, mssw_company_id, time_range)

        # 4b. 接口10：下载文件 B
        log("  4b. 下载漏洞文件...")
        file_b_raw = mssw_api10_download(mssw_cookie, mssw_company_id, mssw_file_name, TEMP_DIR)
        # 重命名（先删旧文件，避免 Windows rename 报错）
        now_str = datetime.now().strftime("%Y%m%d%H%M%S")
        mssw_renamed = f"漏洞-脆弱性导出报告_{now_str}.xlsx"
        file_b_path = os.path.join(TEMP_DIR, mssw_renamed)
        if os.path.exists(file_b_path):
            os.remove(file_b_path)
        os.rename(file_b_raw, file_b_path)
        log(f"  文件B: {os.path.basename(file_b_path)}")

        # 4c. 处理文件B → 统一格式
        log("  4c. 处理MSSW数据（重新计算修复优先级）...")
        mssw_data = process_mssw_file(file_b_path)
        log(f"  MSSW 数据准备完毕: {len(mssw_data)} 行")

    # # ==================== 5. 合并输出 ====================
    log("=" * 50)
    log("步骤5：合并生成最终文件")

    # 文件C 的固定列头
    C_HEADERS = [
        "漏洞名称", "修复优先级", "风险等级", "漏洞类型", "修复建议",
        "风险描述", "威胁标签", "数据源", "检测方式", "CVE 编号",
        "最近发现时间", "首次发现时间", "风险资产", "资产类型",
        "所属资产组", "所属业务", "资产责任人", "资产重要性",
        "资产管理状态", "托管状态", "端口", "url", "互联网暴露", "举证信息",
        "处置状态", "数据来源",
    ]

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "漏洞"

    # 提取 MSSW 表头样式（仅当 MSSW 文件可用时）
    header_styles: Dict[str, object] = {}
    ref_style: object = None
    if not args.skip_mssw and file_b_path and os.path.exists(file_b_path):
        log("  提取 MSSW 表头样式...")
        header_styles, ref_style = extract_mssw_header_styles(file_b_path)
        log(f"  已提取 {len(header_styles)} 个列样式，参考样式={'有' if ref_style else '无'}")

    # 写入表头并应用样式
    for ci, header in enumerate(C_HEADERS, start=1):
        cell = ws_out.cell(row=1, column=ci, value=header)
        style_props = header_styles.get(header)
        if style_props is None:
            style_props = ref_style
        if style_props is not None:
            try:
                cell.font          = copy(style_props["font"])
                cell.fill          = copy(style_props["fill"])
                cell.border        = copy(style_props["border"])
                cell.alignment     = copy(style_props["alignment"])
                cell.number_format = style_props["number_format"]
            except Exception as e:
                log(f"  应用表头样式失败 列={header}: {e}", "WARNING")
    #
    # # 写入 MSSW 数据（文档中 MSSW 在前）
    row_num = 2
    for record in mssw_data:
        for ci, header in enumerate(C_HEADERS, start=1):
            ws_out.cell(row=row_num, column=ci, value=record.get(header, ""))
        row_num += 1

    # 写入 EASM 数据
    for record in easm_data:
        for ci, header in enumerate(C_HEADERS, start=1):
            ws_out.cell(row=row_num, column=ci, value=record.get(header, ""))
        row_num += 1

    wb_out.save(OUTPUT_FILE)
    wb_out.close()

    total_rows = len(easm_data) + len(mssw_data)
    _timer.cancel()
    log(f"完成！")
    log(f"  MSSW  数据: {len(mssw_data)} 行")
    log(f"  EASM 数据: {len(easm_data)} 行")
    log(f"  合计: {total_rows} 行")
    log(f"  输出文件: {OUTPUT_FILE}")


if __name__ == '__main__':
    main()

# ============================================================
#  使用说明
# ============================================================
#
#  python vuln_report.py <客户ID或关键词> [选项]
#
#  必填参数:
#     keyword                 客户ID或客户名称（支持模糊匹配）
#
#  可选参数:
#     --skip-easm              跳过EASM平台（外网），只获取MSSW内网数据
#     --skip-mssw               跳过MSSW平台（内网），只获取EASM外网数据
#     --start-time <时间戳>    过滤开始时间（13位毫秒时间戳）
#     --end-time   <时间戳>    过滤结束时间（13位毫秒时间戳）
#
#  前置条件:
#     1. EASM、MSSW 的 Cookie 分别保存在 cookies.txt / mssw_cookies.txt
#     2. 文件路径在脚本顶部【配置】段修改 TEMP_DIR / OUTPUT_FILE
#
#  示例:
#     python vuln_report.py 深圳市口袋网络              # 两个平台一起跑
#     python vuln_report.py 35690473 --skip-mssw         # 仅 EASM
#     python vuln_report.py 客户名 --skip-easm          # 仅 MSSW
#     python vuln_report.py 客户名 --start-time xxx --end-time yyy  # 时间过滤
#
#  输出:
#     漏洞清单.xlsx，sheet名"漏洞"
#     表头: 漏洞名称 | 修复优先级 | 风险等级 | 漏洞类型 | 修复建议 |
#           风险描述 | 威胁标签 | 数据源 | 检测方式 | CVE 编号 |
#           最近发现时间 | 首次发现时间 | 风险资产 | 资产类型 |
#           所属资产组 | 所属业务 | 资产责任人 | 资产重要性 |
#           资产管理状态 | 端口 | url | 互联网暴露 | 举证信息 |
#           处置状态 | 数据来源
#
#  过滤逻辑:
#     EASM 平台: vulnerability_status=[5(未审核),1(处置中),12(待复测),9(修复失败)]
#                scan_method=0(原理扫描) 和 scan_method=2(人工录入) 各调用一次接口
#                last_time 可选时间过滤
#     MSSW 平台:  fixed_status=[0(待处置),1(处置中)]
#                latest_time_range 可选时间过滤
#
# ============================================================