#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
弱口令清单生成工具
从 EASM 平台（外网）和 MSSW 平台（内网）分别获取弱口令数据，
合并生成统一的「弱口令清单.xlsx」。

用法: python weak_report.py <客户ID或客户名称关键词>
示例: python weak_report.py 深圳市口袋网络
      python weak_report.py 35690473
"""

import os
import re
import sys
import time
import threading
import argparse
import requests
from copy import copy
from openpyxl.styles import Font, PatternFill, Border, Alignment
from datetime import datetime, timezone, timedelta
from typing import Dict, List, Optional, Tuple
import openpyxl
from openpyxl import load_workbook


# ==================== 配置（按需修改） ====================

# --- 通用 ---
TEMP_DIR     = r"C:\Users\User\Downloads\temp_report"
OUTPUT_FILE  = r"C:\Users\User\Downloads\弱口令清单.xlsx"

POLL_INTERVAL   = 5     # 轮询间隔（秒）
SCRIPT_TIMEOUT  = 3600  # 全局超时：1小时
MAX_RETRIES     = 3     # 最大重试次数
RETRY_DELAY     = 3     # 重试等待时间（秒）
PAGE_LIMIT      = 100   # 列表接口每次查询数量

# --- EASM 平台（外网） ---
EASM_BASE_URL    = "https://soar59.sangfor.com.cn"
EASM_COOKIES_FILE = r"C:\Users\User\Downloads\cookies.txt"

# --- MSSW 平台（内网） ---
MSSW_BASE_URL      = "https://pre.soar.sangfor.com"
MSSW_COOKIES_FILE   = r"C:\Users\User\Downloads\mssw_cookies.txt"



# EASM 处置状态中文映射（deal_status）
EASM_DEAL_STATUS_MAP = {
    -1: "超时未审核",
    0: "未审核",
    1: "处置中",
    2: "已修复",
    3: "已标为误判",
    4: "接受风险",
    5: "已防护",
    6: "待复测",
    7: "修复失败",
}
EASM_DEAL_STATUS_FILTER = [0, 1, 6, 7]   # 过滤处置状态：[未审核=0, 处置中=1, 待复测=6, 修复失败=7]
MSSW_FIXED_STATUSES = [0, 1]               # 过滤处置状态：[待处置=0, 处置中=1]
# ==========================================================


# ---------- 日志 ----------

def log(msg: str, level: str = "INFO") -> None:
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}][{level}] {msg}")


# ---------- 工具函数 ----------

def _get_system_timezone() -> str:
    offset = datetime.now().astimezone().utcoffset()
    if offset is None:
        return "+00:00"
    total_seconds = int(offset.total_seconds())
    sign = "+" if total_seconds >= 0 else "-"
    total_seconds = abs(total_seconds)
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    return f"{sign}{hours:02d}:{minutes:02d}"


def extract_cookie_value(cookie_str: str, name: str) -> Optional[str]:
    for part in cookie_str.split(';'):
        part = part.strip()
        if '=' in part:
            k, _, v = part.partition('=')
            if k.strip() == name:
                return v.strip()
    return None


# ---------- Cookie 读取 ----------

def read_cookies_as_string(filepath: str) -> str:
    """
    读取 cookies.txt，返回 Cookie 请求头字符串（name=value; name2=value2 格式）。
    支持两种输入格式：
      1. 浏览器原始 Cookie 字符串
      2. Netscape 格式（EditThisCookie 等工具导出，tab 分隔）
    """
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read().strip()

    lines = [ln.strip() for ln in content.splitlines() if ln.strip()]
    is_netscape = any('\t' in ln for ln in lines if not ln.startswith('#'))

    if is_netscape:
        pairs = []
        for ln in lines:
            if ln.startswith('#'):
                continue
            parts = ln.split('\t')
            if len(parts) >= 7:
                pairs.append(f"{parts[5]}={parts[6]}")
        return '; '.join(pairs)
    else:
        return '; '.join(lines)


# ---------- 请求封装 ----------

def _build_default_headers(base_url: str) -> Dict[str, str]:
    return {
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Content-Type": "application/json",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": f"{base_url}/index.html",
        "timezone": _get_system_timezone(),
    }


def request_with_retry(method: str, url: str, base_url: str,
                       cookie_str: str = "",
                       timeout: int = 30,
                       extra_headers: Optional[Dict] = None,
                       **kwargs) -> Optional[requests.Response]:
    headers = _build_default_headers(base_url)
    if extra_headers:
        headers.update(extra_headers)
    if cookie_str:
        headers["Cookie"] = cookie_str
        csrf_token = extract_cookie_value(cookie_str, "csrf_token")
        if csrf_token:
            headers["X-Csrftoken"] = csrf_token
        mssw_csrf = extract_cookie_value(cookie_str, "x-csrf-token")
        if mssw_csrf:
            headers["x-csrf-token"] = mssw_csrf

    session = requests.Session()
    adapter = requests.adapters.HTTPAdapter(
        pool_connections=10, pool_maxsize=10, max_retries=0
    )
    session.mount('https://', adapter)
    session.mount('http://', adapter)

    for attempt in range(MAX_RETRIES):
        try:
            if method.upper() == "POST":
                resp = session.post(url, headers=headers, timeout=timeout,
                                    verify=False, **kwargs)
            else:
                resp = session.get(url, headers=headers, timeout=timeout,
                                   verify=False, **kwargs)
            return resp
        except (requests.exceptions.Timeout, requests.exceptions.ConnectionError,
                requests.exceptions.SSLError, requests.exceptions.RequestException) as e:
            attempt_num = attempt + 1
            log(f"请求失败 (尝试 {attempt_num}/{MAX_RETRIES}) - {e}", "WARNING")
            if attempt < MAX_RETRIES - 1:
                time.sleep(RETRY_DELAY)
            else:
                log(f"重试 {MAX_RETRIES} 次后仍然失败，返回 None", "ERROR")
                return None
    return None


def _parse_json(resp: requests.Response, api_name: str) -> dict:
    if resp is None:
        raise RuntimeError(f"{api_name}：请求失败（已重试 {MAX_RETRIES} 次）")
    try:
        return resp.json()
    except Exception:
        raise RuntimeError(
            f"{api_name}：响应不是JSON（状态码={resp.status_code}，可能Cookie已过期）"
        )


def date_str_to_ms(date_str: str) -> Optional[int]:
    """
    将 EASM 返回的时间字符串（UTC+8，如 '2023-12-06 00:40'）
    转为 13 位 UTC 毫秒时间戳，用于与命令行参数比较。
    """
    if not date_str:
        return None
    from datetime import timezone, timedelta
    try:
        tz_cst = timezone(timedelta(hours=8))
        dt = datetime.strptime(date_str.strip()[:16], "%Y-%m-%d %H:%M")
        dt = dt.replace(tzinfo=tz_cst)
        return int(dt.timestamp() * 1000)
    except (ValueError, TypeError):
        return None


def parse_date_to_ms(date_str: str, is_end: bool = False) -> int:
    """
    将日期字符串转为13位 UTC+8 毫秒时间戳。
    支持格式: '2016-01-01' 或 '2016年1月1日'
    is_end=False → 当天 00:00:00.000
    is_end=True  → 当天 23:59:59.999
    """
    date_str = date_str.replace('年', '-').replace('月', '-').replace('日', '').strip()
    dt = datetime.strptime(date_str[:10], "%Y-%m-%d")
    dt = dt.replace(tzinfo=timezone(timedelta(hours=8)))
    if is_end:
        return int(dt.timestamp() * 1000) + 86399999
    return int(dt.timestamp() * 1000)


def parse_relative_time_to_ms(relative_str: str) -> int:
    """
    解析相对时间字符串 → UTC+8 毫秒时间戳。
    支持格式: '19小时42分钟前' / '35分钟前' / '30秒前'
    """
    now = datetime.now(tz=timezone(timedelta(hours=8)))
    m = re.match(r'(\d+)小时(\d+)分钟前', relative_str)
    if m:
        dt = now - timedelta(hours=int(m.group(1)), minutes=int(m.group(2)))
        return int(dt.timestamp() * 1000)
    m = re.match(r'(\d+)分钟前', relative_str)
    if m:
        dt = now - timedelta(minutes=int(m.group(1)))
        return int(dt.timestamp() * 1000)
    m = re.match(r'(\d+)秒前', relative_str)
    if m:
        dt = now - timedelta(seconds=int(m.group(1)))
        return int(dt.timestamp() * 1000)
    return 0


def resolve_parent_found_time(parent: dict) -> str:
    """
    从母表取值「最近发现时间」。
    found_time 可能被 calculate_assess_time 转为相对时间（如"3小时前"），
    此时用 parse_relative_time_to_ms 反向计算还原为 UTC+8 日期时间字符串。
    """
    found_time = str(parent.get('found_time', '') or '')
    if not found_time:
        return ''
    relative_keywords = ['小时前', '分钟前', '秒前']
    if any(kw in found_time for kw in relative_keywords):
        ts = parse_relative_time_to_ms(found_time)
        if ts:
            dt = datetime.fromtimestamp(ts / 1000, tz=timezone(timedelta(hours=8)))
            return dt.strftime("%Y-%m-%d %H:%M")
    return found_time


# ---------- 通用：搜索客户结果匹配 ----------

def _pick_exact_match(customers: list, keyword: str):
    """从多个模糊搜索结果中优先选精确匹配。返回匹配项或 None"""
    if len(customers) == 1:
        return customers[0]
    exact = [c for c in customers
             if (c.get('company_name', '') or '').strip() == keyword.strip()
             or (c.get('pms_customer_name', '') or '').strip() == keyword.strip()
             or str(c.get('company_id', '')).strip() == keyword.strip()]
    return exact[0] if exact else None


# ---------- 通用：接口0 搜索客户 ----------

def api0_search_customer(cookie_str: str, keyword: str) -> list:
    """根据关键词模糊搜索客户，返回列表"""
    url = f"{EASM_BASE_URL}/gateway/customer-mgr-service/order/v1/user?_method=GET"
    # 正确入参
    payload = {
        "order": "asc", "offset": 0, "limit": 20, "keyword": keyword,
        "share_ids": [], "delivery_channel_id": [], "service_code": [],
        "industry": [], "industry_segmentation": [], "customer_type": [],
        "customer_stratification": [], "protection_type": [], "service_group": [],
        "delivery_method": [], "platform_type": [], "service_status": 0, "my_customer": 0,
    }
    # 35环境入参（废弃）
    # payload = {
    #     "order": "asc", "offset": 0, "limit": 10, "keyword": keyword,
    #     "share_ids": [], "delivery_channel_id": [], "phase": [],
    #     "industry": [], "industry_segmentation": [], "auth_status": 2,
    #     "customer_type": [], "customer_stratification": [],
    #     "protection_type": [], "delivery_method": [], "platform_type": [],
    #     "operate_status": [], "service_status": 0, "my_customer": 0,
    # }
    resp = request_with_retry("POST", url, EASM_BASE_URL, cookie_str, json=payload, timeout=120)
    data = _parse_json(resp, "接口0")
    if data.get('code') != 0:
        raise RuntimeError(f"接口0失败: {data.get('msg')}")
    return data['data']['list']


# ---------- MSSW：接口8-0 搜索客户 ----------

def mssw_api0_search_customer(cookie_str: str, keyword: str) -> list:
    """接口8-0：MSSW 平台客户搜索，根据关键词（名称或ID）模糊搜索"""
    url = f"{MSSW_BASE_URL}/gateway/customer-mgr-service/order/v1/user?_method=GET"
    key_field = "company_id_keyword" if keyword.isdigit() else "company_name_keyword"
    payload = {
        "my_customer": 0,
        key_field: keyword,
        "offset": 0, "limit": 20,
    }
    resp = request_with_retry("POST", url, MSSW_BASE_URL, cookie_str, json=payload, timeout=120)
    data = _parse_json(resp, "接口8-0（MSSW客户搜索）")
    if data.get('code') != 0:
        raise RuntimeError(f"接口8-0失败: {data.get('msg')}")
    return data['data']['list']


# ====================================================================
#  EASM 平台（外网）接口调用
# ====================================================================

def easm_api7_1_weak_pwd_list(cookie_str: str, company_id: str) -> List[dict]:
    """接口7-1：分页获取母表弱口令列表，返回所有数据"""
    url = f"{EASM_BASE_URL}/gateway/vuln-manager/vm/order/v1/weak_pwd/easm/summary_list"
    all_items: List[dict] = []
    offset = 0
    total = None

    while True:
        payload = {
            "order": {},
            "offset": offset,
            "limit": PAGE_LIMIT,
            "keyword": "",
            "target_company_id": [],
            "task_id": [],
            "found_time": [],
            "verified_time": [],
            "closed_loop_time": [],
            "deal_status": EASM_DEAL_STATUS_FILTER,
            "company_id": company_id,
        }
        resp = request_with_retry("POST", url, EASM_BASE_URL, cookie_str, json=payload, timeout=120)
        data = _parse_json(resp, "接口7-1（EASM母表弱口令列表）")
        if data.get('code') != 0:
            raise RuntimeError(f"接口7-1失败: {data.get('msg')}")

        list_data = data.get('data', {})
        if total is None:
            total = list_data.get('total', 0)
            log(f"EASM 弱口令母表总数: {total}")

        page_list = list_data.get('list', [])
        all_items.extend(page_list)
        offset += PAGE_LIMIT

        log(f"  已获取母表 {len(all_items)}/{total} 条")
        if len(page_list) < PAGE_LIMIT or len(all_items) >= total:
            break

    return all_items


def easm_api7_2_weak_pwd_sub_list(cookie_str: str, company_id: str, ip: str) -> List[dict]:
    """接口7-2：获取指定 IP 下的子表弱口令列表（分页全量）"""
    url = (f"{EASM_BASE_URL}/gateway/vuln-manager/vm/order/v1/weak_pwd/easm/list"
           f"?ip={ip}")
    all_items: List[dict] = []
    offset = 0
    total = None

    while True:
        payload = {
            "order": {},
            "offset": offset,
            "limit": PAGE_LIMIT,
            "keyword": "",
            "target_company_id": [],
            "task_id": [],
            "found_time": [],
            "verified_time": [],
            "closed_loop_time": [],
            "deal_status": EASM_DEAL_STATUS_FILTER,
            "company_id": company_id,
            "ip": ip,
        }
        resp = request_with_retry("POST", url, EASM_BASE_URL, cookie_str, json=payload, timeout=120)
        data = _parse_json(resp, f"接口7-2（子表 ip={ip}）")
        if data.get('code') != 0:
            log(f"  接口7-2失败 ip={ip}: {data.get('msg')}", "WARNING")
            break

        list_data = data.get('data', {})
        if total is None:
            total = list_data.get('total', 0)

        page_list = list_data.get('list', [])
        all_items.extend(page_list)
        offset += PAGE_LIMIT

        if len(page_list) < PAGE_LIMIT or len(all_items) >= (total or 0):
            break

    return all_items


def easm_weak_pwd_detail(cookie_str: str, pwd_id: str) -> dict:
    """补充接口：弱密码详情，返回 detail dict（含 found_time / login_time / proof）"""
    url = (f"{EASM_BASE_URL}/gateway/vuln-manager/vm/order/v1/weak_pwd/"
           f"weak_pwd_info?_method=GET")
    payload = {"pwd_id": pwd_id}
    resp = request_with_retry("POST", url, EASM_BASE_URL, cookie_str, json=payload, timeout=60)
    data = _parse_json(resp, f"弱密码详情（pwd_id={pwd_id}）")
    if data.get('code') != 0:
        log(f"  弱密码详情失败 pwd_id={pwd_id}: {data.get('msg')}", "WARNING")
        return {}
    return data.get('data', {})


# ====================================================================
#  MSSW 平台（内网）接口调用
# ====================================================================

def mssw_api9_export_weak_pwd(cookie_str: str, company_id: str, latest_time_range: list) -> str:
    """接口9：触发弱口令导出，返回 file_name"""
    url = f"{MSSW_BASE_URL}/order/v1/vul_manage/vul_risk_export"
    payload = {
        "asset_ip": {"op": "=", "val": ""},
        "asset_manager": {"op": "=", "val": ""},
        "asset_status": [],
        "asset_tags": [],
        "asset_type": "all",
        "attack_state": [],
        "branch_ids": [],
        "disposal_tag": [],
        "exposure": [],
        "fix_priority": [],
        "fixed_status": MSSW_FIXED_STATUSES,
        "group_ids": [],
        "keyword": "",
        "keyword_all": "",
        "magnitude": [],
        "name": {"op": "=", "val": ""},
        "order_status": [],
        "platform_ids": [],
        "platform_filter": [],
        "retest_status": [],
        "source_device": [],
        "data_type": ["weak_pwd"],
        "whitelisted_status": [],
        "latest_time_range": latest_time_range,
        "is_show": 1,
        "custom_headers": {
            "asset_info": [
                {"disabled": True,  "key": "asset",                "label": "风险资产",      "selected": True},
                {"disabled": False, "key": "asset_type",           "label": "资产类型",      "selected": True},
                {"disabled": False, "key": "business_name",        "label": "所属资产组",    "selected": True},
                {"disabled": False, "key": "group_name",           "label": "所属业务",      "selected": True},
                {"disabled": False, "key": "manager",              "label": "资产责任人",    "selected": True},
                {"disabled": False, "key": "magnitude",            "label": "资产重要性",    "selected": True},
                {"disabled": True,  "key": "port",                 "label": "端口",          "selected": True},
                {"disabled": False, "key": "exposure",             "label": "互联网暴露",    "selected": True},
                {"disabled": False, "key": "evidence_information", "label": "举证信息",      "selected": True},
                {"disabled": False, "key": "asset_status",         "label": "资产管理状态",  "selected": True},
                {"disabled": False, "key": "platform_name",        "label": "来源平台",      "selected": True},
                {"disabled": False, "key": "managed_level",      "label": "托管状态",      "selected": True},
            ],
            "base_info": [
                {"disabled": True,  "key": "name",              "label": "弱密码名称",    "selected": True},
                {"disabled": True,  "key": "fix_priority_level","label": "修复优先级",    "selected": True},
                {"disabled": True,  "key": "risk_Level",        "label": "风险等级",      "selected": True},
                {"disabled": True,  "key": "user",              "label": "账号",          "selected": True},
                {"disabled": True,  "key": "pwd",               "label": "密码",          "selected": True},
                {"disabled": True,  "key": "url",               "label": "url",           "selected": True},
                {"disabled": False, "key": "refer",             "label": "refer",         "selected": True},
                {"disabled": False, "key": "process_path",      "label": "进程路径",      "selected": True},
                {"disabled": False, "key": "src_type",          "label": "数据源",        "selected": True},
                {"disabled": False, "key": "last_time",         "label": "最近发现时间",  "selected": True},
                {"disabled": False, "key": "found_time",        "label": "首次发现时间",  "selected": True},
                {"disabled": False, "key": "is_gpt",            "label": "GPT检测",       "selected": True},
            ],
            "disposal_info": [
                {"disabled": False, "key": "whitelisted_status","label": "加白状态",      "selected": True},
                {"disabled": False, "key": "fixed_status",      "label": "处置状态",      "selected": True},
                {"disabled": False, "key": "fixed_tag",         "label": "处置标签",      "selected": True},
                {"disabled": False, "key": "order_progress",    "label": "最新工单进展",  "selected": True},
                {"disabled": False, "key": "retest_status",     "label": "验证状态",      "selected": True},
            ],
        },
        "header_id": "week_pass_1",
        "is_all": False,
        "multiple_choice": [],
        "exclude_multiple_choice": [],
    }
    extra_hdrs = {"X-MSSW-Company-Id": company_id} if company_id else None
    resp = request_with_retry("POST", url, MSSW_BASE_URL, cookie_str,
                               json=payload, timeout=120, extra_headers=extra_hdrs)
    data = _parse_json(resp, "接口9（MSSW弱口令导出）")
    if data.get('code') != 0:
        raise RuntimeError(f"接口9失败: {data.get('message') or data.get('msg')}")
    file_name = data['data']['file_name']
    log(f"MSSW 弱口令导出文件名: {file_name}")
    return file_name


def mssw_api10_download(cookie_str: str, company_id: str, file_name: str, save_dir: str) -> str:
    """接口10：下载 MSSW 导出文件，返回本地文件路径"""
    url = f"{MSSW_BASE_URL}/order/v1/vul_manage/download_file?file={file_name}"
    extra_hdrs = {"X-MSSW-Company-Id": company_id} if company_id else None
    resp = request_with_retry("GET", url, MSSW_BASE_URL, cookie_str, stream=True, timeout=120, extra_headers=extra_hdrs)
    if resp is None:
        raise RuntimeError("MSSW 弱口令文件下载失败")

    filepath = os.path.join(save_dir, file_name)
    total = 0
    try:
        with open(filepath, 'wb') as f:
            for chunk in resp.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
                    total += len(chunk)
    finally:
        resp.close()

    # 检查是否为有效xlsx
    with open(filepath, 'rb') as f:
        magic = f.read(4)
    if magic != b'PK\x03\x04':
        with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
            content = f.read(500)
        raise RuntimeError(f"MSSW 下载文件不是有效xlsx（{total}字节），服务器响应: {content}")

    log(f"MSSW 弱口令文件已保存: {os.path.basename(filepath)} ({total / 1024:.1f} KB)")
    return filepath


# ====================================================================
#  Excel 处理
# ====================================================================

# 文件C 固定列头（从左到右）
C_HEADERS = [
    "弱密码名称", "账号", "密码", "url",
    "数据源", "最近发现时间", "首次发现时间", "风险资产",
    "资产类型", "所属资产组", "所属业务", "资产责任人",
    "资产重要性", "资产管理状态", "托管状态", "端口", "refer",
    "互联网暴露", "举证信息", "加白状态", "处置状态", "数据来源",
]


def _proof_to_str(proof) -> str:
    """将 proof 字段（list 或 str）转换为字符串"""
    if isinstance(proof, list):
        return "\n".join(str(p) for p in proof if p)
    return str(proof) if proof else ""


def build_easm_rows(cookie_str: str, company_id: str,
                    time_range: list = None) -> List[dict]:
    """
    按照接口7-1 → 接口7-2 → 详情接口 三级调用，
    构建文件A（EASM 弱口令数据），返回 list[dict]，每个 dict 对应文件C 的一行。
    time_range: [start_ms, end_ms] 或 None/空列表（不过滤）
    """
    # 3a. 获取母表全量数据
    log("  3a. 获取EASM弱口令母表数据（接口7-1）...")
    parent_list = easm_api7_1_weak_pwd_list(cookie_str, company_id)
    log(f"  母表共 {len(parent_list)} 条")

    # 时间过滤：标准化 time_range
    if time_range and len(time_range) >= 2 and time_range[0] and time_range[1]:
        t_start, t_end = int(time_range[0]), int(time_range[1])
    else:
        t_start, t_end = None, None

    rows_out: List[dict] = []
    total_parents = len(parent_list)
    skipped_count = 0

    for p_idx, parent in enumerate(parent_list):
        parent_ip = parent.get('ip', '')

        # 时间过滤：母表 found_time 不在范围内则跳过此母表及其子表
        if t_start is not None:
            found_ms = date_str_to_ms(parent.get('found_time', ''))
            if found_ms is None:
                found_ms = parse_relative_time_to_ms(parent.get('found_time', ''))
            if found_ms is None or found_ms < t_start or found_ms > t_end:
                skipped_count += 1
                continue

        log(f"  3b. [{p_idx+1}/{total_parents}] 获取子表 ip={parent_ip}（接口7-2）...")

        # 3b. 获取子表数据
        sub_list = easm_api7_2_weak_pwd_sub_list(cookie_str, company_id, parent_ip)
        log(f"      子表 {len(sub_list)} 条")

        for sub in sub_list:
            pwd_id = sub.get('pwd_id', '')
            log(f"      获取详情 pwd_id={pwd_id}（补充接口）...")

            # 3c. 获取详情
            detail = {}
            if pwd_id:
                detail = easm_weak_pwd_detail(cookie_str, pwd_id)
                time.sleep(0.2)

            # 详情中的 proof
            proof = detail.get('proof') or sub.get('proof')
            proof_str = _proof_to_str(proof)

            row = {
                "弱密码名称":   sub.get('weak_pwd_type', ''),
                "账号":         sub.get('account', ''),
                "密码":         sub.get('passwd', ''),
                "url":          sub.get('url', ''),
                "数据源":       "bjx",
                "最近发现时间": resolve_parent_found_time(parent),
                "首次发现时间": sub.get('found_time', ''),
                "风险资产":     parent_ip,
                "资产类型":     "",
                "所属资产组":   "",
                "所属业务":     "",
                "资产责任人":   "",
                "资产重要性":   "",
                "资产管理状态": "",
                "端口":         str(sub.get('port', '')),
                "refer":        "",
                "互联网暴露":   "暴露",
                "举证信息":     proof_str,
                "加白状态":     "",
                "处置状态":     EASM_DEAL_STATUS_MAP.get(
                                    sub.get('deal_status', ''),
                                    str(sub.get('deal_status', ''))
                                ),
                "数据来源":     "外网",
                "托管状态":     "",
            }
            rows_out.append(row)

    if skipped_count:
        log(f"  时间过滤：跳过 {skipped_count} 条母表（found_time 不在范围内）")
    log(f"  EASM 数据准备完毕: {len(rows_out)} 行")
    return rows_out


def process_mssw_file(file_b_path: str) -> List[dict]:
    """
    处理 MSSW 导出的弱口令文件B：
    按文件C 的列头从文件B 中提取同名字段。
    返回 list[dict]
    """
    wb = load_workbook(file_b_path, data_only=True)
    try:
        ws = wb.active

        col_map: Dict[str, int] = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value is not None:
                col_map[str(cell.value).strip()] = col_idx

        log(f"MSSW 文件B 列数: {len(col_map)}, 数据行数: {ws.max_row - 1}")
        log(f"  文件B 列名: {list(col_map.keys())}")

        rows_out: List[dict] = []

        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for header in C_HEADERS:
                src_col = col_map.get(header)
                if src_col is not None:
                    val = ws.cell(row=row_idx, column=src_col).value
                    row_data[header] = val if val is not None else ""
                else:
                    row_data[header] = ""
            row_data["数据来源"] = "内网"  # MSSW 固定为内网
            rows_out.append(row_data)
    finally:
        wb.close()

    log(f"MSSW 处理后共 {len(rows_out)} 行")
    return rows_out


def extract_mssw_header_styles(file_b_path: str) -> Tuple[Dict[str, object], object]:
    """
    从 MSSW 导出文件第 1 行表头提取样式信息。
    返回:
      header_styles: {列名 → {'font','fill','border','alignment','number_format'}}
      ref_style:    参考样式（取第 1 个有显式样式的表头单元格），用于额外字段列回退
    异常或空表头时返回 ({}, None)，不抛错。

    注意：不保存 cell._style（那是工作簿内部索引，不可跨簿复制），
    而是保存 Font/Fill/Border/Alignment 等可移植的样式属性对象。
    """
    try:
        wb = load_workbook(file_b_path, data_only=True)
    except Exception as e:
        log(f"  提取MSSW表头样式失败（文件读取失败）: {e}", "WARNING")
        return {}, None

    try:
        ws = wb.active
        header_styles: Dict[str, object] = {}
        ref_style: object = None

        for cell in ws[1]:
            if cell.value is None:
                continue
            if not cell.has_style:          # 跳过无显式样式的列，避免默认样式污染 ref_style
                continue
            col_name = str(cell.value).strip()
            # 保存样式属性对象（可跨工作簿使用），而非 _style（工作簿内部索引）
            style_props = {
                "font":          copy(cell.font),
                "fill":          copy(cell.fill),
                "border":        copy(cell.border),
                "alignment":     copy(cell.alignment),
                "number_format": cell.number_format,  # 字符串不可变，无需 copy
            }
            header_styles[col_name] = style_props
            if ref_style is None:
                ref_style = dict(style_props)       # 浅拷贝 dict，避免与第一列共享同一对象

        if not header_styles:
            log("  MSSW 表头为空或无样式，最终 Excel 表头将使用默认样式", "WARNING")

        return header_styles, ref_style
    except Exception as e:
        log(f"  提取MSSW表头样式失败（遍历异常）: {e}", "WARNING")
        return {}, None
    finally:
        wb.close()


# ====================================================================
#  主流程
# ====================================================================

def _on_script_timeout():
    log(f"错误：脚本执行超时（超过 {SCRIPT_TIMEOUT // 60} 分钟），强制退出", "ERROR")
    os._exit(1)


def main():
    _timer = threading.Timer(SCRIPT_TIMEOUT, _on_script_timeout)
    _timer.daemon = True
    _timer.start()

    parser = argparse.ArgumentParser(
        description='弱口令清单生成工具',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            '示例:\n'
            '  python weak_report.py 深圳市口袋网络\n'
            '  python weak_report.py 35690473\n'
            '  python weak_report.py 客户名 --skip-easm\n'
            '  python weak_report.py 客户名 --skip-mssw\n'
            '  python weak_report.py 客户名 --start-time 1714492800000 --end-time 1782403199000'
        )
    )
    parser.add_argument('keyword', nargs='?', default=None,
                        help='客户ID或客户名称（支持模糊匹配）')
    parser.add_argument('--skip-mssw', action='store_true',
                        help='跳过MSSW平台，只生成EASM数据')
    parser.add_argument('--skip-easm', action='store_true',
                        help='跳过EASM平台，只生成MSSW数据')
    parser.add_argument('--mssw-file', default=None,
                        help='（测试用）指定本地MSSW文件路径，跳过接口调用')
    parser.add_argument('--start-time', default=None,
                        help='（可选）过滤开始日期，如 2016-01-01 或 2016年1月1日')
    parser.add_argument('--end-time', default=None,
                        help='（可选）过滤结束日期，如 2026-06-30 或 2026年6月30日')
    args = parser.parse_args()

    if not args.skip_easm and not args.keyword:
        sys.exit("错误：请提供客户ID或客户名称关键词\n示例: python weak_report.py 35690473")

    os.makedirs(TEMP_DIR, exist_ok=True)

    # 时间过滤：传入则过滤，不传则 []（全部数据）
    if args.start_time and args.end_time:
        start_ms = parse_date_to_ms(args.start_time, is_end=False)
        end_ms   = parse_date_to_ms(args.end_time,   is_end=True)
        time_range = [start_ms, end_ms]
        log(f"时间过滤: {args.start_time} → {start_ms}  /  {args.end_time} → {end_ms}")
    else:
        time_range = []

    # ==================== 1. 加载 Cookie ====================
    log("=" * 50)
    log("步骤1：加载 Cookie")

    easm_cookie = ""
    if not args.skip_easm:
        if not os.path.exists(EASM_COOKIES_FILE):
            sys.exit(f"错误：EASM cookies文件不存在 → {EASM_COOKIES_FILE}")
        easm_cookie = read_cookies_as_string(EASM_COOKIES_FILE)
        if not easm_cookie:
            sys.exit(f"错误：EASM cookies解析结果为空 → {EASM_COOKIES_FILE}")
        log(f"  EASM Cookie: {EASM_COOKIES_FILE}")

    mssw_cookie = ""
    if not args.skip_mssw and not args.mssw_file:
        if not os.path.exists(MSSW_COOKIES_FILE):
            sys.exit(f"错误：MSSW cookies文件不存在 → {MSSW_COOKIES_FILE}")
        mssw_cookie = read_cookies_as_string(MSSW_COOKIES_FILE)
        if not mssw_cookie:
            sys.exit(f"错误：MSSW cookies解析结果为空 → {MSSW_COOKIES_FILE}")
        log(f"  MSSW Cookie: {MSSW_COOKIES_FILE}")
    elif args.mssw_file:
        log(f"  MSSW Cookie: 已跳过（使用本地文件）")
    else:
        log(f"  MSSW Cookie: 已跳过（--skip-mssw）")

    # ==================== 2. 搜索客户（分平台独立搜索） ====================
    log("=" * 50)
    log("步骤2：搜索客户")

    # 2a. EASM 客户搜索
    easm_company_id = ""
    if not args.skip_easm:
        easm_keyword = args.keyword + "[影子]"
        log(f"  步骤2a：EASM 搜索客户「{easm_keyword}」...")
        easm_customers = api0_search_customer(easm_cookie, easm_keyword)
        if not easm_customers:
            sys.exit(f"错误：EASM 未找到匹配客户，请检查关键词「{easm_keyword}」")
        if len(easm_customers) > 1:
            exact = _pick_exact_match(easm_customers, easm_keyword)
            if exact:
                easm_customers = [exact]
            else:
                log(f"错误：EASM 找到 {len(easm_customers)} 个匹配客户，请使用更精确的关键词：", "ERROR")
                for c in easm_customers:
                    name = c.get('company_name') or c.get('pms_customer_name', '未知')
                    print(f"  ID={c['company_id']}  名称={name}")
                sys.exit(1)
        easm_company_id = easm_customers[0]['company_id']

    # 2b. MSSW 客户搜索
    mssw_company_id = ""
    if not args.skip_mssw and not args.mssw_file:
        log(f"  步骤2b：MSSW 搜索客户「{args.keyword}」...")
        mssw_customers = mssw_api0_search_customer(mssw_cookie, args.keyword)
        if not mssw_customers:
            sys.exit(f"错误：MSSW 未找到匹配客户，请检查关键词「{args.keyword}」")
        if len(mssw_customers) > 1:
            exact = _pick_exact_match(mssw_customers, args.keyword)
            if exact:
                mssw_customers = [exact]
            else:
                log(f"错误：MSSW 找到 {len(mssw_customers)} 个匹配客户，请使用更精确的关键词：", "ERROR")
                for c in mssw_customers:
                    name = c.get('company_name') or c.get('pms_customer_name', '未知')
                    print(f"  ID={c['company_id']}  名称={name}")
                sys.exit(1)
        mssw_company_id = mssw_customers[0]['company_id']

    # ==================== 3. MSSW 平台处理 ====================
    log("=" * 50)
    file_b_path = ""  # 预声明，确保步骤5 可访问（--skip-mssw 时保持为空）
    if args.skip_mssw:
        log("步骤3：MSSW 平台（内网）已跳过（--skip-mssw）")
        mssw_data: List[dict] = []
    else:
        log("步骤3：MSSW 平台（内网）弱口令数据获取")

        if args.mssw_file:
            # 测试用：直接读取本地文件
            file_b_path = args.mssw_file
            if not os.path.exists(file_b_path):
                sys.exit(f"错误：指定的MSSW文件不存在 → {file_b_path}")
            log(f"  [测试用] 使用本地文件: {file_b_path}")
        else:
            # 3a. 接口9：触发导出
            log("  3a. 触发MSSW弱口令导出...")
            mssw_file_name = mssw_api9_export_weak_pwd(mssw_cookie, mssw_company_id, time_range)

            # 3b. 接口10：下载文件B
            log("  3b. 下载弱口令文件...")
            file_b_raw = mssw_api10_download(mssw_cookie, mssw_company_id, mssw_file_name, TEMP_DIR)
            # 文件名已是"弱密码-脆弱性导出报告_导出时间.xlsx"格式，直接使用
            file_b_path = file_b_raw
            log(f"  文件B: {os.path.basename(file_b_path)}")

        # 3c. 处理文件B → 统一格式
        log("  3c. 处理MSSW数据...")
        mssw_data = process_mssw_file(file_b_path)
        log(f"  MSSW 数据准备完毕: {len(mssw_data)} 行")

    # ==================== 4. EASM 平台处理 ====================
    log("=" * 50)
    if args.skip_easm:
        log("步骤4：EASM 平台（外网）已跳过（--skip-easm）")
        easm_data: List[dict] = []
    else:
        log("步骤4：EASM 平台（外网）弱口令数据获取")
        easm_data = build_easm_rows(easm_cookie, easm_company_id, time_range)

    # ==================== 5. 合并输出 ====================
    log("=" * 50)
    log("步骤5：合并生成最终文件")

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "弱口令"

    # 提取 MSSW 表头样式（仅当 MSSW 文件可用时）
    header_styles: Dict[str, object] = {}
    ref_style: object = None
    if not args.skip_mssw and file_b_path and os.path.exists(file_b_path):
        log("  提取 MSSW 表头样式...")
        header_styles, ref_style = extract_mssw_header_styles(file_b_path)
        log(f"  已提取 {len(header_styles)} 个列样式，参考样式={'有' if ref_style else '无'}")

    # 写入表头并应用样式
    for ci, header in enumerate(C_HEADERS, start=1):
        cell = ws_out.cell(row=1, column=ci, value=header)
        style_props = header_styles.get(header)
        if style_props is None:
            style_props = ref_style
        if style_props is not None:
            try:
                cell.font          = copy(style_props["font"])
                cell.fill          = copy(style_props["fill"])
                cell.border        = copy(style_props["border"])
                cell.alignment     = copy(style_props["alignment"])
                cell.number_format = style_props["number_format"]
            except Exception as e:
                log(f"  应用表头样式失败 列={header}: {e}", "WARNING")

    # 先写 MSSW 数据，再写 EASM 数据
    row_num = 2
    for record in mssw_data:
        for ci, header in enumerate(C_HEADERS, start=1):
            ws_out.cell(row=row_num, column=ci, value=record.get(header, ""))
        row_num += 1

    for record in easm_data:
        for ci, header in enumerate(C_HEADERS, start=1):
            ws_out.cell(row=row_num, column=ci, value=record.get(header, ""))
        row_num += 1

    wb_out.save(OUTPUT_FILE)
    wb_out.close()

    total_rows = len(mssw_data) + len(easm_data)
    _timer.cancel()
    log(f"完成！")
    log(f"  MSSW  数据: {len(mssw_data)} 行")
    log(f"  EASM 数据: {len(easm_data)} 行")
    log(f"  合计: {total_rows} 行")
    log(f"  输出文件: {OUTPUT_FILE}")


if __name__ == '__main__':
    main()

# ============================================================
#  使用说明
# ============================================================
#
#  python weak_report.py <客户ID或关键词> [选项]
#
#  可选参数:
#     --skip-easm              跳过EASM平台（外网），只获取MSSW内网数据
#     --skip-mssw               跳过MSSW平台（内网），只获取EASM外网数据
#     --mssw-file <路径>        测试用，直接读取本地MSSW文件
#     --start-time <时间戳>    过滤开始时间（13位UTC毫秒时间戳）
#     --end-time   <时间戳>    过滤结束时间（13位UTC毫秒时间戳）
#
#  示例:
#     python weak_report.py 35690473                                      # 两个平台一起
#     python weak_report.py 客户名 --skip-easm                            # 仅 MSSW
#     python weak_report.py 客户名 --skip-mssw                             # 仅 EASM
#     python weak_report.py 客户名 --skip-mssw --start-time x --end-time y # 仅EASM+时间
#     python weak_report.py 客户名 --skip-easm --start-time x --end-time y # 仅MSSW+时间
#
#  输出:
#     弱口令清单.xlsx，sheet名"弱口令"
#
#  过滤逻辑:
#     EASM 平台: deal_status=[0(未审核),1(处置中),6(待复测),7(修复失败)]
#                found_time 内存过滤（UTC+8→UTC毫秒时间戳比较）
#     MSSW 平台:  fixed_status=[0(待处置),1(处置中)]
#                latest_time_range 接口层时间过滤
#
# ============================================================
