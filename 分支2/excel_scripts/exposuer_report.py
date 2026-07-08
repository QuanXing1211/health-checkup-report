#!/usr/bin/env python3
# -*- coding: utf-8 -*-


import os
import sys
import time
import threading
import zipfile
import argparse
import requests
from copy import copy
from datetime import datetime
from typing import Dict, Optional
import openpyxl
from openpyxl import load_workbook

# ==================== 配置（按需修改） ====================
COOKIES_FILE    = r"C:\Users\User\Downloads\cookies.txt"
BASE_URL        = "https://soar59.sangfor.com.cn"
TEMP_DIR        = r"C:\Users\User\Downloads\temp_report"
OUTPUT_FILE     = r"C:\Users\User\Downloads\temp_report\暴露面清单.xlsx"
POLL_INTERVAL  = 5     # 轮询间隔（秒）
REPORT_LIMIT   = 100   # 接口3每次查询的报告数量
SCRIPT_TIMEOUT = 3600  # 全局超时：1小时
MAX_RETRIES    = 3     # 最大重试次数
RETRY_DELAY    = 3     # 重试等待时间（秒）
# ==========================================================


# ---------- 日志 ----------

def log(msg: str, level: str = "INFO") -> None:
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}][{level}] {msg}")


# ---------- 工具函数 ----------

def _get_system_timezone() -> str:
    offset = datetime.now().astimezone().utcoffset()
    if offset is None:
        return "+00:00"
    total_seconds = int(offset.total_seconds())
    sign = "+" if total_seconds >= 0 else "-"
    total_seconds = abs(total_seconds)
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    return f"{sign}{hours:02d}:{minutes:02d}"


def extract_cookie_value(cookie_str: str, name: str) -> Optional[str]:
    for part in cookie_str.split(';'):
        part = part.strip()
        if '=' in part:
            k, _, v = part.partition('=')
            if k.strip() == name:
                return v.strip()
    return None


# ---------- Cookie 读取 ----------

def read_cookies_as_string(filepath: str) -> str:
    """
    读取 cookies.txt，返回 Cookie 请求头字符串（name=value; name2=value2 格式）。
    支持两种输入格式：
      1. 浏览器原始 Cookie 字符串
      2. Netscape 格式（EditThisCookie 等工具导出，tab 分隔）
    """
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read().strip()

    lines = [ln.strip() for ln in content.splitlines() if ln.strip()]
    is_netscape = any('\t' in ln for ln in lines if not ln.startswith('#'))

    if is_netscape:
        pairs = []
        for ln in lines:
            if ln.startswith('#'):
                continue
            parts = ln.split('\t')
            if len(parts) >= 7:
                pairs.append(f"{parts[5]}={parts[6]}")
        return '; '.join(pairs)
    else:
        return '; '.join(lines)


# ---------- 请求封装 ----------

DEFAULT_HEADERS: Dict[str, str] = {
    "Accept": "application/json, text/javascript, */*; q=0.01",
    "Accept-Language": "zh-CN,zh;q=0.9",
    "Content-Type": "application/json",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "X-Requested-With": "XMLHttpRequest",
    "Referer": f"{BASE_URL}/index.html",
    "timezone": _get_system_timezone(),
}


def request_with_retry(method: str, url: str, headers: Optional[Dict] = None,
                       timeout: int = 60, **kwargs) -> Optional[requests.Response]:
    if headers is None:
        headers = DEFAULT_HEADERS.copy()

    cookie = headers.get("Cookie", "")
    if cookie:
        csrf_token = extract_cookie_value(cookie, "csrf_token")
        if csrf_token:
            headers["X-Csrftoken"] = csrf_token

    session = requests.Session()
    adapter = requests.adapters.HTTPAdapter(
        pool_connections=10, pool_maxsize=10, max_retries=0
    )
    session.mount('https://', adapter)
    session.mount('http://', adapter)

    for attempt in range(MAX_RETRIES):
        try:
            if method.upper() == "POST":
                resp = session.post(url, headers=headers, timeout=timeout,
                                    verify=False, **kwargs)
            else:
                resp = session.get(url, headers=headers, timeout=timeout,
                                   verify=False, **kwargs)
            return resp
        except (requests.exceptions.Timeout, requests.exceptions.ConnectionError,
                requests.exceptions.SSLError, requests.exceptions.RequestException) as e:
            attempt_num = attempt + 1
            log(f"请求失败 (尝试 {attempt_num}/{MAX_RETRIES}) - {e}", "WARNING")
            if attempt < MAX_RETRIES - 1:
                log(f"等待 {RETRY_DELAY}s 后重试...", "INFO")
                time.sleep(RETRY_DELAY)
            else:
                log(f"重试 {MAX_RETRIES} 次后仍然失败，返回 None", "ERROR")
                return None
    return None


def _build_headers(cookie_str: str) -> Dict[str, str]:
    headers = DEFAULT_HEADERS.copy()
    headers["Cookie"] = cookie_str
    return headers


def _parse_json(resp: requests.Response, api_name: str) -> dict:
    if resp is None:
        raise RuntimeError(f"{api_name}：请求失败（已重试 {MAX_RETRIES} 次）")
    try:
        return resp.json()
    except Exception:
        raise RuntimeError(
            f"{api_name}：响应不是JSON（状态码={resp.status_code}，可能Cookie已过期）"
        )


# ---------- 接口调用 ----------

def api0_search_customer(headers: Dict, keyword: str) -> list:
    """接口0：根据关键词模糊搜索客户，返回列表"""
    url = f"{BASE_URL}/gateway/customer-mgr-service/order/v1/user?_method=GET"
    payload = {
        "order": "asc", "offset": 0, "limit": 20, "keyword": keyword,
        "share_ids": [], "delivery_channel_id": [], "service_code": [],
        "industry": [], "industry_segmentation": [], "customer_type": [],
        "customer_stratification": [], "protection_type": [], "service_group": [],
        "delivery_method": [], "platform_type": [], "service_status": 0, "my_customer": 0,
    }
    resp = request_with_retry("POST", url, headers=headers, json=payload)
    data = _parse_json(resp, "接口0")
    if data.get('code') != 0:
        raise RuntimeError(f"接口0失败: {data.get('msg')}")
    return data['data']['list']


def _pick_exact_match(customers: list, keyword: str):
    """从多个模糊搜索结果中优先选精确匹配。返回匹配项或 None"""
    if len(customers) == 1:
        return customers[0]
    exact = [c for c in customers
             if (c.get('company_name', '') or '').strip() == keyword.strip()
             or (c.get('pms_customer_name', '') or '').strip() == keyword.strip()
             or str(c.get('company_id', '')).strip() == keyword.strip()]
    return exact[0] if exact else None


def api1_get_template(headers: Dict) -> tuple:
    """接口1：获取报告模板列表，返回 (template_id, template_name)"""
    url = f"{BASE_URL}/order/v1/report/template_list"
    resp = request_with_retry("POST", url, headers=headers, json={"template_format": ""})
    data = _parse_json(resp, "接口1")
    if data.get('code') != 0:
        raise RuntimeError(f"接口1失败: {data.get('msg')}")

    for tpl in data['data']['template_list']:
        if (tpl.get('template_source') == 'easm'
                and tpl.get('easm_report_type') == 1
                and tpl.get('template_format') == 'excel'):
            log(f"模板: {tpl['template_name']} (id={tpl['template_id']})")
            return tpl['template_id'], tpl['template_name']

    raise RuntimeError("接口1：未找到「EASM单次服务成果清单」Excel模板")


def api2_generate_report(headers: Dict, company_id: str,
                         template_id: str, template_name: str) -> str:
    """接口2：触发报告生成，返回 task_id"""
    url = f"{BASE_URL}/order/v1/report/generate_easm_report"
    payload = {
        "customer_id": company_id,
        "target_company_id": ["all"],
        "need_split": False,
        "template_id": template_id,
        "template_name": template_name,
        "params": {}, "page_data": {}, "reports": [],
    }
    resp = request_with_retry("POST", url, headers=headers, json=payload)
    data = _parse_json(resp, "接口2")
    if data.get('code') != 0:
        raise RuntimeError(f"接口2失败: {data.get('msg')}")
    task_id = data['data']['_id']
    log(f"task_id={task_id}")
    return task_id


def api3_poll_status(headers: Dict, task_id: str) -> None:
    """接口3：轮询报告生成状态，直到 task_status=1（成功）"""
    url = f"{BASE_URL}/order/v1/report/report_status"
    payload = {
        "order": "desc", "offset": 0, "limit": REPORT_LIMIT,
        "keyword": "", "start_time": "", "end_time": "",
        "template_name": ["EASM单次服务成果清单"],
    }

    attempt = 0
    while True:
        attempt += 1
        resp = request_with_retry("POST", url, headers=headers, json=payload)
        data = _parse_json(resp, "接口3")
        if data.get('code') != 0:
            raise RuntimeError(f"接口3失败: {data.get('msg')}")

        matched = next(
            (item for item in data['data']['list'] if item.get('task_id') == task_id),
            None
        )
        if matched:
            status = matched.get('task_status')
            if status == 1:
                log(f"第{attempt}次轮询：报告生成成功")
                return
            elif status == 2:
                raise RuntimeError("报告生成失败（task_status=2）")
            else:
                log(f"第{attempt}次轮询：生成中，{POLL_INTERVAL}s后重试...")
        else:
            log(f"第{attempt}次轮询：任务尚未出现，{POLL_INTERVAL}s后重试...")

        time.sleep(POLL_INTERVAL)


def api4_download_report(headers: Dict, task_id: str, save_dir: str) -> str:
    """接口4：下载报告压缩包，返回本地 zip 路径"""
    url = f"{BASE_URL}/order/v1/report/report_download?task_id={task_id}"
    # 下载用流式GET，stream参数通过kwargs传入
    resp = request_with_retry("GET", url, headers=headers, stream=True)
    if resp is None:
        raise RuntimeError("接口4：下载报告失败")

    content_disp = resp.headers.get('Content-Disposition', '')
    if 'filename=' in content_disp:
        filename = content_disp.split('filename=')[-1].strip().strip('"\'')
    else:
        filename = f"easm_report_{task_id}.zip"

    zip_path = os.path.join(save_dir, filename)
    total = 0
    with open(zip_path, 'wb') as f:
        for chunk in resp.iter_content(chunk_size=8192):
            if chunk:
                f.write(chunk)
                total += len(chunk)

    log(f"已保存: {os.path.basename(zip_path)} ({total / 1024:.1f} KB)")
    return zip_path


# ---------- 文件处理 ----------

def _extract_zip_fix_encoding(zip_path: str, extract_dir: str) -> None:
    """解压 zip，自动修复 Windows 下 GBK 编码的文件名乱码。"""
    with zipfile.ZipFile(zip_path, 'r') as zf:
        for info in zf.infolist():
            # flag_bits bit 11 = 1 表示文件名已是 UTF-8，否则尝试 GBK 解码
            if info.flag_bits & 0x800:
                fname = info.filename
            else:
                try:
                    fname = info.filename.encode('cp437').decode('gbk')
                except (UnicodeDecodeError, UnicodeEncodeError):
                    fname = info.filename
            target = os.path.join(extract_dir, fname)
            if fname.endswith('/') or fname.endswith('\\'):
                os.makedirs(target, exist_ok=True)
            else:
                os.makedirs(os.path.dirname(target), exist_ok=True)
                with zf.open(info) as src, open(target, 'wb') as dst:
                    dst.write(src.read())


def extract_report(zip_path: str, extract_dir: str) -> tuple:
    """解压报告zip，返回 (外部风险路径, 资产台账路径)"""
    _extract_zip_fix_encoding(zip_path, extract_dir)

    file_b = file_c = None
    for root, _, files in os.walk(extract_dir):
        for fname in files:
            if not fname.endswith('.xlsx'):
                continue
            path = os.path.join(root, fname)
            if '外部风险' in fname:
                file_b = path
            elif '资产台账' in fname:
                file_c = path

    if not file_b:
        raise FileNotFoundError(f"解压后未找到「外部风险」文件，解压目录: {extract_dir}")
    if not file_c:
        raise FileNotFoundError(f"解压后未找到「资产台账」文件，解压目录: {extract_dir}")

    log(f"外部风险(B): {os.path.basename(file_b)}")
    log(f"资产台账(C): {os.path.basename(file_c)}")
    return file_b, file_c


def _copy_sheet(src_ws, dst_wb: openpyxl.Workbook, name: str):
    """将 src_ws 的数据、样式、列宽行高、合并单元格复制到 dst_wb 的新 sheet"""
    # --- 诊断日志：源 sheet 基本信息 ---
    log(f"  [copy] 源sheet={src_ws.title!r}  dimensions={src_ws.dimensions}"
        f"  max_row={src_ws.max_row}  max_col={src_ws.max_column}"
        f"  sheet_state={src_ws.sheet_state}")
    # # 打印前3行前5列的实际值，确认数据是否存在
    # preview_rows = 0
    # for row in src_ws.iter_rows(min_row=1, max_row=3, max_col=5):
    #     vals = [repr(cell.value) for cell in row]
    #     log(f"  [copy] 预览行{row[0].row}: {vals}")
    #     preview_rows += 1
    # if preview_rows == 0:
    #     log(f"  [copy] 警告：iter_rows 未遍历到任何行，源sheet可能为空", "WARNING")

    dst_ws = dst_wb.create_sheet(title=name)

    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width
        dst_ws.column_dimensions[col_letter].hidden = dim.hidden
    for row_idx, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_idx].height = dim.height
        dst_ws.row_dimensions[row_idx].hidden = dim.hidden

    # 第一次遇到某种样式时做完整赋值，缓存目标 workbook 已注册的 _style 索引数组；
    # 后续相同样式直接复制索引数组，绕过6次属性 setter 的注册开销
    style_cache = {}  # src _style key → dst _style array
    cell_count = 0
    for row in src_ws.iter_rows():
        for cell in row:
            dst_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.value is not None:
                cell_count += 1
            if cell.has_style:
                key = tuple(cell._style)
                if key not in style_cache:
                    dst_cell.font          = copy(cell.font)
                    dst_cell.border        = copy(cell.border)
                    dst_cell.fill          = copy(cell.fill)
                    dst_cell.number_format = cell.number_format
                    dst_cell.alignment     = copy(cell.alignment)
                    dst_cell.protection    = copy(cell.protection)
                    style_cache[key] = copy(dst_cell._style)
                else:
                    dst_cell._style = copy(style_cache[key])

    for merge_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merge_range))

    log(f"  [copy] 复制完成 → 目标sheet={name!r}，非空单元格数={cell_count}")
    return dst_ws


def _delete_columns_by_header(ws, names_to_delete: list) -> int:
    to_delete = [
        cell.column for cell in ws[1]
        if cell.value in names_to_delete
    ]
    for col_idx in sorted(to_delete, reverse=True):
        ws.delete_cols(col_idx)
    return len(to_delete)


# ---------- 时间过滤辅助函数 ----------

def _parse_arg_time(s: str, is_end: bool = False) -> datetime:
    """将命令行时间参数字符串解析为 datetime。纯日期的结束时间默认取当天末尾。"""
    s = s.strip()
    s = s.replace('年', '-').replace('月', '-').replace('日', '')  # 支持 2026年1月1日 格式
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S', '%Y-%m-%d', '%Y/%m/%d'):
        try:
            dt = datetime.strptime(s, fmt)
            if is_end and fmt in ('%Y-%m-%d', '%Y/%m/%d'):
                dt = dt.replace(hour=23, minute=59, second=59)
            return dt
        except ValueError:
            continue
    raise ValueError(f"无法解析时间参数 {s!r}，支持格式：2025-01-01 或 2025-01-01 00:00:00")


def _parse_cell_time(val) -> Optional[datetime]:
    """将单元格值解析为 datetime，无法解析返回 None。"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        s = val.strip()
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S', 
                    '%Y-%m-%d %H:%M', '%Y/%m/%d %H:%M',  # 新增不带秒的格式
                    '%Y-%m-%d', '%Y/%m/%d'):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
    return None


def _find_col_idx(ws, header: str) -> Optional[int]:
    """在第1行查找列标题（精确匹配），返回1-based列号，未找到返回 None。"""
    for cell in ws[1]:
        if cell.value is not None and str(cell.value).strip() == header:
            return cell.column
    return None


def _build_single_key_lookup(ws, key_col: str,
                              time_col: str = '最近更新时间',
                              connectivity_col: Optional[str] = None) -> Dict[str, Optional[datetime]]:
    """从 ws 构建 {key_col值: 最近更新时间} 字典（同一键取最后出现行的时间）。"""
    kid = _find_col_idx(ws, key_col)
    tid = _find_col_idx(ws, time_col)
    if kid is None or tid is None:
        log(f"  [lookup] {ws.title!r}: 未找到列 {key_col!r} 或 {time_col!r}", "WARNING")
        return {}
    cid = _find_col_idx(ws, connectivity_col) if connectivity_col else None
    result: Dict[str, Optional[datetime]] = {}
    not_open_count = 0
    for row in ws.iter_rows(min_row=2):
        if cid is not None:
            cv = row[cid - 1].value
            if cv is None or str(cv).strip() != '开放':
                not_open_count += 1
                continue
        kv = row[kid - 1].value
        if kv is not None:
            result[str(kv).strip()] = _parse_cell_time(row[tid - 1].value)
    print("_build_single_key_lookup not open count:", not_open_count)
    return result


def _build_pair_key_lookup(ws, host_col: str, port_col: str,
                           time_col: str = '最近更新时间',
                           connectivity_col: Optional[str] = None) -> Dict[tuple, Optional[datetime]]:
    """从 ws 构建 {(host, port): 最近更新时间} 字典（同一键取最后出现行的时间）。"""
    hid = _find_col_idx(ws, host_col)
    pid = _find_col_idx(ws, port_col)
    tid = _find_col_idx(ws, time_col)
    if hid is None or pid is None or tid is None:
        log(f"  [lookup] {ws.title!r}: 未找到列 {host_col!r}/{port_col!r} 或 {time_col!r}", "WARNING")
        return {}
    cid = _find_col_idx(ws, connectivity_col) if connectivity_col else None
    result: Dict[tuple, Optional[datetime]] = {}
    not_open_count = 0
    for row in ws.iter_rows(min_row=2):
        if cid is not None:
            cv = row[cid - 1].value
            if cv is None or str(cv).strip() != '开放':
                not_open_count += 1
                continue
        hv = row[hid - 1].value
        pv = row[pid - 1].value
        if hv is not None and pv is not None:
            result[(str(hv).strip(), str(pv).strip())] = _parse_cell_time(row[tid - 1].value)

    print("_build_single_key_lookup not open count:", not_open_count)
    return result


def _in_range(t: Optional[datetime],
              start: Optional[datetime], end: Optional[datetime]) -> bool:
    """判断 t 是否在 [start, end] 内。t=None 时视为不在范围内。"""
    if t is None:
        return False
    if start and t < start:
        return False
    if end and t > end:
        return False
    return True


def _remap_row(r: int, deleted_sorted: list) -> int:
    """原始行号 r → 删除 deleted_sorted 中各行后的新行号。"""
    return r - sum(1 for d in deleted_sorted if d < r)


def _delete_out_of_range_rows(ws, time_getter, start: Optional[datetime],
                               end: Optional[datetime],
                               debug_wb: Optional[openpyxl.Workbook] = None) -> int:
    """
    逐行检查（跳过第1行标题），用 time_getter(row_cells) 取时间，
    删除时间不在范围内的行。保留合并单元格样式。返回删除行数。
    """
    to_del = set()
    null_time_rows_data = []

    # 公有云资产有跨列合并，在副本上解除合并再遍历，避免 time_getter 取到 MergedCell，原 ws 不受影响
    import copy as _copy_mod
    prefix = ws.title.split('（')[0].split('(')[0].strip()
    if prefix == '公有云资产':
        tmp_ws = _copy_mod.deepcopy(ws)
        for r in list(tmp_ws.merged_cells.ranges):
            min_row, min_col, max_row, max_col = r.min_row, r.min_col, r.max_row, r.max_col
            anchor_val = tmp_ws.cell(row=min_row, column=min_col).value
            tmp_ws.unmerge_cells(str(r))
            for row_idx in range(min_row, max_row + 1):
                for col_idx in range(min_col, max_col + 1):
                    tmp_ws.cell(row=row_idx, column=col_idx).value = anchor_val
        # tmp_save_wb = openpyxl.Workbook()
        # tmp_save_ws = tmp_save_wb.active
        # tmp_save_ws.title = tmp_ws.title
        # for row in tmp_ws.iter_rows():
        #     for cell in row:
        #         tmp_save_ws.cell(row=cell.row, column=cell.column, value=cell.value)
        # tmp_save_path = os.path.join(TEMP_DIR, 'debug_公有云资产_unmerged.xlsx')
        # tmp_save_wb.save(tmp_save_path)
        # log(f"  [debug] 公有云资产副本已保存: {tmp_save_path}")
        iter_ws = tmp_ws
    else:
        iter_ws = ws

    for row in iter_ws.iter_rows(min_row=2):
        t = time_getter(row)
        # if ws.title == "公有云资产":
        #     print(t)
        if t is None and debug_wb is not None:
            null_time_rows_data.append([cell.value for cell in row])
        if not _in_range(t, start, end):
            to_del.add(row[0].row)
    # if ws.title == "公有云资产":
    #     print(null_time_rows_data)
    if not to_del and not null_time_rows_data:
        return 0

    # 表名带(0)或（0）说明数据为空，跳过 debug 写入，但仍需执行后续的 to_del 删除逻辑
    skip_debug = ('(0)' in ws.title or '（0）' in ws.title)
    if null_time_rows_data and debug_wb is not None and not skip_debug:
        safe_name = ws.title.split('（')[0].split('(')[0].strip()[:28].replace('/', '_').replace('\\', '_')
        sheet_name = f"null_{safe_name}"
        dbg_ws = debug_wb.create_sheet(title=sheet_name)
        # 写入表头（第1行）
        header_vals = [cell.value for cell in iter_ws[1]]
        dbg_ws.append(['' if v is None else v for v in header_vals])
        # 写入时间为空的数据行
        for row_vals in null_time_rows_data:
            dbg_ws.append(['' if v is None else v for v in row_vals])
        log(f"  [debug] 时间为空的行已写入 debug sheet: {sheet_name!r}")

    if not to_del:
        return 0
    # 解除合并后将值和样式填入区域内每一行，避免起始行被删导致值丢失
    saved_merges = [
        (r.min_row, r.min_col, r.max_row, r.max_col)
        for r in ws.merged_cells.ranges
    ]
    for r in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(r))
    for min_row, min_col, max_row, max_col in saved_merges:
        anchor = ws.cell(row=min_row, column=min_col)
        anchor_val   = anchor.value
        anchor_style = copy(anchor._style) if anchor.has_style else None
        for row_idx in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = anchor_val
                if anchor_style is not None:
                    cell._style = copy(anchor_style)
    # 原地压缩：把保留行依次写到前面，最后一次性删除末尾多余行（O(n) 替代逐行删除的 O(n²)）
    max_col = ws.max_column
    keep_rows = [r for r in range(2, ws.max_row + 1) if r not in to_del]
    for new_idx, old_idx in enumerate(keep_rows, start=2):
        if new_idx == old_idx:
            continue
        for col in range(1, max_col + 1):
            src = ws.cell(row=old_idx, column=col)
            dst = ws.cell(row=new_idx, column=col)
            dst.value  = src.value
            dst._style = copy(src._style)

    first_extra = len(keep_rows) + 2
    trailing = ws.max_row - first_extra + 1
    if trailing > 0:
        ws.delete_rows(first_extra, trailing)
    # 按新行号重新映射并恢复合并区域
    row_map = {old: new for new, old in enumerate(keep_rows, start=2)}
    for min_row, min_col, max_row, max_col in saved_merges:
        surviving = [r for r in range(min_row, max_row + 1) if r not in to_del]
        if not surviving:
            continue
        new_min = row_map.get(surviving[0], surviving[0])
        new_max = row_map.get(surviving[-1], surviving[-1])
        if new_min == new_max and min_col == max_col:
            continue
        ws.merge_cells(start_row=new_min, start_column=min_col,
                       end_row=new_max, end_column=max_col)
    # 第一列序号从1重新排列
    for i, row in enumerate(ws.iter_rows(min_row=2, min_col=1, max_col=1), start=1):
        row[0].value = i
    return len(to_del)


def _apply_time_filter(ws, start: Optional[datetime], end: Optional[datetime],
                       rule: str, debug_wb: Optional[openpyxl.Workbook] = None,
                       **kwargs) -> None:
    """
    对已复制到目标 workbook 的 ws 应用时间过滤。
    rule:
      'direct'       — 直接读取本表【最近更新时间】列
      'host_lookup'  — 通过单键关联另一张表（kwargs: key_col, lookup）
      'pair_lookup'  — 通过双键 (host+port) 关联端口表（kwargs: host_col, port_col, lookup）
    """
    if start is None and end is None:
        return

    deleted = 0

    if rule == 'direct':
        tid = _find_col_idx(ws, '最近更新时间')
        if tid is None:
            log(f"  [filter] {ws.title!r}: 未找到【最近更新时间】，跳过", "WARNING")
            return
        deleted = _delete_out_of_range_rows(
            ws, lambda row: _parse_cell_time(row[tid - 1].value), start, end, debug_wb)

    elif rule == 'host_lookup':
        key_col = kwargs['key_col']
        lookup  = kwargs['lookup']
        kid = _find_col_idx(ws, key_col)
        if kid is None:
            log(f"  [filter] {ws.title!r}: 未找到列【{key_col}】，跳过", "WARNING")
            return
        deleted = _delete_out_of_range_rows(
            ws,
            lambda row: lookup.get(str(row[kid - 1].value).strip())
                        if row[kid - 1].value is not None else None,
            start, end, debug_wb)

    elif rule == 'pair_lookup':
        host_col = kwargs['host_col']
        port_col = kwargs['port_col']
        lookup   = kwargs['lookup']
        hid = _find_col_idx(ws, host_col)
        pid = _find_col_idx(ws, port_col)
        if hid is None or pid is None:
            log(f"  [filter] {ws.title!r}: 未找到列【{host_col}】或【{port_col}】，跳过", "WARNING")
            return
        def _pair_getter(row):
            hv = row[hid - 1].value
            pv = row[pid - 1].value
            if hv is None or pv is None:
                return None
            return lookup.get((str(hv).strip(), str(pv).strip()))
        deleted = _delete_out_of_range_rows(ws, _pair_getter, start, end, debug_wb)

    log(f"  [filter] {ws.title!r}: rule={rule}，删除 {deleted} 行")


# ---------- 报告构建 ----------

def build_output_excel(file_b: str, file_c: str, output_path: str,
                       start_dt: Optional[datetime] = None,
                       end_dt: Optional[datetime] = None,
                       debug_mode: bool = False) -> None:
    """
    生成暴露面清单：
      C（资产台账）所有sheet + 总表去掉「风险数量」「网站监测授权」列
      B（外部风险）「重要组件」→「Web服务风险分布」
                   「高危端口&服务」→「非Web服务风险分布」

    start_dt / end_dt 不为 None 时，对各 sheet 按最近更新时间进行数据过滤：
      - 根域名/SSL证书/公众号&小程序资产/APP资产：直接读本表【最近更新时间】
      - 子域名：通过【子域名】→资产总表【Host（IP/子域名）】取时间
      - IP C段：通过【对应主机】→资产总表【Host（IP/子域名）】取时间
      - 公有云资产：通过【IP地址/子域名】→资产总表【Host（IP/子域名）】取时间
      - 登录入口/网络&安全设备/WEB资产/Web服务风险分布：通过【访问路径】→端口表【访问路径】取时间
      - 非WEB资产/非Web服务风险分布：通过【IP地址/子域名】+【端口】→端口表【Host】+【端口】取时间
    """
    wb_c = load_workbook(file_c, data_only=True)
    wb_b = load_workbook(file_b, data_only=True)
    wb_new = openpyxl.Workbook()
    wb_new.remove(wb_new.active)

    debug_wb: Optional[openpyxl.Workbook] = None
    if debug_mode:
        debug_wb = openpyxl.Workbook()
        debug_wb.remove(debug_wb.active)

    def _find_sheet(wb, prefix: str) -> Optional[str]:
        """按前缀匹配sheet名，忽略末尾的中文括号数字，如「重要组件（0）」"""
        for name in wb.sheetnames:
            stripped = name.split('（')[0].split('(')[0].strip()
            if stripped == prefix:
                return name
        return None

    # --- 预构建关联时间查找表 ---
    need_filter = start_dt is not None or end_dt is not None
    asset_lookup: Dict[str, Optional[datetime]] = {}
    port_url_lookup: Dict[str, Optional[datetime]] = {}
    port_pair_lookup: Dict[tuple, Optional[datetime]] = {}

    if need_filter:
        asset_sheet_name = _find_sheet(wb_c, '资产总表')
        port_sheet_name  = _find_sheet(wb_c, '端口表')
        if asset_sheet_name:
            asset_lookup = _build_single_key_lookup(
                wb_c[asset_sheet_name], 'Host（IP/子域名）')
            log(f"  [lookup] 资产总表 条目数={len(asset_lookup)}")
        else:
            log("  [lookup] 警告：未找到「资产总表」sheet", "WARNING")
        if port_sheet_name:
            port_url_lookup  = _build_single_key_lookup(
                wb_c[port_sheet_name], '访问路径', connectivity_col='端口连通性')
            port_pair_lookup = _build_pair_key_lookup(
                wb_c[port_sheet_name], 'Host', '端口', connectivity_col='端口连通性')
            log(f"  [lookup] 端口表 URL条目={len(port_url_lookup)}，"
                f"IP+端口条目={len(port_pair_lookup)}")
        else:
            log("  [lookup] 警告：未找到「端口」sheet", "WARNING")


    # 各 sheet 前缀对应的过滤规则
    # DIRECT_FILTER = {'根域名', 'SSL证书', '公众号&小程序资产', 'APP资产'}
    DIRECT_FILTER = {'根域名', 'SSL证书', '公众号&小程序资产', 'APP资产', '资产总表', '端口表'}
    ASSET_LOOKUP  = {
        '子域名':    '子域名',
        'IP C段':    '对应主机',
        '公有云资产': 'IP地址/子域名',
    }
    PORT_URL_FILTER  = {'登录入口', '网络&安全设备', 'WEB资产'}
    PORT_PAIR_FILTER = {'非WEB资产'}

    for name in wb_c.sheetnames:
        print("start")
        print(name)
        prefix = name.split('（')[0].split('(')[0].strip()
        print(prefix)

        if prefix == '文档说明':
            log(f"[C] {name} — 已跳过（文档说明）")
            continue

        dst_ws = _copy_sheet(wb_c[name], wb_new, name)
        log(f"[C] {name}")

        if prefix == '资产总表':
            n = _delete_columns_by_header(dst_ws, ['风险数量', '网站监测授权'])
            log(f"[C] {name} — 删除 {n} 列（风险数量、网站监测授权）")
        elif prefix == '端口表':
            cid = _find_col_idx(dst_ws, '端口连通性')
            if cid is not None:
                to_del_conn = [
                    cell.row for cell in dst_ws['A'][1:]
                    if dst_ws.cell(row=cell.row, column=cid).value is None
                    or str(dst_ws.cell(row=cell.row, column=cid).value).strip() != '开放'
                ]
                for ri in sorted(to_del_conn, reverse=True):
                    dst_ws.delete_rows(ri)
                log(f"  [filter] {name!r}: 端口连通性过滤，删除 {len(to_del_conn)} 行")

        if need_filter:
            if prefix in DIRECT_FILTER:
                _apply_time_filter(dst_ws, start_dt, end_dt, 'direct')  # 自身有时间字段，不记录 debug
            elif prefix in ASSET_LOOKUP:
                _apply_time_filter(dst_ws, start_dt, end_dt, 'host_lookup',
                                   debug_wb=debug_wb,
                                   key_col=ASSET_LOOKUP[prefix], lookup=asset_lookup)
            elif prefix in PORT_URL_FILTER:
                _apply_time_filter(dst_ws, start_dt, end_dt, 'host_lookup',
                                   debug_wb=debug_wb,
                                   key_col='访问路径', lookup=port_url_lookup)
            elif prefix in PORT_PAIR_FILTER:
                _apply_time_filter(dst_ws, start_dt, end_dt, 'pair_lookup',
                                   debug_wb=debug_wb,
                                   host_col='IP地址/子域名', port_col='端口',
                                   lookup=port_pair_lookup)
            else:
                print("未找到时间处理方式")

    # return None

    B_WEB_PREFIX     = '重要组件'
    B_NON_WEB_PREFIX = '高危端口&服务'

    sheet_web = _find_sheet(wb_b, B_WEB_PREFIX)
    if not sheet_web:
        raise ValueError(f"文件B中未找到以「{B_WEB_PREFIX}」开头的sheet，实际sheets: {wb_b.sheetnames}")
    dst_web = _copy_sheet(wb_b[sheet_web], wb_new, 'Web服务风险分布')
    log(f"[B] {sheet_web} → Web服务风险分布")
    if need_filter:
        _apply_time_filter(dst_web, start_dt, end_dt, 'host_lookup',
                           debug_wb=debug_wb,
                           key_col='访问路径', lookup=port_url_lookup)

    sheet_non_web = _find_sheet(wb_b, B_NON_WEB_PREFIX)
    if not sheet_non_web:
        raise ValueError(f"文件B中未找到以「{B_NON_WEB_PREFIX}」开头的sheet，实际sheets: {wb_b.sheetnames}")
    dst_non_web = _copy_sheet(wb_b[sheet_non_web], wb_new, '非Web服务风险分布')
    log(f"[B] {sheet_non_web} → 非Web服务风险分布")
    if need_filter:
        _apply_time_filter(dst_non_web, start_dt, end_dt, 'pair_lookup',
                           debug_wb=debug_wb,
                           host_col='IP地址/子域名', port_col='端口',
                           lookup=port_pair_lookup)

    # 自动单元格大小
    # for ws in wb_new.worksheets:
    #     for rd in ws.row_dimensions.values():
    #         rd.height = None

    wb_new.save(output_path)

    if debug_wb is not None and debug_wb.worksheets:
        debug_path = os.path.join(os.path.dirname(output_path), 'debug_not_find_time.xlsx')
        debug_wb.save(debug_path)
        log(f"[debug] 未找到时间数据已保存: {debug_path}")


def _on_script_timeout():
    log(f"错误：脚本执行超时（超过 {SCRIPT_TIMEOUT // 60} 分钟），强制退出", "ERROR")
    os._exit(1)


# ---------- 主流程 ----------

def main():
    _timer = threading.Timer(SCRIPT_TIMEOUT, _on_script_timeout)
    _timer.daemon = True
    _timer.start()

    parser = argparse.ArgumentParser(
        description='EASM暴露面清单生成工具',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            '示例:\n'
            '  python generate_report.py 深圳市口袋网络\n'
            '  python generate_report.py 35690473\n'
            '  python generate_report.py --zip report.zip --start-time 2025-01-01 --end-time 2025-03-31'
        )
    )
    parser.add_argument('keyword', nargs='?', default=None,
                        help='客户ID或客户名称（支持模糊匹配）；传入 --zip 时可省略')
    parser.add_argument('--zip', dest='zip_path', default=None,
                        help='直接指定本地zip压缩包路径，跳过接口调用（步骤1-6），从解压开始')
    parser.add_argument('--start-time', default=None,
                        help='数据过滤起始时间，如 2025-01-01 或 "2025-01-01 00:00:00"')
    parser.add_argument('--end-time', default=None,
                        help='数据过滤结束时间，如 2025-03-31 或 "2025-03-31 23:59:59"'
                             '（纯日期时自动取当天末尾 23:59:59）')
    parser.add_argument('--debug', action='store_true',
                        help='（可选）输出 debug_not_find_time.xlsx，记录未匹配到时间的行')
    args = parser.parse_args()

    # 解析时间过滤参数
    start_dt: Optional[datetime] = None
    end_dt:   Optional[datetime] = None
    if args.start_time:
        try:
            start_dt = _parse_arg_time(args.start_time, is_end=False)
            log(f"过滤起始时间: {start_dt}")
        except ValueError as e:
            sys.exit(f"错误：{e}")
    if args.end_time:
        try:
            end_dt = _parse_arg_time(args.end_time, is_end=True)
            log(f"过滤结束时间: {end_dt}")
        except ValueError as e:
            sys.exit(f"错误：{e}")

    # 传入 --zip 时直接跳到第7步
    if args.zip_path:
        if not os.path.exists(args.zip_path):
            sys.exit(f"错误：指定的zip文件不存在 → {args.zip_path}")
        zip_path = args.zip_path
        log(f"[直接模式] 使用本地zip: {zip_path}")
    else:
        if not args.keyword:
            sys.exit("错误：未传入 --zip 时必须提供客户ID或客户名称关键词")

        # 1. 加载 cookies
        if not os.path.exists(COOKIES_FILE):
            sys.exit(f"错误：cookies文件不存在 → {COOKIES_FILE}")
        cookie_str = read_cookies_as_string(COOKIES_FILE)
        if not cookie_str:
            sys.exit(f"错误：cookies解析结果为空，请检查文件格式 → {COOKIES_FILE}")
        headers = _build_headers(cookie_str)
        log("[1/7] 已加载cookies")

        os.makedirs(TEMP_DIR, exist_ok=True)

        # 2. 接口0：搜索客户
        easm_keyword = args.keyword + "[影子]"
        log(f"[2/7] 搜索客户「{easm_keyword}」...")
        customers = api0_search_customer(headers, easm_keyword)

        if not customers:
            sys.exit(f"错误：未找到匹配客户，请检查关键词「{easm_keyword}」")
        if len(customers) > 1:
            exact = _pick_exact_match(customers, easm_keyword)
            if exact:
                customers = [exact]
            else:
                log(f"错误：找到 {len(customers)} 个匹配客户，请使用更精确的关键词：", "ERROR")
                for c in customers:
                    name = c.get('company_name') or c.get('pms_customer_name', '未知')
                    print(f"  ID={c['company_id']}  名称={name}")
                sys.exit(1)

        customer     = customers[0]
        company_id   = customer['company_id']
        company_name = customer.get('company_name') or customer.get('pms_customer_name', '未知')
        log(f"确认客户: {company_name}（ID={company_id}）")

        # 3. 接口1：获取模板
        log("[3/7] 获取报告模板...")
        template_id, template_name = api1_get_template(headers)

        # 4. 接口2：触发报告生成
        log("[4/7] 触发报告生成...")
        task_id = api2_generate_report(headers, company_id, template_id, template_name)

        # 5. 接口3：轮询状态
        log("[5/7] 等待报告生成...")
        api3_poll_status(headers, task_id)

        # 6. 接口4：下载压缩包
        log("[6/7] 下载报告...")
        zip_path = api4_download_report(headers, task_id, TEMP_DIR)

    # 7. 解压 + 构建新报告
    extract_dir = os.path.join(TEMP_DIR, 'extracted')
    os.makedirs(extract_dir, exist_ok=True)
    log("[7/7] 构建暴露面清单...")
    file_b, file_c = extract_report(zip_path, extract_dir)
    log(file_b)
    log(file_c)
    build_output_excel(file_b, file_c, OUTPUT_FILE, start_dt=start_dt, end_dt=end_dt,
                       debug_mode=args.debug)

    _timer.cancel()
    log(f"完成！输出文件: {OUTPUT_FILE}")


if __name__ == '__main__':
    main()
