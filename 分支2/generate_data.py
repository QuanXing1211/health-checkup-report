import openpyxl
import json
from collections import defaultdict

BASE_DIR = "数据清单"
RISK_LEVEL_ORDER = {"严重": 4, "高危": 3, "中危": 2, "低危": 1}


# ════════════════════════════════════════════════════════════════
# 层1：工具函数
# ════════════════════════════════════════════════════════════════

def load_wb(filename):
    """加载 Excel 工作簿（data_only 模式，读取公式计算结果）"""
    return openpyxl.load_workbook(f"{BASE_DIR}/{filename}", data_only=True)


def get_rows(ws):
    """将 sheet 转为 [{列名: 值, ...}] 列表，跳过空行"""
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            row_dict = {}
            for h, v in zip(headers, row):
                if h is not None:
                    row_dict[h] = v
            rows.append(row_dict)
    return rows


def count_data_rows(ws):
    """统计 sheet 有效数据行数（不构建字典，轻量版）"""
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            count += 1
    return count


def s(v):
    """安全字符串转换：None → ''"""
    return str(v).strip() if v is not None else ""


def unique_count(rows, field):
    """统计某字段的去重数量"""
    return len(set(s(r.get(field)) for r in rows if r.get(field)))


def filter_count(rows, field, value):
    """统计某字段 == value 的行数"""
    return sum(1 for r in rows if s(r.get(field)) == value)


def filter_count_multi(rows, field, values):
    """统计某字段 in values 的行数"""
    return sum(1 for r in rows if s(r.get(field)) in values)


def top_by_risk_level(rows, key_field):
    """按风险等级（严重>高危>中危>低危）取最严重的 key，
       等级相同时取数量最多的。返回 key 名或空字符串"""
    data = defaultdict(list)
    for r in rows:
        key = s(r.get(key_field))
        level = s(r.get("风险等级"))
        if key:
            data[key].append(level)
    if not data:
        return ""

    def sort_key(k):
        levels = data[k]
        max_level = max((RISK_LEVEL_ORDER.get(l, 0) for l in levels), default=0)
        return (-max_level, -len(levels))

    return sorted(data.keys(), key=sort_key)[0]


def top_priority_rows(rows, key_field, n=5):
    """按修复优先级（急需 > 尽快 > 建议）分组统计，取 top n。
       返回 [{asset, urgent, soon, suggest}, ...]"""
    stats = defaultdict(lambda: [0, 0, 0])
    for r in rows:
        k = s(r.get(key_field))
        p = s(r.get("修复优先级"))
        if not k:
            continue
        if p == "急需修复":
            stats[k][0] += 1
        elif p == "尽快修复":
            stats[k][1] += 1
        elif p == "建议修复":
            stats[k][2] += 1
    sorted_keys = sorted(
        stats.keys(),
        key=lambda k: (-stats[k][0], -stats[k][1], -stats[k][2])
    )[:n]
    return [
        {"asset": k, "urgent": stats[k][0], "soon": stats[k][1], "suggest": stats[k][2]}
        for k in sorted_keys
    ]


def fill_forward(rows, field):
    """向前填充 Excel 合并单元格导致的 None 值。
       修改 rows 原地，返回 rows 自身。"""
    last_val = ""
    for r in rows:
        cur = s(r.get(field))
        if cur:
            last_val = cur
        elif last_val:
            r[field] = last_val
    return rows


# ════════════════════════════════════════════════════════════════
# 层2：数据加载（纯 IO，无计算）
# ════════════════════════════════════════════════════════════════

def load_all_data():
    """加载所有 Excel 工作簿，返回原始 rows 字典"""
    wb_exp = load_wb("暴露面清单.xlsx")
    wb_vuln = load_wb("漏洞清单.xlsx")
    wb_weak = load_wb("弱口令清单.xlsx")
    wb_event = load_wb("安全事件表.xlsx")
    wb_asset = load_wb("资产表.xlsx")

    try:
        ws_asset_sheet = wb_asset["资产表"]
    except KeyError:
        ws_asset_sheet = wb_asset.worksheets[0]

    return {
        "wb_exp": wb_exp,
        "rows_web_risk":    get_rows(wb_exp["Web服务风险分布"]),
        "rows_nonweb_risk": get_rows(wb_exp["非Web服务风险分布"]),
        "rows_port":        get_rows(wb_exp["端口表"]),
        "rows_vuln":        get_rows(wb_vuln["漏洞"]),
        "rows_weak":        get_rows(wb_weak["弱口令"]),
        "rows_event":       get_rows(wb_event["事件表"]),
        "rows_asset":       get_rows(ws_asset_sheet),
    }


# ════════════════════════════════════════════════════════════════
# 层3：数据准备（过滤 + 预计算辅助结构）
# ════════════════════════════════════════════════════════════════

def prepare_datasets(raw):
    """基于原始 rows 做过滤、预计算，返回增强后的 datasets 字典"""
    rows_vuln  = raw["rows_vuln"]
    rows_weak  = raw["rows_weak"]
    rows_port  = raw["rows_port"]

    # ── 向前填充合并单元格的 None 值 ──
    fill_forward(raw["rows_web_risk"], "组件名称")
    fill_forward(raw["rows_nonweb_risk"], "服务")

    # ── 过滤：按数据来源 / 端口状态 ──
    vuln_net   = [r for r in rows_vuln if s(r.get("数据来源")) == "外网"]
    vuln_intra = [r for r in rows_vuln if s(r.get("数据来源")) == "内网"]
    weak_net   = [r for r in rows_weak if s(r.get("数据来源")) == "外网"]
    weak_intra = [r for r in rows_weak if s(r.get("数据来源")) == "内网"]
    open_ports = [r for r in rows_port if s(r.get("端口连通性")) == "开放"]

    # ── 辅助映射：端口 · 访问路径 → Host ──
    port_path_to_host = {}
    for r in rows_port:
        path = s(r.get("访问路径"))
        host = s(r.get("Host"))
        if path and host:
            port_path_to_host[path] = host

    def web_risk_host(row):
        return port_path_to_host.get(s(row.get("访问路径")), "")

    # ── 暴露面 host / IP 集合 ──
    web_hosts = set()
    for r in raw["rows_web_risk"]:
        h = web_risk_host(r)
        if h:
            web_hosts.add(h)

    nonweb_ips = set()
    for r in raw["rows_nonweb_risk"]:
        ip = s(r.get("IP地址/子域名"))
        if ip:
            nonweb_ips.add(ip)

    # ── 弱口令资产级预计算 ──
    weak_asset_count = defaultdict(int)
    asset_importance = {}
    asset_own_group = {}
    for r in rows_weak:
        a = s(r.get("风险资产"))
        if a:
            weak_asset_count[a] += 1
            if a not in asset_importance:
                asset_importance[a] = s(r.get("资产重要性"))
            if a not in asset_own_group:
                asset_own_group[a] = s(r.get("所属资产组"))

    # ── 安全事件资产集合 ──
    event_assets = set()
    for r in raw["rows_event"]:
        a = s(r.get("影响资产"))
        if a:
            event_assets.add(a)

    # ── 资产表 IP → 资产名称映射 ──
    asset_ip_to_name = {}
    for r in raw["rows_asset"]:
        ip = s(r.get("IP地址"))
        name = s(r.get("资产名称"))
        if ip:
            asset_ip_to_name[ip] = name

    asset_ip_to_group = {}
    for r in raw["rows_asset"]:
        ip = s(r.get("IP地址"))
        group = s(r.get("资产组名"))
        if ip:
            asset_ip_to_group[ip] = group

    # ── 合并返回 ──
    return {
        **raw,
        "vuln_net": vuln_net,
        "vuln_intra": vuln_intra,
        "weak_net": weak_net,
        "weak_intra": weak_intra,
        "open_ports": open_ports,
        "web_hosts": web_hosts,
        "nonweb_ips": nonweb_ips,
        "weak_asset_count": weak_asset_count,
        "asset_importance": asset_importance,
        "asset_own_group": asset_own_group,
        "event_assets": event_assets,
        "asset_ip_to_name": asset_ip_to_name,
        "asset_ip_to_group": asset_ip_to_group,
        "web_risk_host": web_risk_host,
    }


# ════════════════════════════════════════════════════════════════
# 层4：分模块计算函数
#     每个函数输入 ds（datasets 字典），返回该模块的 dict 片段
# ════════════════════════════════════════════════════════════════

def calc_summary(ds):
    """计算 data["summary"] 全部字段"""
    return {
        "internet": {
            "exposure": {
                "risk_ports":  len(ds["rows_web_risk"]) + len(ds["rows_nonweb_risk"]),
                "total_ports": len(ds["rows_port"]),
                "risk_assets": len(ds["web_hosts"] | ds["nonweb_ips"]),
            },
            "vuln": {
                "priority_urgent": filter_count(ds["vuln_net"], "修复优先级", "急需修复"),
                "total":           len(ds["vuln_net"]),
                "risk_assets":     unique_count(ds["vuln_net"], "风险资产"),
            },
            "weak_pwd": {
                "risk_assets": unique_count(ds["weak_net"], "风险资产"),
                "total":       len(ds["weak_net"]),
            },
        },
        "intranet": {
            "vuln": {
                "priority_urgent": filter_count(ds["vuln_intra"], "修复优先级", "急需修复"),
                "total":           len(ds["vuln_intra"]),
                "risk_assets":     unique_count(ds["vuln_intra"], "风险资产"),
            },
            "weak_pwd": {
                "risk_assets": unique_count(ds["weak_intra"], "风险资产"),
                "total":       len(ds["weak_intra"]),
            },
        },
    }


def calc_key_risks(ds):
    """计算 data["key_risks"] 全部字段"""
    # ── vuln ──
    _biz = top_by_risk_level(ds["vuln_intra"], "所属业务")
    biz_example = ("（如" + _biz + "）") if _biz else ""
    high_count  = filter_count_multi(ds["rows_vuln"], "风险等级", {"严重", "高危"})

    # ── weak_pwd ──
    _weak_biz = top_by_risk_level(ds["weak_intra"], "所属业务")
    weak_biz_example = ("（如" + _weak_biz + "）") if _weak_biz else ""
    weak_total = len(ds["rows_weak"])

    # example_asset: 弱口令数量 top3 资产
    top3_by_count = sorted(ds["weak_asset_count"], key=lambda a: -ds["weak_asset_count"][a])[:3]
    example_asset = ("（如" + "、".join(top3_by_count) + "）") if top3_by_count else ""

    # priority_assets: 关联事件资产 > 核心资产 > 弱口令数量，取 top3
    all_weak_assets = list(set(s(r.get("风险资产")) for r in ds["rows_weak"] if r.get("风险资产")))

    def pri_sort_key(a):
        has_event = a in ds["event_assets"]
        is_core = ds["asset_importance"].get(a, "") == "核心资产"
        group = 0 if has_event else (1 if is_core else 2)
        return (group, -ds["weak_asset_count"].get(a, 0))

    top3_pri = sorted(all_weak_assets, key=pri_sort_key)[:3]
    priority_assets = ("建议优先修改资产：" + "、".join(top3_pri) + "的弱口令，并添加多因素认证。") if top3_pri else ""

    # priority_assets_reason: 说明优先原因
    top3_types = set()
    for a in top3_pri:
        if a in ds["event_assets"]:
            top3_types.add("关联事件资产")
        elif ds["asset_importance"].get(a, "") == "核心资产":
            top3_types.add("核心资产")
    if top3_types == {"关联事件资产"}:
        priority_assets_reason = "因为这些资产关联事件。"
    elif top3_types == {"核心资产"}:
        priority_assets_reason = "因为这些资产属于核心业务。"
    elif "关联事件资产" in top3_types and "核心资产" in top3_types:
        priority_assets_reason = "因为这些资产关联事件或有核心资产。"
    else:
        priority_assets_reason = ""

    # ── exposure ──
    web_cnt    = len(ds["rows_web_risk"])
    nonweb_cnt = len(ds["rows_nonweb_risk"])

    # nonweb_services: 最多 2 个不重复的服务名
    svc_list, svc_seen = [], set()
    for r in ds["rows_nonweb_risk"]:
        sv = s(r.get("服务"))
        if sv and sv not in svc_seen:
            svc_seen.add(sv)
            svc_list.append(sv)
        if len(svc_list) >= 2:
            break
    nonweb_services = ("（如" + "、".join(svc_list) + "等）") if svc_list else ""

    # example_service: 第 1 条非 Web 风险的资产:端口（服务）
    example_service = ""
    if ds["rows_nonweb_risk"]:
        r0 = ds["rows_nonweb_risk"][0]
        asset_val = s(r0.get("风险资产")) or s(r0.get("IP地址/子域名"))
        example_service = "（如" + f"{asset_val}:{s(r0.get('端口'))}（{s(r0.get('服务'))}）" + "）"

    return {
        "vuln": {
            "biz_example": biz_example,
            "high_count":  high_count,
        },
        "weak_pwd": {
            "biz_example":     weak_biz_example,
            "total":           weak_total,
            "example_asset":   example_asset,
            "priority_assets": priority_assets,
            "priority_assets_reason": priority_assets_reason,
        },
        "exposure": {
            "web_count":       web_cnt,
            "nonweb_count":    nonweb_cnt,
            "total":           web_cnt + nonweb_cnt,
            "nonweb_services": nonweb_services,
            "example_service": example_service,
        },
    }


def calc_risk_detail(ds):
    """计算 data["risk_detail"] 全部字段"""
    rd_exp      = len(ds["rows_web_risk"]) + len(ds["rows_nonweb_risk"])
    rd_vuln_net = len(ds["vuln_net"])
    rd_weak_net = len(ds["weak_net"])

    A = filter_count_multi(ds["rows_web_risk"], "风险等级", {"严重", "高危"})
    B = filter_count_multi(ds["rows_nonweb_risk"], "风险等级", {"严重", "高危"})
    C = filter_count_multi(ds["vuln_net"], "风险等级", {"严重", "高危"})

    return {
        "internet": {
            "exposure":   rd_exp,
            "vuln":       rd_vuln_net,
            "weak_pwd":   rd_weak_net,
            "total":      rd_exp + rd_vuln_net + rd_weak_net,
            "high_above": A + B + C,
        },
        "intranet": {
            "vuln":     len(ds["vuln_intra"]),
            "weak_pwd": len(ds["weak_intra"]),
            "total":    len(ds["vuln_intra"]) + len(ds["weak_intra"]),
            "high":     filter_count_multi(ds["vuln_intra"], "风险等级", {"严重", "高危"}),
        },
    }


def calc_internet_exposure(ds):
    """计算 data["internet"]["exposure"] 全部字段"""
    web_count    = len(ds["rows_web_risk"])
    nonweb_count = len(ds["rows_nonweb_risk"])

    # ── dist：暴露面清单中除 3 个 sheet 外每个 sheet 的数据行数 ──
    EXCLUDED_SHEETS = {"文档说明", "Web服务风险分布", "非Web服务风险分布"}
    dist_list = []
    for sn in ds["wb_exp"].sheetnames:
        if sn not in EXCLUDED_SHEETS:
            cnt = count_data_rows(ds["wb_exp"][sn])
            dist_list.append({
                "name": sn, "value": cnt,
            })

    # ── web_top5：按组件名称统计 ──
    comp_cnt = defaultdict(int)
    for r in ds["rows_web_risk"]:
        c = s(r.get("组件名称"))
        if c:
            comp_cnt[c] += 1
    web_top5 = [
        {"name": n, "value": v}
        for n, v in sorted(comp_cnt.items(), key=lambda x: -x[1])[:5]
    ]

    # ── nonweb_top5：按服务统计 ──
    svc_cnt = defaultdict(int)
    for r in ds["rows_nonweb_risk"]:
        sv = s(r.get("服务"))
        if sv:
            svc_cnt[sv] += 1
    nonweb_top5 = [
        {"name": n, "value": v}
        for n, v in sorted(svc_cnt.items(), key=lambda x: -x[1])[:5]
    ]

    # ── stack_rows：按 host 聚合 web + nonWeb 数量，取 riskTotal top5 ──
    h_web_cnt    = defaultdict(int)
    h_nonweb_cnt = defaultdict(int)
    for r in ds["rows_web_risk"]:
        h = ds["web_risk_host"](r)
        if h:
            h_web_cnt[h] += 1
    for r in ds["rows_nonweb_risk"]:
        h = s(r.get("IP地址/子域名"))
        if h:
            h_nonweb_cnt[h] += 1
    all_h = set(h_web_cnt) | set(h_nonweb_cnt)
    stack_rows = sorted(
        [{"host": h, "web": h_web_cnt[h], "nonWeb": h_nonweb_cnt[h],
          "riskTotal": h_web_cnt[h] + h_nonweb_cnt[h]} for h in all_h],
        key=lambda x: -x["riskTotal"]
    )[:5]

    return {
        "risk_asset_count": unique_count(ds["open_ports"], "Host"),
        "port_count":       len(ds["open_ports"]),
        "vuln_count":       len(ds["vuln_net"]),
        "total_exposure":   len(ds["open_ports"]),
        "risk_exposure":    web_count + nonweb_count,
        "total_assets":     unique_count(ds["open_ports"], "Host"),
        "risk_assets":      len(ds["web_hosts"] | ds["nonweb_ips"]),
        "dist":             dist_list,
        "web_top5":         web_top5,
        "nonweb_top5":      nonweb_top5,
        "stack_rows":       stack_rows,
    }


def calc_internet_vuln(ds):
    """计算 data["internet"]["vuln"] 全部字段"""
    return {
        "total":           len(ds["vuln_net"]),
        "critical":        filter_count(ds["vuln_net"], "风险等级", "严重"),
        "high":            filter_count(ds["vuln_net"], "风险等级", "高危"),
        "medium":          filter_count(ds["vuln_net"], "风险等级", "中危"),
        "low":             filter_count(ds["vuln_net"], "风险等级", "低危"),
        "related_assets":  unique_count(ds["vuln_net"], "风险资产"),
        "priority_urgent":  filter_count(ds["vuln_net"], "修复优先级", "急需修复"),
        "priority_soon":    filter_count(ds["vuln_net"], "修复优先级", "尽快修复"),
        "priority_suggest": filter_count(ds["vuln_net"], "修复优先级", "建议修复"),
        "top_rows":        top_priority_rows(ds["vuln_net"], "风险资产"),
    }


def calc_internet_weak_pwd(ds):
    """计算 data["internet"]["weak_pwd"] 全部字段"""
    net_asset_cnt = defaultdict(int)
    for r in ds["weak_net"]:
        a = s(r.get("风险资产"))
        if a:
            net_asset_cnt[a] += 1
    top5_net_a = sorted(net_asset_cnt.items(), key=lambda x: -x[1])[:5]

    return {
        "affected_assets": unique_count(ds["weak_net"], "风险资产"),
        "total_count":     len(ds["weak_net"]),
        "asset_rows":      [{"asset": a, "count": c} for a, c in top5_net_a],
    }


def calc_intranet_vuln(ds):
    """计算 data["intranet"]["vuln"] 全部字段"""
    return {
        "total":            len(ds["vuln_intra"]),
        "critical":         filter_count(ds["vuln_intra"], "风险等级", "严重"),
        "high":             filter_count(ds["vuln_intra"], "风险等级", "高危"),
        "medium":           filter_count(ds["vuln_intra"], "风险等级", "中危"),
        "low":              filter_count(ds["vuln_intra"], "风险等级", "低危"),
        "related_biz":      unique_count(ds["vuln_intra"], "所属业务"),
        "related_assets":   unique_count(ds["vuln_intra"], "风险资产"),
        "priority_urgent":  filter_count(ds["vuln_intra"], "修复优先级", "急需修复"),
        "priority_soon":    filter_count(ds["vuln_intra"], "修复优先级", "尽快修复"),
        "priority_suggest": filter_count(ds["vuln_intra"], "修复优先级", "建议修复"),
        "biz_top_rows":     top_priority_rows(ds["vuln_intra"], "所属业务"),
        "asset_top_rows":   top_priority_rows(ds["vuln_intra"], "风险资产"),
    }


def calc_intranet_weak_pwd(ds):
    """计算 data["intranet"]["weak_pwd"] 全部字段"""
    # biz_rows: 按所属业务 top5
    biz_cnt = defaultdict(int)
    for r in ds["weak_intra"]:
        b = s(r.get("所属业务"))
        if b:
            biz_cnt[b] += 1
    top5_biz = sorted(biz_cnt.items(), key=lambda x: -x[1])[:5]

    # asset_rows: 按风险资产 top5，附带 asset_name/asset_group
    intra_asset_cnt = defaultdict(int)
    intra_asset_group = {}
    for r in ds["weak_intra"]:
        a = s(r.get("风险资产"))
        if a:
            intra_asset_cnt[a] += 1
            if a not in intra_asset_group:
                intra_asset_group[a] = s(r.get("所属资产组"))
    top5_intra_a = sorted(intra_asset_cnt.items(), key=lambda x: -x[1])[:5]

    return {
        "total_count":     len(ds["weak_intra"]),
        "affected_assets": unique_count(ds["weak_intra"], "风险资产"),
        "risk_count":      unique_count(ds["weak_intra"], "所属业务"),
        "biz_rows":        [{"asset": b, "count": c} for b, c in top5_biz],
        "asset_rows": [
            {
                "asset": a, "count": c,
                "asset_name":  ds["asset_ip_to_name"].get(a, ""),
                "asset_group": ds["asset_ip_to_group"].get(a, ""),
            }
            for a, c in top5_intra_a
        ],
    }


# ════════════════════════════════════════════════════════════════
# 层5：主函数 —— 串联所有层
# ════════════════════════════════════════════════════════════════

def main():
    # 1. 加载所有原始数据
    raw = load_all_data()

    # 2. 过滤 + 预计算辅助结构
    ds = prepare_datasets(raw)

    # 3. 读取或初始化 data.json
    try:
        with open("data.json", encoding="utf-8") as f:
            data = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        data = {}

    # 确保顶层结构存在
    for key in ("summary", "key_risks", "risk_detail", "internet", "intranet"):
        if key not in data:
            data[key] = {}
    for key in ("exposure", "vuln", "weak_pwd"):
        if key not in data["internet"]:
            data["internet"][key] = {}
    for key in ("vuln", "weak_pwd"):
        if key not in data["intranet"]:
            data["intranet"][key] = {}

    # 4. 分模块计算，直接覆盖写入 data 中
    data["summary"]                 = calc_summary(ds)
    data["key_risks"]               = calc_key_risks(ds)
    data["risk_detail"]             = calc_risk_detail(ds)
    data["internet"]["exposure"]    = calc_internet_exposure(ds)
    data["internet"]["vuln"]        = calc_internet_vuln(ds)
    data["internet"]["weak_pwd"]    = calc_internet_weak_pwd(ds)
    data["intranet"]["vuln"]        = calc_intranet_vuln(ds)
    data["intranet"]["weak_pwd"]    = calc_intranet_weak_pwd(ds)

    # 5. 写出 data.json
    with open("data.json", "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print("data.json 更新完成")


if __name__ == "__main__":
    main()
