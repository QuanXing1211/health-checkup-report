"""
策略检查数据采集与导出
======================
从 XDR 策略检查结果查询接口分页获取指定 time_range 时间范围内的数据，写入 Excel。
用法:
    python policy_check_export.py
    python policy_check_export.py --company_id 57229265 --start "2026-06-01" --end "2026-06-22"
    python policy_check_export.py --company_id 57229265 --start "2026-06-01" --end "2026-06-22" --status at_risk
"""

import argparse
import datetime
import json
import os
import ssl
import urllib.request
import urllib.error

import openpyxl


# ──────────────────────────────────────────────
# 常量
# ──────────────────────────────────────────────

API_URL = "https://sitmssw.soar.sangfor.com/gateway/idps/order/v1/tools/task/xdr_policy_check/result"
PAGE_SIZE = 100
DEFAULT_OUTPUT_PATH = "策略检查清单.xlsx"
TMP_DIR = "tmp"

COLUMN_MAP = [
    ("序号", "seq"),
    ("设备", "dev_name"),
    ("策略名称", "name"),
    ("策略状态", "policy_status"),
    ("风险状态", "risk_status"),
    ("策略描述", "description"),
    ("最近检查时间", "latest_time"),
    ("风险描述", "risk_desc"),
]

SHEET_NAME = "策略检查"

DEV_TYPE_DICT = {
    3: "AF",
    9: "SIP",
    12: "EDR",
}


# ──────────────────────────────────────────────
# 工具函数
# ──────────────────────────────────────────────

def _extract_first(value):
    """从可能是列表的值中提取第一个元素，若为字符串则直接返回。"""
    if isinstance(value, list) and value:
        return str(value[0])
    if isinstance(value, str):
        return value
    return ""


def _utc_to_local_str(utc_str):
    """
    将 UTC 时间字符串（如 "2026-07-13T01:31:32Z"）转换为当前时区的本地时间字符串。

    优先使用系统当前时区，获取不到时回退到东八区（UTC+8）。
    输出格式为 "YYYY-MM-DD HH:MM:SS"。解析失败时原样返回。
    """
    if not utc_str:
        return ""
    try:
        normalized = utc_str.replace("Z", "+00:00")
        dt = datetime.datetime.fromisoformat(normalized)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=datetime.timezone.utc)
    except ValueError:
        return utc_str
    local_tz = datetime.datetime.now().astimezone().tzinfo or datetime.timezone(datetime.timedelta(hours=8))
    return dt.astimezone(local_tz).strftime("%Y-%m-%d %H:%M:%S")


def _parse_datetime(text):
    """
    解析日期时间字符串，按本地时间解释并附加本地时区信息。

    输入字符串按本地时间解释，支持 YYYY-MM-DD 和 YYYY-MM-DD HH:MM:SS 两种格式。
    返回带本地时区 tzinfo 的 datetime（保持本地时间，不转 UTC）。
    调用接口需要 UTC 时再由调用方 astimezone(UTC) 转换。
    """
    naive = None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            naive = datetime.datetime.strptime(text, fmt)
            break
        except ValueError:
            continue
    if naive is None:
        raise ValueError(f"无法解析日期: {text}，请使用 YYYY-MM-DD 或 YYYY-MM-DD HH:MM:SS 格式")
    # 附加本地时区；获取不到本地时区时回退到东八区
    local_tz = datetime.datetime.now().astimezone().tzinfo or datetime.timezone(datetime.timedelta(hours=8))
    return naive.replace(tzinfo=local_tz)


# ──────────────────────────────────────────────
# 核心类
# ──────────────────────────────────────────────

class PolicyCheckExporter:
    """
    策略检查数据采集与导出器。

    封装获取数据、格式转换、写入 Excel 三个步骤。
    配置（客户ID、时间范围、状态、cookie、输出路径）通过构造函数传入，
    调用 run() 执行完整流程，也可单独调用各步骤方法。

    用法:
        exporter = PolicyCheckExporter(
            company_id="57229265",
            start_time="2026-06-01",
            end_time="2026-06-22",
        )
        exporter.run()
    """

    def __init__(
        self,
        company_id,
        start_time,
        end_time,
        status="",
        cookie=None,
        cookie_path=None,
        output_path=None,
        json_output_path=None,
    ):
        self.company_id = company_id
        self.start_time = _parse_datetime(start_time)
        self.end_time = _parse_datetime(end_time)
        # 若 end 只有日期，自动补到当天最后一秒
        if self.end_time.hour == 0 and self.end_time.minute == 0 and self.end_time.second == 0:
            self.end_time = self.end_time.replace(hour=23, minute=59, second=59)
        self.status = status
        self.cookie_path = cookie_path
        self.cookie = cookie if cookie is not None else self._load_cookie(cookie_path)
        self.output_path = output_path or DEFAULT_OUTPUT_PATH
        self.json_output_path = json_output_path

    # ── 第一步：获取数据 ──

    def _load_cookie(self, cookie_path=None):
        """
        从文件读取并提取 cookie 内容。

        cookie_path 必须在创建时提供，否则抛出 ValueError。
        文件格式如果是以 { 或 [ 开头的 JSON，提取 cookieString 或 cookie 字段；
        否则直接作为原始 cookie 字符串使用。
        """
        if not cookie_path:
            raise ValueError("必须提供 cookie_path 以加载 cookie")

        if not os.path.exists(cookie_path):
            raise FileNotFoundError(f"Cookie 文件不存在: {cookie_path}")

        with open(cookie_path, "r", encoding="utf-8") as f:
            raw = f.read()

        stripped = raw.strip()
        if not stripped:
            raise ValueError(f"Cookie 文件为空: {cookie_path}")

        if stripped.startswith("{") or stripped.startswith("["):
            data = json.loads(stripped)
            if isinstance(data, dict):
                return data.get("cookieString", "") or data.get("cookie", "")

        return stripped

    def fetch_data(self):
        """
        从策略检查结果查询接口分页获取指定 time_range 范围内的数据。

        依据 策略检查接口对齐.md 中的接口 3：
          入参: company_id / dev_id_list / limit / time_range / offset
          time_range 为 ISO 日期字符串数组 [start, end]（YYYY-MM-DD）
          返回: code == 0 时，data 为结果条目数组
        本端按 offset+limit 常规分页循环拉取，直至无数据。
        """
        all_records = []
        offset = 0

        # 接口要求 time_range 为 [start, end] ISO 8601 UTC 字符串数组
        # self.start_time / self.end_time 保存的是本地时间，此处转换为 UTC
        utc_tz = datetime.timezone.utc
        time_range = [
            self.start_time.astimezone(utc_tz).strftime("%Y-%m-%dT%H:%M:%SZ"),
            self.end_time.astimezone(utc_tz).strftime("%Y-%m-%dT%H:%M:%SZ"),
        ]

        while True:
            payload = {
                "company_id": self.company_id,
                "limit": PAGE_SIZE,
                "offset": offset,
                "latest_time": time_range,
            }
            if self.status:
                payload["status"] = self.status

            body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
            headers = {"Content-Type": "application/json"}
            if self.cookie:
                headers["Cookie"] = self.cookie
            req = urllib.request.Request(
                API_URL,
                data=body,
                headers=headers,
                method="POST",
            )
            ctx = ssl.create_default_context()
            ctx.check_hostname = False
            ctx.verify_mode = ssl.CERT_NONE

            try:
                with urllib.request.urlopen(req, context=ctx, timeout=30) as resp:
                    result = json.loads(resp.read().decode("utf-8"))
            except urllib.error.HTTPError as e:
                print(f"[ERROR] HTTP 请求失败: {e.code} {e.reason}")
                break
            except urllib.error.URLError as e:
                print(f"[ERROR] 网络请求失败: {e.reason}")
                break

            if result.get("code") != 0:
                print(f"[ERROR] 接口返回错误: code={result.get('code')}, 数据获取中断")
                break

            data = result.get("data", [])
            records = data if isinstance(data, list) else data.get("list", []) if isinstance(data, dict) else []

            if not records:
                break

            all_records.extend(records)

            if len(records) < PAGE_SIZE:
                break

            offset += PAGE_SIZE

        print(f"[INFO] 接口获取 {len(all_records)} 条")
        for item in all_records:
            item["dev_type"] = DEV_TYPE_DICT.get(item.get("dev_type")) or item.get("dev_type") or ""

        return all_records

    # ── 第二步：格式转换 ──

    def transform(self, records):
        """
        将接口原始记录转换为 Excel 行格式。

        每条记录转为一个字典，key 为中文表头名，value 为格式化后的值。
        - 序号在转换时按顺序生成，映射为 seq 字段
        - latest_time / event_time 为列表时取第一个元素
        - 缺失字段填 "-"
        """
        rows = []
        for idx, record in enumerate(records, 1):
            row = {}
            for col_name, field_name in COLUMN_MAP:
                if field_name == "seq":
                    row[col_name] = idx
                elif field_name == "latest_time":
                    row[col_name] = _utc_to_local_str(_extract_first(record.get(field_name, [])))
                elif field_name == "risk_status":
                    raw = record.get("risk_status", "")
                    row[col_name] = "存在风险" if raw == "at_risk" else "无风险"
                else:
                    value = record.get(field_name)
                    row[col_name] = str(value) if value is not None else "-"
            rows.append(row)
        return rows

    # ── 第三步：写入 Excel ──

    def write_excel(self, rows, output_path=None):
        """
        将行字典列表写入 Excel 文件。

        包含表头行和数据行，表头微软雅黑白字深灰底，自动调整列宽。
        """
        path = output_path or self.output_path
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME

        headers = [col_name for col_name, _ in COLUMN_MAP]
        ws.append(headers)

        # 表头样式：微软雅黑 白色字 深灰底（与漏洞清单保持一致）
        header_font = openpyxl.styles.Font(name="微软雅黑", color="FFFFFFFF")
        header_fill = openpyxl.styles.PatternFill(start_color="FF333333", end_color="FF333333", fill_type="solid")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill

        for row in rows:
            values = [row.get(col_name, "") for col_name, _ in COLUMN_MAP]
            ws.append(values)

        for col_idx, header in enumerate(headers, 1):
            max_len = len(str(header))
            for row_idx in range(2, ws.max_row + 1):
                cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "")
                max_len = max(max_len, len(cell_value))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 60)

        wb.save(path)
        print(f"[INFO] 已写入 Excel: {path} ({len(rows)} 行数据)")

    # ── 完整流程 ──

    def _save_json(self, records):
        """将原始数据保存为 JSON 文件到 tmp 目录。"""
        json_path = self.json_output_path
        if not json_path:
            os.makedirs(TMP_DIR, exist_ok=True)
            json_path = os.path.join(TMP_DIR, "policy_check.json")
        else:
            os.makedirs(os.path.dirname(os.path.abspath(json_path)), exist_ok=True)
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(records, f, ensure_ascii=False, indent=2)
        print(f"[INFO] 已保存 JSON: {json_path} ({len(records)} 条记录)")
        return json_path

    def run(self):
        """执行获取数据 → 保存JSON → 格式转换 → 写入 Excel 完整流程。"""
        print("=" * 50)
        print("策略检查数据采集与导出")
        print("=" * 50)
        print(f"  客户ID:    {self.company_id}")
        print(f"  时间范围:  {self.start_time} ~ {self.end_time}")
        print(f"  状态过滤:  {self.status or '(不过滤)'}")
        print(f"  输出文件:  {self.output_path}")
        if self.json_output_path:
            print(f"  JSON输出:  {self.json_output_path}")
        print(f"  Cookie:    {'已加载' if self.cookie else '未加载'}")
        print()

        records = self.fetch_data()

        json_path = self._save_json(records)
        rows = self.transform(records)
        self.write_excel(rows)

        print("\n完成!")
        print("=" * 50)
        return {
            "recordCount": len(records),
            "excelPath": self.output_path,
            "jsonPath": json_path,
        }


# ──────────────────────────────────────────────
# 命令行入口
# ──────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(description="策略检查数据采集与导出")
    parser.add_argument("--company_id", required=True, help="客户ID")
    parser.add_argument("--start", required=True, help="时间范围起始，格式 YYYY-MM-DD 或 YYYY-MM-DD HH:MM:SS")
    parser.add_argument("--end", required=True, help="时间范围结束，格式 YYYY-MM-DD 或 YYYY-MM-DD HH:MM:SS")
    parser.add_argument("--status", default="", help="策略状态过滤，如 at_risk、no_risk，默认不过滤")
    parser.add_argument("--cookie-path", required=True, help="Cookie 文件路径（必填），支持纯文本或含 cookieString 的 JSON")
    parser.add_argument("--output", default=None, help="输出 Excel 路径，默认为当前目录下 策略检查.xlsx")
    parser.add_argument("--json-output", default=None, help="输出 JSON 路径，默认为 tmp/policy_check.json")
    return parser.parse_args()


def main():
    args = parse_args()
    exporter = PolicyCheckExporter(
        company_id=args.company_id,
        start_time=args.start,
        end_time=args.end,
        status=args.status,
        cookie_path=args.cookie_path,
        output_path=args.output,
        json_output_path=args.json_output,
    )
    result = exporter.run()
    if result:
        print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
