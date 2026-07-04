"""
Excel/JSON 读取模块
支持根据表头名称读取 Excel 数据，将列名映射为字段名，返回结构化数据。
支持读取 JSON 文件（标准对象或嵌套 data.list 结构），返回字典列表。
"""

import json
import os
import openpyxl
import re
from typing import Optional


def read_excel(
    filepath: str,
    sheet_name: Optional[str] = None,
    header_row: int = 1,
    data_start_row: int = 2,
    column_alias: Optional[dict[str, str]] = None,
) -> list[dict]:
    """
    读取 Excel 文件，根据表头名称自动映射列，返回字典列表。

    Args:
        filepath: Excel 文件路径
        sheet_name: 工作表名称，为 None 时取第一个 sheet
        header_row: 表头所在行号（从 1 开始）
        data_start_row: 数据起始行号（从 1 开始）
        column_alias: 列名别名映射，将表头名替换为语义正确名称
                       例如 {"事件ID": "等级"} 将"事件ID"列重命名为"等级"

    Returns:
        每行为一个字典，key 为表头名称（或别名），value 为单元格值
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    alias = column_alias or {}

    # 读取表头
    headers = {}
    for cell in ws[header_row]:
        if cell.value is not None:
            raw_name = str(cell.value).strip()
            mapped_name = alias.get(raw_name, raw_name)
            headers[cell.column] = mapped_name

    # 读取数据行
    rows = []
    for row in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row):
        row_data = {}
        for cell in row:
            col_num = cell.column
            if col_num in headers and cell.value is not None:
                row_data[headers[col_num]] = cell.value
        if row_data:
            rows.append(row_data)

    wb.close()
    return rows


def read_excel_by_columns(
    filepath: str,
    column_map: dict[str, str],
    sheet_name: Optional[str] = None,
    header_row: int = 1,
    data_start_row: int = 2,
) -> list[dict]:
    """
    按列字母映射读取 Excel 文件。

    Args:
        filepath: Excel 文件路径
        column_map: 列字母到字段名的映射，如 {"A": "ip", "B": "level"}
        sheet_name: 工作表名称
        header_row: 表头所在行号
        data_start_row: 数据起始行号

    Returns:
        每行为一个字典，key 为 column_map 中的字段名
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    col_letter_to_num = {}
    for letter, field_name in column_map.items():
        col_num = openpyxl.utils.column_index_from_string(letter)
        col_letter_to_num[col_num] = field_name

    rows = []
    for row_idx in range(data_start_row, ws.max_row + 1):
        row_data = {}
        has_data = False
        for col_num, field_name in col_letter_to_num.items():
            val = ws.cell(row=row_idx, column=col_num).value
            if val is not None:
                has_data = True
                row_data[field_name] = val
        if has_data:
            rows.append(row_data)

    wb.close()
    return rows


def read_json(
    filepath: str,
    data_path: Optional[str] = None,
) -> list[dict]:
    """
    读取 JSON 文件，返回字典列表。

    Args:
        filepath: JSON 文件路径
        data_path: 数据在 JSON 中的路径，用点号分隔。
                   例如 "data.list" 表示 json_obj["data"]["list"]。
                   为 None 时直接使用整个 JSON 作为列表。

    Returns:
        每个元素为一个字典，key 为原始字段名，value 为对应值
    """
    with open(filepath, "r", encoding="utf-8") as f:
        raw = json.load(f)

    # 按 data_path 提取列表
    if data_path:
        obj = raw
        for key in data_path.split("."):
            if isinstance(obj, dict) and key in obj:
                obj = obj[key]
            else:
                print(f"[WARNING] JSON 路径 '{data_path}' 在 '{key}' 处断开")
                return []
        items = obj if isinstance(obj, list) else [obj]
    else:
        items = raw if isinstance(raw, list) else [raw]

    return [item for item in items if isinstance(item, dict)]


def filter_rows(
    data: list[dict],
    field: str,
    value,
    match_mode: str = "exact",
) -> list[dict]:
    """
    按字段筛选数据行。

    Args:
        data: 数据行列表
        field: 筛选字段名
        value: 筛选值
        match_mode: 匹配模式
            - "exact": 精确匹配
            - "contains": 包含匹配
            - "in": value 为列表，字段值在列表中
    Returns:
        筛选后的数据行列表
    """
    result = []
    for row in data:
        if field not in row:
            continue
        cell_val = str(row[field]).strip()
        if match_mode == "exact" and cell_val == str(value):
            result.append(row)
        elif match_mode == "contains" and str(value) in cell_val:
            result.append(row)
        elif match_mode == "in" and cell_val in [str(v) for v in value]:
            result.append(row)
    return result


def count_rows(
    data: list[dict],
    field: str,
    value,
    match_mode: str = "exact",
) -> int:
    """
    统计满足条件的数据行数。
    """
    return len(filter_rows(data, field, value, match_mode))


def count_by_field(data: list[dict], field: str) -> dict[str, int]:
    """
    按某字段分组计数。

    Returns:
        字段值 -> 计数的字典
    """
    result = {}
    for row in data:
        if field in row:
            val = str(row[field]).strip()
            result[val] = result.get(val, 0) + 1
    return result


def resolve_data_file(base_path: str, file_cfg: dict) -> str:
    """
    解析数据源文件路径。

    支持两种方式：
      1. filename: 固定文件名/路径
      2. latest_match: 在 base_path 下按正则匹配最新文件
    """
    latest_match = file_cfg.get("latest_match")
    if latest_match:
        pattern = re.compile(latest_match, re.IGNORECASE)
        if not os.path.isdir(base_path):
            return ""

        candidates = []
        for entry in os.scandir(base_path):
            if entry.is_file() and pattern.search(entry.name):
                candidates.append((entry.path, entry.stat().st_mtime))

        if not candidates:
            return ""

        candidates.sort(key=lambda item: item[1], reverse=True)
        return candidates[0][0]

    filename = file_cfg.get("filename", "")
    if not filename:
        return ""
    if os.path.isabs(filename):
        return filename
    return os.path.join(base_path, filename)


def load_data_sources(config: dict) -> dict[str, list[dict]]:
    """
    根据配置文件加载所有数据源，返回 {数据源名称: 数据行列表}。

    支持 Excel 和 JSON 两种数据源类型：
      - 默认为 Excel（无需指定 type）
      - JSON 类型需在配置中指定 type: json

    Args:
        config: 解析后的配置字典

    Returns:
        各数据源名称到其数据行列表的映射
    """
    ds_cfg = config["data_source"]
    base_path = ds_cfg["base_path"]
    result = {}

    for name, file_cfg in ds_cfg["files"].items():
        source_type = file_cfg.get("type", "excel")
        filepath = resolve_data_file(base_path, file_cfg)

        if not os.path.exists(filepath):
            print(f"[WARNING] 文件不存在: {filepath}")
            result[name] = []
            continue

        if source_type == "json":
            data_path = file_cfg.get("data_path")
            try:
                data = read_json(filepath, data_path=data_path)
                result[name] = data
                print(f"[INFO] 已加载 {name} (JSON): {filepath} ({len(data)} 行)")
            except Exception as e:
                print(f"[ERROR] 加载 {name} (JSON) 失败: {e}")
                result[name] = []
        else:
            sheet = file_cfg.get("sheet")
            header_row = file_cfg.get("header_row", 1)
            data_start_row = file_cfg.get("data_start_row", header_row + 1)
            column_alias = file_cfg.get("column_alias")
            try:
                data = read_excel(
                    filepath,
                    sheet_name=sheet,
                    header_row=header_row,
                    data_start_row=data_start_row,
                    column_alias=column_alias,
                )
                result[name] = data
                print(f"[INFO] 已加载 {name}: {filepath} ({len(data)} 行)")
            except Exception as e:
                print(f"[ERROR] 加载 {name} 失败: {e}")
                result[name] = []

    return result
